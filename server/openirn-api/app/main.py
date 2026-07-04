from __future__ import annotations

import asyncio
import hashlib
import hmac
import json
import os
import re
import secrets
import shutil
import subprocess
import tempfile
import uuid
import urllib.error
import urllib.parse
import urllib.request
from contextlib import contextmanager
from datetime import datetime, timedelta, timezone
from pathlib import Path
from io import BytesIO
from typing import Any, Iterator

from fastapi import FastAPI, HTTPException, Query, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse

APP_VERSION = "0.10.0"
DATA_DIR = Path(os.environ.get("OPENIRN_API_DATA_DIR", "/var/lib/openirn-api"))
MYSQL_URL = os.environ.get("OPENIRN_API_MYSQL_URL", "").strip()
BACKUP_DIR = Path(os.environ.get("OPENIRN_API_BACKUP_DIR", str(DATA_DIR / "backups")))
BACKUP_KEEP = int(os.environ.get("OPENIRN_API_BACKUP_KEEP", "30"))
BACKUP_AUTO_ENABLED = os.environ.get("OPENIRN_API_BACKUP_AUTO_ENABLED", "true").strip().lower() not in {"0", "false", "no", "off"}
BACKUP_PROTECTIVE_ENABLED = os.environ.get("OPENIRN_API_BACKUP_PROTECTIVE_ENABLED", "true").strip().lower() not in {"0", "false", "no", "off"}
BACKUP_PROTECTIVE_MIN_INTERVAL_MINUTES = int(os.environ.get("OPENIRN_API_BACKUP_PROTECTIVE_MIN_INTERVAL_MINUTES", "30"))
BACKUP_SIGNATURE_SECRET = os.environ.get("OPENIRN_API_BACKUP_SIGNATURE_SECRET", "").strip()
MARIADB_SCHEMA_PATH = Path(os.environ.get("OPENIRN_API_MARIADB_SCHEMA", str(Path(__file__).resolve().parents[1] / "sql" / "schema_mariadb.sql")))
MARIADB_SCHEMA_COMPAT_PATH = Path(os.environ.get("OPENIRN_API_MARIADB_SCHEMA_COMPAT", str(Path(__file__).resolve().parents[1] / "sql" / "165a1_widen_sync_events.sql")))
MARIADB_DUMP_BIN = os.environ.get("OPENIRN_MARIADB_DUMP_BIN", "").strip()
TENANT_RE = re.compile(r"[^a-zA-Z0-9_.-]+")
DEVICE_ID_RE = re.compile(r"[^a-zA-Z0-9_.:-]+")
DEFAULT_TENANT_ID = "default"
SOLUTION_ADMIN_TENANT_ID = (
    TENANT_RE.sub("_", os.environ.get("OPENIRN_SOLUTION_ADMIN_TENANT_ID", "archoad").strip())[:80]
    or "archoad"
)
PIN_DEFAULT = os.environ.get("OPENIRN_DEFAULT_USER_PIN", "0000")
PIN_ITERATIONS = int(os.environ.get("OPENIRN_PIN_ITERATIONS", "200000"))
OFFICIAL_ADRI_GITLAB_API = os.environ.get("OPENIRN_ADRI_GITLAB_API", "https://gitlab.com/api/v4").rstrip("/")
OFFICIAL_ADRI_PROJECT_PATH = os.environ.get("OPENIRN_ADRI_PROJECT_PATH", "digitalresilienceinitiative/adri-irn")
OFFICIAL_ADRI_TREE_PATH = os.environ.get("OPENIRN_ADRI_TREE_PATH", "Grille d'évaluation IRN (FR)/xlsx")
OFFICIAL_ADRI_DEFAULT_BRANCH = os.environ.get("OPENIRN_ADRI_DEFAULT_BRANCH", "main")
OFFICIAL_ADRI_SOURCE_URL = os.environ.get("OPENIRN_ADRI_SOURCE_URL", "https://gitlab.com/digitalresilienceinitiative/adri-irn")
OFFICIAL_ADRI_LICENSE = os.environ.get("OPENIRN_ADRI_LICENSE", "CC BY-NC-ND 4.0")
OFFICIAL_REFERENTIAL_DIR = Path(os.environ.get("OPENIRN_REFERENTIAL_DIR", str(DATA_DIR / "referentials")))
AUTH_ATTEMPT_WINDOW_MINUTES = int(os.environ.get("OPENIRN_AUTH_ATTEMPT_WINDOW_MINUTES", "15"))
AUTH_MAX_FAILED_BY_DEVICE = int(os.environ.get("OPENIRN_AUTH_MAX_FAILED_BY_DEVICE", "5"))
AUTH_MAX_FAILED_BY_USER = int(os.environ.get("OPENIRN_AUTH_MAX_FAILED_BY_USER", "5"))
AUTH_MAX_FAILED_BY_IP = int(os.environ.get("OPENIRN_AUTH_MAX_FAILED_BY_IP", "20"))
AUTH_ATTEMPT_RETENTION_DAYS = int(os.environ.get("OPENIRN_AUTH_ATTEMPT_RETENTION_DAYS", "30"))
SESSION_TTL_MINUTES = int(os.environ.get("OPENIRN_SESSION_TTL_MINUTES", "480"))
SESSION_IDLE_TIMEOUT_MINUTES = int(os.environ.get("OPENIRN_SESSION_IDLE_TIMEOUT_MINUTES", "30"))
LEGACY_GLOBAL_BEARER_ENABLED = os.environ.get("OPENIRN_LEGACY_GLOBAL_BEARER_ENABLED", "").strip().lower() in {"1", "true", "yes", "on"}

# Unified API authorization policy.
#
# The device id is deliberately public: it identifies an enrolled terminal but
# does not prove an authenticated user session. Read-only connectivity and
# synchronization checks may use an active terminal id. Every write or
# administration operation must use a short-lived server session with the
# appropriate user role. The historical global bearer is disabled by default
# and never grants write or administration rights.
API_ROLE_ADMIN = {"administrator"}
API_ROLE_CAMPAIGN_MANAGER = {"administrator", "campaign_manager"}
API_ROLE_WRITE = {"administrator", "campaign_manager", "evaluator", "reviewer"}
API_ROLE_READ = {"administrator", "campaign_manager", "evaluator", "reviewer", "reader"}


def _db_backend() -> str:
    return "mariadb"


def _using_mariadb() -> bool:
    return True


def _load_pymysql() -> Any:
    try:
        import pymysql  # type: ignore
    except ImportError as exc:
        raise RuntimeError(
            "Backend MariaDB demandé mais PyMySQL n'est pas installé. "
            "Installez server/openirn-api/requirements-mariadb.txt dans le venv."
        ) from exc
    return pymysql


def _parse_mysql_url() -> dict[str, Any]:
    if not MYSQL_URL:
        raise RuntimeError("OPENIRN_API_MYSQL_URL est requis pour démarrer OpenIRN")
    parsed = urllib.parse.urlparse(MYSQL_URL)
    if parsed.scheme not in {"mysql", "mysql+pymysql", "mariadb", "mariadb+pymysql"}:
        raise RuntimeError("OPENIRN_API_MYSQL_URL doit utiliser mysql+pymysql:// ou mariadb+pymysql://")
    database = parsed.path.lstrip("/")
    if not database:
        raise RuntimeError("OPENIRN_API_MYSQL_URL ne contient pas de nom de base")
    query = urllib.parse.parse_qs(parsed.query)
    return {
        "host": parsed.hostname or "127.0.0.1",
        "port": int(parsed.port or 3306),
        "user": urllib.parse.unquote(parsed.username or ""),
        "password": urllib.parse.unquote(parsed.password or ""),
        "database": database,
        "charset": (query.get("charset") or ["utf8mb4"])[0] or "utf8mb4",
    }


def _mariadb_target_label() -> str:
    try:
        config = _parse_mysql_url()
        return f"{config['user']}@{config['host']}:{config['port']}/{config['database']}"
    except Exception:
        return "mariadb:configuration-incomplete"


class _DbRow:
    def __init__(self, values: tuple[Any, ...], columns: list[str]):
        self._values = tuple(values)
        self._columns = list(columns)
        self._mapping = {column: self._values[index] for index, column in enumerate(self._columns)}

    def __getitem__(self, key: int | str) -> Any:
        if isinstance(key, str):
            return self._mapping[key]
        return self._values[key]

    def __iter__(self):
        return iter(self._values)

    def __len__(self) -> int:
        return len(self._values)

    def keys(self) -> list[str]:
        return list(self._columns)

    def get(self, key: str, default: Any = None) -> Any:
        return self._mapping.get(key, default)

    def __repr__(self) -> str:
        return repr(self._mapping)


class _MySQLCursorResult:
    def __init__(self, cursor: Any):
        self._cursor = cursor
        self._columns = [item[0] for item in (cursor.description or [])]

    def _wrap(self, row: Any) -> _DbRow | None:
        if row is None:
            return None
        if isinstance(row, dict):
            columns = list(row.keys())
            return _DbRow(tuple(row[column] for column in columns), columns)
        return _DbRow(tuple(row), self._columns)

    def fetchone(self) -> _DbRow | None:
        return self._wrap(self._cursor.fetchone())

    def fetchall(self) -> list[_DbRow]:
        return [row for raw in self._cursor.fetchall() if (row := self._wrap(raw)) is not None]

    def __iter__(self):
        return iter(self.fetchall())


class DbError(RuntimeError):
    pass


class _MySQLConnection:
    def __init__(self):
        pymysql = _load_pymysql()
        config = _parse_mysql_url()
        try:
            self._con = pymysql.connect(
                host=config["host"],
                port=int(config["port"]),
                user=config["user"],
                password=config["password"],
                database=config["database"],
                charset=config.get("charset") or "utf8mb4",
                autocommit=False,
            )
        except Exception as exc:
            raise DbError(f"Connexion MariaDB impossible: {exc}") from exc

    def execute(self, sql: str, parameters: Any = None) -> _MySQLCursorResult:
        translated = _translate_mysql_sql(sql)
        cursor = self._con.cursor()
        try:
            cursor.execute(translated, parameters or None)
        except Exception as exc:
            try:
                self._con.rollback()
            except Exception:
                pass
            raise DbError(f"Erreur MariaDB: {exc}; SQL={translated}") from exc
        return _MySQLCursorResult(cursor)

    def executescript(self, script: str) -> None:
        for statement in _split_sql_script(_strip_sql_comments(script)):
            if statement.strip():
                self.execute(statement)

    def commit(self) -> None:
        self._con.commit()

    def rollback(self) -> None:
        self._con.rollback()

    def close(self) -> None:
        self._con.close()

    def __enter__(self) -> "_MySQLConnection":
        return self

    def __exit__(self, exc_type: Any, exc: Any, tb: Any) -> None:
        if exc_type is None:
            self.commit()
        else:
            self.rollback()


def _strip_sql_comments(script: str) -> str:
    lines: list[str] = []
    for line in script.splitlines():
        if line.strip().startswith("--"):
            continue
        lines.append(line)
    return "\n".join(lines)


def _split_sql_script(script: str) -> list[str]:
    statements: list[str] = []
    current: list[str] = []
    in_single = False
    in_double = False
    escape = False
    for char in script:
        current.append(char)
        if escape:
            escape = False
            continue
        if char == "\\":
            escape = True
            continue
        if char == "'" and not in_double:
            in_single = not in_single
            continue
        if char == '"' and not in_single:
            in_double = not in_double
            continue
        if char == ";" and not in_single and not in_double:
            statement = "".join(current).strip()
            current = []
            if statement:
                statements.append(statement[:-1].strip())
    tail = "".join(current).strip()
    if tail:
        statements.append(tail)
    return statements


def _translate_mysql_sql(sql: str) -> str:
    translated = sql.strip()
    low = translated.lower()
    if low.startswith("pragma foreign_keys"):
        return "SET FOREIGN_KEY_CHECKS = 0" if "off" in low else "SET FOREIGN_KEY_CHECKS = 1"
    if low.startswith("pragma journal_mode"):
        return "SELECT 1"

    translated = re.sub(r"\bINSERT\s+OR\s+IGNORE\s+INTO\b", "INSERT IGNORE INTO", translated, flags=re.IGNORECASE)
    translated = re.sub(r"\bINSERT\s+OR\s+REPLACE\s+INTO\b", "REPLACE INTO", translated, flags=re.IGNORECASE)

    if re.search(r"ON\s+CONFLICT\s*\([^)]*\)\s*DO\s+NOTHING", translated, flags=re.IGNORECASE | re.DOTALL):
        translated = re.sub(r"\bINSERT\s+INTO\b", "INSERT IGNORE INTO", translated, count=1, flags=re.IGNORECASE)
        translated = re.sub(r"\s*ON\s+CONFLICT\s*\([^)]*\)\s*DO\s+NOTHING\s*", "", translated, flags=re.IGNORECASE | re.DOTALL)

    if re.search(r"ON\s+CONFLICT\s*\([^)]*\)\s*DO\s+UPDATE\s+SET", translated, flags=re.IGNORECASE | re.DOTALL):
        translated = re.sub(
            r"\s*ON\s+CONFLICT\s*\([^)]*\)\s*DO\s+UPDATE\s+SET\s*",
            " ON DUPLICATE KEY UPDATE ",
            translated,
            flags=re.IGNORECASE | re.DOTALL,
        )
        translated = re.sub(r"\bexcluded\.([A-Za-z_][A-Za-z0-9_]*)\b", r"VALUES(\1)", translated, flags=re.IGNORECASE)

    return translated.replace("?", "%s")


app = FastAPI(
    title="OpenIRN API",
    version=APP_VERSION,
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "https://www.archoad.io",
    ],
    allow_credentials=True,
    allow_methods=["GET", "POST", "DELETE", "OPTIONS"],
    allow_headers=["Authorization", "Content-Type", "Accept"],
)


def _utc_now() -> datetime:
    return datetime.now(timezone.utc)


def _safe_segment(value: Any, fallback: str) -> str:
    raw = str(value or "").strip()
    if not raw:
        return fallback
    cleaned = TENANT_RE.sub("_", raw)
    return cleaned[:80] or fallback


def _normalize_device_id(value: Any) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    cleaned = DEVICE_ID_RE.sub("_", raw)
    return cleaned[:160].strip("._-:") or ""


UUID_RE = re.compile(
    r"^[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[1-5][0-9a-fA-F]{3}-[89abAB][0-9a-fA-F]{3}-[0-9a-fA-F]{12}$"
)


def _new_uuid() -> str:
    return str(uuid.uuid4())


def _is_uuid(value: Any) -> bool:
    return bool(UUID_RE.match(str(value or "").strip()))


def _normalize_uuid(value: Any) -> str:
    raw = str(value or "").strip().lower()
    return raw if _is_uuid(raw) else ""


def _canonical_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def _pretty_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, indent=2)


def _json_sha256(value: Any) -> str:
    return hashlib.sha256(_canonical_json(value).encode("utf-8")).hexdigest()


def _backup_signing_secret() -> str:
    # Backups use their own dedicated secret. OPENIRN_API_TOKEN is deliberately
    # not reused here: the historical global bearer is deprecated and must not
    # remain a hidden signing dependency.
    return BACKUP_SIGNATURE_SECRET


def _backup_signature_payload(metadata: dict[str, Any]) -> dict[str, Any]:
    return {
        key: value
        for key, value in metadata.items()
        if key not in {"signature", "signatureAlgorithm", "signatureStatus"}
    }


def _backup_metadata_signature(metadata: dict[str, Any], secret: str | None = None) -> str:
    effective_secret = (secret or _backup_signing_secret()).strip()
    if not effective_secret:
        return ""
    message = _canonical_json(_backup_signature_payload(metadata)).encode("utf-8")
    return hmac.new(effective_secret.encode("utf-8"), message, hashlib.sha256).hexdigest()


def _backup_signature_status(metadata: dict[str, Any]) -> str:
    signature = str(metadata.get("signature") or "").strip()
    if not signature:
        return "unsigned"
    secret = _backup_signing_secret()
    if not secret:
        return "unverified_no_secret"
    expected = _backup_metadata_signature(metadata, secret)
    return "valid" if expected and hmac.compare_digest(signature, expected) else "invalid"


def _chmod_private(path: Path, mode: int) -> None:
    try:
        path.chmod(mode)
    except OSError:
        # Best effort: permissions may be constrained by the filesystem.
        pass


def _ensure_private_backup_dir() -> None:
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    _chmod_private(BACKUP_DIR, 0o700)


def _parse_json(raw: str | None, fallback: Any) -> Any:
    if not raw:
        return fallback
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        return fallback


def _parse_datetime(value: Any) -> datetime:
    raw = str(value or "").strip()
    if not raw:
        return datetime.fromtimestamp(0, timezone.utc)
    try:
        normalized = raw.replace("Z", "+00:00")
        parsed = datetime.fromisoformat(normalized)
        if parsed.tzinfo is None:
            return parsed.replace(tzinfo=timezone.utc)
        return parsed.astimezone(timezone.utc)
    except ValueError:
        return datetime.fromtimestamp(0, timezone.utc)


def _raw_configured_api_token() -> str:
    return os.environ.get("OPENIRN_API_TOKEN", "").strip()


def _configured_api_token() -> str:
    if not LEGACY_GLOBAL_BEARER_ENABLED:
        return ""
    return _raw_configured_api_token()


def _extract_bearer_token(request: Request) -> str:
    authorization = request.headers.get("authorization", "").strip()
    if not authorization:
        return ""
    scheme, _, token = authorization.partition(" ")
    if scheme.lower() != "bearer" or not token.strip():
        return ""
    return token.strip()


def _secret_hash(value: str) -> str:
    return hashlib.sha256(str(value or "").encode("utf-8")).hexdigest()


def _enrollment_code_hash_with_pepper(tenant_id: str, code: str, pepper: str) -> str:
    normalized = _normalize_enrollment_code(code)
    return hmac.new(
        pepper.encode("utf-8"),
        f"{tenant_id}:{normalized}".encode("utf-8"),
        hashlib.sha256,
    ).hexdigest()


def _enrollment_code_hash(tenant_id: str, code: str) -> str:
    # New enrollment codes no longer depend on OPENIRN_API_TOKEN. This keeps
    # terminal pairing independent from the deprecated global bearer.
    return _enrollment_code_hash_with_pepper(
        tenant_id,
        code,
        "openirn-device-enrollment-v2",
    )


def _bootstrap_enrollment_code_hash(tenant_id: str, code: str) -> str:
    return _enrollment_code_hash_with_pepper(
        tenant_id,
        code,
        "openirn-device-enrollment-bootstrap-v1",
    )


def _enrollment_code_hash_candidates(tenant_id: str, code: str) -> list[str]:
    hashes = [
        _enrollment_code_hash(tenant_id, code),
        _bootstrap_enrollment_code_hash(tenant_id, code),
        _enrollment_code_hash_with_pepper(tenant_id, code, "openirn-device-enrollment"),
    ]
    legacy_api_token = _raw_configured_api_token()
    if legacy_api_token:
        # Compatibility only for already-issued codes created before the
        # bearer-global deprecation. This does not re-enable API bearer access.
        hashes.append(_enrollment_code_hash_with_pepper(tenant_id, code, legacy_api_token))
    return list(dict.fromkeys(hashes))


def _normalize_enrollment_code(value: Any) -> str:
    return re.sub(r"[^A-Z0-9]", "", str(value or "").upper())


def _format_enrollment_code(value: str) -> str:
    normalized = _normalize_enrollment_code(value)
    return "-".join(normalized[index : index + 4] for index in range(0, len(normalized), 4))


def _new_enrollment_code() -> str:
    # Crockford-like alphabet without easily confused characters. 10 chars ≈ 50 bits.
    alphabet = "ABCDEFGHJKMNPQRSTUVWXYZ23456789"
    return "".join(secrets.choice(alphabet) for _ in range(10))


def _new_device_token() -> str:
    return "odt_" + secrets.token_urlsafe(36)


def _new_session_token() -> str:
    return "ost_" + secrets.token_urlsafe(36)


def _session_auth_context(provided_token: str) -> dict[str, Any] | None:
    token_hash = _secret_hash(provided_token)
    now = _utc_now()
    try:
        with _db() as con:
            row = con.execute(
                """
                SELECT s.tenant_id, s.session_id, s.device_id, s.user_id,
                       s.expires_at, s.last_seen_at, s.revoked_at,
                       u.role AS user_role, u.active AS user_active
                FROM api_sessions s
                LEFT JOIN users u
                  ON u.tenant_id = s.tenant_id AND u.user_id = s.user_id
                WHERE s.token_hash = ? AND s.revoked_at IS NULL
                """,
                (token_hash,),
            ).fetchone()
            if row is None:
                return None
            if _parse_datetime(row["expires_at"]) < now:
                return None
            last_seen_at = _parse_datetime(row["last_seen_at"])
            idle_timeout = max(1, SESSION_IDLE_TIMEOUT_MINUTES)
            if now - last_seen_at >= timedelta(minutes=idle_timeout):
                con.execute(
                    """
                    UPDATE api_sessions
                    SET revoked_at = ?
                    WHERE tenant_id = ? AND session_id = ?
                    """,
                    (now.isoformat(), row["tenant_id"], row["session_id"]),
                )
                _record_device_audit(
                    con,
                    row["tenant_id"],
                    "session.idle_timeout",
                    device_id=row["device_id"],
                    payload={"sessionId": row["session_id"], "idleTimeoutMinutes": idle_timeout},
                )
                con.commit()
                return None
            device = con.execute(
                """
                SELECT 1
                FROM authorized_devices
                WHERE tenant_id = ? AND device_id = ?
                  AND status = 'active' AND revoked_at IS NULL
                """,
                (row["tenant_id"], row["device_id"]),
            ).fetchone()
            if device is None:
                return None
            if row["user_active"] is not None and int(row["user_active"] or 0) != 1:
                return None

            con.execute(
                """
                UPDATE api_sessions
                SET last_seen_at = ?
                WHERE tenant_id = ? AND session_id = ?
                """,
                (now.isoformat(), row["tenant_id"], row["session_id"]),
            )
            con.execute(
                """
                UPDATE authorized_devices
                SET last_seen_at = ?
                WHERE tenant_id = ? AND device_id = ?
                """,
                (now.isoformat(), row["tenant_id"], row["device_id"]),
            )
            _touch_terminal(con, str(row["device_id"] or ""), now.isoformat())
            con.commit()
            return {
                "authMode": "session",
                "tenantId": row["tenant_id"],
                "sessionId": row["session_id"],
                "deviceId": row["device_id"],
                "userId": row["user_id"],
                "userRole": _role_normalize(row["user_role"]),
            }
    except DbError:
        return None


def _is_active_session_token(provided_token: str) -> bool:
    return _session_auth_context(provided_token) is not None


def _device_token_auth_context(provided_token: str) -> dict[str, Any] | None:
    token_hash = _secret_hash(provided_token)
    now = _utc_now().isoformat()
    try:
        with _db() as con:
            row = con.execute(
                """
                SELECT tenant_id, device_id
                FROM authorized_devices
                WHERE token_hash = ? AND status = 'active' AND revoked_at IS NULL
                """,
                (token_hash,),
            ).fetchone()
            if row is None:
                return None
            con.execute(
                """
                UPDATE authorized_devices
                SET last_seen_at = ?
                WHERE tenant_id = ? AND device_id = ?
                """,
                (now, row["tenant_id"], row["device_id"]),
            )
            _touch_terminal(con, str(row["device_id"] or ""), now)
            con.commit()
            return {
                "authMode": "legacy_device_token",
                "tenantId": row["tenant_id"],
                "deviceId": row["device_id"],
                "userId": "",
                "userRole": "",
            }
    except DbError:
        return None


def _request_auth_context(request: Request) -> dict[str, Any] | None:
    provided_token = _extract_bearer_token(request)
    if not provided_token:
        return None

    session_context = _session_auth_context(provided_token)
    if session_context is not None:
        return session_context

    device_context = _device_token_auth_context(provided_token)
    if device_context is not None:
        return device_context

    expected_token = _configured_api_token()
    if expected_token and hmac.compare_digest(provided_token, expected_token):
        return {
            "authMode": "legacy_global_bearer",
            "tenantId": "",
            "deviceId": "",
            "userId": "",
            "userRole": "",
        }

    return None


def _is_administrator_context(context: dict[str, Any] | None) -> bool:
    if context is None:
        return False
    return (
        str(context.get("authMode") or "") == "session"
        and _role_normalize(context.get("userRole")) == "administrator"
    )


def _is_solution_admin_context(context: dict[str, Any] | None) -> bool:
    # Depuis la v1, le rôle Administrateur représente l’administration globale
    # OpenIRN, quel que soit l’espace de travail utilisé pour ouvrir la session.
    # Le nom historique "solution admin" reste conservé pour compatibilité avec
    # les réponses API et l’interface existante.
    return _is_administrator_context(context)


def _request_has_solution_admin_authorization(request: Request) -> bool:
    return _is_solution_admin_context(_request_auth_context(request))


def _request_has_api_authorization(request: Request) -> bool:
    return _request_auth_context(request) is not None


def _authorization_unavailable_exception() -> HTTPException:
    return HTTPException(status_code=403, detail="Session expirée ou autorisation OpenIRN invalide")


def _require_role_authorization(
    request: Request,
    tenant_id: str,
    allowed_roles: set[str],
    *,
    detail: str,
) -> dict[str, Any]:
    context = _request_auth_context(request)
    if context is None:
        raise _authorization_unavailable_exception()

    auth_mode = str(context.get("authMode") or "")
    if auth_mode == "legacy_device_token":
        raise HTTPException(
            status_code=403,
            detail="Ce terminal n’est pas autorisé pour modifier les données ou administrer OpenIRN",
        )
    if auth_mode == "legacy_global_bearer":
        raise HTTPException(
            status_code=403,
            detail="Le bearer global legacy ne donne plus de droits d’écriture ou d’administration",
        )

    role = _role_normalize(context.get("userRole"))
    if role not in allowed_roles:
        raise HTTPException(status_code=403, detail=detail)

    if tenant_id and str(context.get("tenantId") or "") != tenant_id:
        if _is_administrator_context(context):
            return context
        raise HTTPException(status_code=403, detail="La session ne correspond pas à l’espace de travail demandé")

    return context


def _require_admin_authorization(request: Request, tenant_id: str) -> dict[str, Any]:
    return _require_role_authorization(
        request,
        tenant_id,
        API_ROLE_ADMIN,
        detail="Cette action API est réservée aux administrateurs OpenIRN",
    )


def _require_campaign_manager_authorization(request: Request, tenant_id: str) -> dict[str, Any]:
    return _require_role_authorization(
        request,
        tenant_id,
        API_ROLE_CAMPAIGN_MANAGER,
        detail="Cette action API est réservée aux administrateurs et pilotes IRN",
    )


def _require_write_authorization(request: Request, tenant_id: str) -> dict[str, Any]:
    return _require_role_authorization(
        request,
        tenant_id,
        API_ROLE_WRITE,
        detail="Cette action API exige une session utilisateur active avec droits d’écriture",
    )


def _require_device_or_authorized_read(request: Request, tenant_id: str) -> dict[str, Any] | None:
    context = _request_auth_context(request)
    if context is not None:
        auth_mode = str(context.get("authMode") or "")
        if auth_mode == "legacy_global_bearer":
            return context
        if not tenant_id or str(context.get("tenantId") or "") == tenant_id:
            return context
        if _is_administrator_context(context):
            return context
        raise HTTPException(status_code=403, detail="L’autorisation ne correspond pas à l’espace de travail demandé")
    _require_active_device(request, tenant_id)
    return None


def _request_device_id(request: Request, payload: dict[str, Any] | None = None) -> str:
    header_value = request.headers.get("x-openirn-device-id", "").strip()
    if header_value:
        return header_value[:160]
    if payload:
        body_value = str(payload.get("deviceId") or "").strip()
        if body_value:
            return body_value[:160]
    return ""


def _request_client_ip(request: Request) -> str:
    forwarded_for = request.headers.get("x-forwarded-for", "").strip()
    if forwarded_for:
        return forwarded_for.split(",", 1)[0].strip()[:80]
    real_ip = request.headers.get("x-real-ip", "").strip()
    if real_ip:
        return real_ip[:80]
    if request.client and request.client.host:
        return request.client.host[:80]
    return "unknown"


def _record_auth_attempt(
    con: Any,
    tenant_id: str,
    *,
    device_id: str,
    user_id: str,
    ip_address: str,
    successful: bool,
    reason: str,
) -> None:
    now = _utc_now()
    retention_start = now - timedelta(days=max(1, AUTH_ATTEMPT_RETENTION_DAYS))
    con.execute(
        """
        DELETE FROM auth_attempts
        WHERE tenant_id = ? AND created_at < ?
        """,
        (tenant_id, retention_start.isoformat()),
    )
    con.execute(
        """
        INSERT INTO auth_attempts(
            tenant_id, attempt_id, device_id, user_id, ip_address,
            successful, reason, created_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            tenant_id,
            "auth-" + secrets.token_urlsafe(18),
            device_id[:160],
            user_id[:160],
            ip_address[:80],
            1 if successful else 0,
            reason[:120],
            now.isoformat(),
        ),
    )


def _recent_auth_failures(
    con: Any,
    tenant_id: str,
    column: str,
    value: str,
) -> int:
    allowed_columns = {"device_id", "user_id", "ip_address"}
    if column not in allowed_columns or not value:
        return 0
    window_start = (_utc_now() - timedelta(minutes=max(1, AUTH_ATTEMPT_WINDOW_MINUTES))).isoformat()
    row = con.execute(
        f"""
        SELECT COUNT(*) AS total
        FROM auth_attempts
        WHERE tenant_id = ?
          AND successful = 0
          AND created_at >= ?
          AND {column} = ?
        """,
        (tenant_id, window_start, value),
    ).fetchone()
    return int(row["total"] if row is not None else 0)


def _enforce_auth_rate_limit(
    con: Any,
    tenant_id: str,
    *,
    device_id: str,
    user_id: str,
    ip_address: str,
) -> None:
    checks = [
        ("device_id", device_id, AUTH_MAX_FAILED_BY_DEVICE, "Trop de codes invalides pour ce terminal"),
        ("user_id", user_id, AUTH_MAX_FAILED_BY_USER, "Trop de codes invalides pour ce profil"),
        ("ip_address", ip_address, AUTH_MAX_FAILED_BY_IP, "Trop de tentatives depuis cette adresse réseau"),
    ]
    for column, value, limit, message in checks:
        if limit <= 0:
            continue
        if _recent_auth_failures(con, tenant_id, column, value) >= limit:
            _record_auth_attempt(
                con,
                tenant_id,
                device_id=device_id,
                user_id=user_id,
                ip_address=ip_address,
                successful=False,
                reason=f"rate_limited:{column}",
            )
            _record_device_audit(
                con,
                tenant_id,
                "auth.rate_limited",
                device_id=device_id,
                payload={
                    "userId": user_id,
                    "ipAddress": ip_address,
                    "scope": column,
                    "windowMinutes": AUTH_ATTEMPT_WINDOW_MINUTES,
                    "limit": limit,
                },
            )
            con.commit()
            raise HTTPException(
                status_code=429,
                detail=f"{message}. Réessayez dans quelques minutes.",
            )


def _require_active_device(
    request: Request,
    tenant_id: str,
    payload: dict[str, Any] | None = None,
) -> str:
    device_id = _request_device_id(request, payload)
    if not device_id:
        raise HTTPException(status_code=401, detail="Missing OpenIRN device id")
    now = _utc_now().isoformat()
    with _db() as con:
        row = con.execute(
            """
            SELECT 1
            FROM authorized_devices
            WHERE tenant_id = ? AND device_id = ?
              AND status = 'active' AND revoked_at IS NULL
            """,
            (tenant_id, device_id),
        ).fetchone()
        if row is None:
            raise HTTPException(status_code=403, detail="Terminal non autorisé ou révoqué")
        con.execute(
            """
            UPDATE authorized_devices
            SET last_seen_at = ?
            WHERE tenant_id = ? AND device_id = ?
            """,
            (now, tenant_id, device_id),
        )
        con.commit()
    return device_id


def _create_api_session(
    con: Any,
    tenant_id: str,
    device_id: str,
    user_id: str,
    ttl_minutes: int | None = None,
) -> tuple[str, str, datetime]:
    token = _new_session_token()
    session_id = "session-" + secrets.token_urlsafe(18)
    now = _utc_now()
    effective_ttl_minutes = max(5, ttl_minutes or SESSION_TTL_MINUTES)
    expires_at = now + timedelta(minutes=effective_ttl_minutes)
    con.execute(
        """
        INSERT INTO api_sessions(
            tenant_id, session_id, token_hash, device_id, user_id,
            created_at, expires_at, last_seen_at, revoked_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, NULL)
        """,
        (
            tenant_id,
            session_id,
            _secret_hash(token),
            device_id,
            user_id,
            now.isoformat(),
            expires_at.isoformat(),
            now.isoformat(),
        ),
    )
    return session_id, token, expires_at


def _is_active_device_token(provided_token: str) -> bool:
    return _device_token_auth_context(provided_token) is not None


def _require_api_token(request: Request) -> None:
    # Kept for compatibility with transition endpoints.  New endpoint code
    # should prefer the explicit helpers below:
    #   _require_admin_authorization
    #   _require_campaign_manager_authorization
    #   _require_write_authorization
    #   _require_sync_read_access
    if _request_has_api_authorization(request):
        return
    raise _authorization_unavailable_exception()


def _require_sync_read_access(request: Request, tenant_id: str) -> None:
    """Authorize read-only synchronization endpoints.

    Read-only endpoints accept a session token, a legacy device token, an
    enrolled active terminal id in X-OpenIRN-Device-Id, or, only when explicitly
    re-enabled server-side, the deprecated global bearer. A terminal id is not a
    secret and never grants write or administration rights.
    """
    _require_device_or_authorized_read(request, tenant_id)


def _pin_hash(pin: str, salt: str, iterations: int = PIN_ITERATIONS) -> str:
    return hashlib.pbkdf2_hmac(
        "sha256",
        str(pin or "").encode("utf-8"),
        salt.encode("utf-8"),
        iterations,
    ).hex()


@contextmanager
def _db() -> Iterator[Any]:
    con = _MySQLConnection()
    try:
        yield con
    finally:
        con.close()


def _apply_schema() -> None:
    if not MARIADB_SCHEMA_PATH.exists():
        raise RuntimeError(f"OpenIRN MariaDB schema not found: {MARIADB_SCHEMA_PATH}")
    with _db() as con:
        con.executescript(MARIADB_SCHEMA_PATH.read_text(encoding="utf-8"))
        if MARIADB_SCHEMA_COMPAT_PATH.exists():
            con.executescript(MARIADB_SCHEMA_COMPAT_PATH.read_text(encoding="utf-8"))
        _migrate_tenants_schema(con)
        _migrate_existing_entity_ids_to_uuid(con)
        _delete_legacy_revoked_authorized_devices(con)
        _migrate_authorized_device_identity_schema(con)
        _migrate_terminal_identity_schema(con)
        _ensure_tenant(con, DEFAULT_TENANT_ID)
        _backfill_default_tenant_display_name(con)
        _sync_solution_administrators_to_all_tenants(con)
        _backfill_official_referential_history(con)
        con.commit()


@app.on_event("startup")
def _startup() -> None:
    _apply_schema()



def _table_exists(con: Any, table_name: str) -> bool:
    row = con.execute(
        """
        SELECT table_name
        FROM information_schema.tables
        WHERE table_schema = DATABASE() AND table_name = ?
        """,
        (table_name,),
    ).fetchone()
    return row is not None


def _table_columns(con: Any, table_name: str) -> set[str]:
    rows = con.execute(
        """
        SELECT column_name
        FROM information_schema.columns
        WHERE table_schema = DATABASE() AND table_name = ?
        ORDER BY ordinal_position
        """,
        (table_name,),
    ).fetchall()
    return {str(row[0]) for row in rows}


def _migrate_tenants_schema(con: Any) -> None:
    columns = _table_columns(con, "tenants")
    if "display_name" not in columns:
        con.execute("ALTER TABLE tenants ADD COLUMN display_name VARCHAR(255) NOT NULL DEFAULT ''" if isinstance(con, _MySQLConnection) else "ALTER TABLE tenants ADD COLUMN display_name TEXT NOT NULL DEFAULT ''")
    if "description" not in columns:
        con.execute("ALTER TABLE tenants ADD COLUMN description TEXT NOT NULL" if isinstance(con, _MySQLConnection) else "ALTER TABLE tenants ADD COLUMN description TEXT NOT NULL DEFAULT ''")
    if "permanent" not in columns:
        con.execute("ALTER TABLE tenants ADD COLUMN permanent TINYINT(1) NOT NULL DEFAULT 0" if isinstance(con, _MySQLConnection) else "ALTER TABLE tenants ADD COLUMN permanent INTEGER NOT NULL DEFAULT 0")


def _migration_applied(con: Any, version: int) -> bool:
    row = con.execute(
        "SELECT 1 FROM schema_migrations WHERE version = ?",
        (version,),
    ).fetchone()
    return row is not None


def _record_migration(con: Any, version: int, name: str) -> None:
    con.execute(
        "INSERT OR IGNORE INTO schema_migrations(version, name) VALUES (?, ?)",
        (version, name),
    )


def _alias_target(con: Any, entity_type: str, old_id: str, scope_id: str = "") -> str:
    row = con.execute(
        """
        SELECT new_id FROM id_aliases
        WHERE entity_type = ? AND scope_id = ? AND old_id = ?
        """,
        (entity_type, scope_id, old_id),
    ).fetchone()
    return str(row["new_id"] or "") if row else ""


def _store_alias(
    con: Any,
    *,
    entity_type: str,
    old_id: str,
    new_id: str,
    scope_id: str = "",
) -> None:
    if not old_id or not new_id or old_id == new_id:
        return
    con.execute(
        """
        INSERT OR IGNORE INTO id_aliases(entity_type, scope_id, old_id, new_id, created_at)
        VALUES (?, ?, ?, ?, ?)
        """,
        (entity_type, scope_id, old_id, new_id, _utc_now().isoformat()),
    )


def _rewrite_json_ids(raw: Any, mappings: dict[str, str]) -> Any:
    if not mappings:
        return raw
    if isinstance(raw, str):
        return mappings.get(raw, raw)
    if isinstance(raw, list):
        return [_rewrite_json_ids(item, mappings) for item in raw]
    if isinstance(raw, dict):
        rewritten: dict[str, Any] = {}
        for key, value in raw.items():
            new_key = mappings.get(str(key), str(key))
            rewritten[new_key] = _rewrite_json_ids(value, mappings)
        return rewritten
    return raw


def _rewrite_json_text(raw_text: str, mappings: dict[str, str]) -> str:
    if not raw_text or not mappings:
        return raw_text
    try:
        payload = json.loads(raw_text)
    except Exception:
        return raw_text
    return _canonical_json(_rewrite_json_ids(payload, mappings))


def _rewrite_json_column(
    con: Any,
    table: str,
    pk_columns: list[str],
    json_column: str,
    mappings: dict[str, str],
) -> None:
    if not mappings:
        return
    columns = _table_columns(con, table)
    if json_column not in columns:
        return
    selected_columns = pk_columns + [json_column]
    rows = con.execute(
        f"SELECT {', '.join(selected_columns)} FROM {table}"
    ).fetchall()
    for row in rows:
        original = str(row[json_column] or "")
        rewritten = _rewrite_json_text(original, mappings)
        if rewritten == original:
            continue
        where_clause = " AND ".join(f"{column} = ?" for column in pk_columns)
        con.execute(
            f"UPDATE {table} SET {json_column} = ? WHERE {where_clause}",
            (rewritten, *[row[column] for column in pk_columns]),
        )


def _resolve_tenant_id(con: Any, tenant_id: Any, fallback: str = DEFAULT_TENANT_ID) -> str:
    requested = _safe_segment(tenant_id, fallback)
    row = con.execute("SELECT id FROM tenants WHERE id = ?", (requested,)).fetchone()
    if row is not None:
        return str(row["id"] or requested)
    aliased = _alias_target(con, "tenant", requested)
    if aliased:
        return aliased
    if requested != fallback:
        fallback_alias = _alias_target(con, "tenant", fallback)
        if fallback_alias and requested == fallback:
            return fallback_alias
    return requested


def _resolve_tenant_id_for_request(value: Any, fallback: str = DEFAULT_TENANT_ID) -> str:
    requested = _safe_segment(value, fallback)
    try:
        with _db() as con:
            return _resolve_tenant_id(con, requested, fallback)
    except Exception:
        return requested


def _default_tenant_id(con: Any) -> str:
    row = con.execute(
        "SELECT id FROM tenants WHERE permanent = 1 ORDER BY created_at ASC LIMIT 1"
    ).fetchone()
    if row is not None:
        return str(row["id"] or DEFAULT_TENANT_ID)
    aliased = _alias_target(con, "tenant", DEFAULT_TENANT_ID)
    return aliased or DEFAULT_TENANT_ID


def _tenant_id_for_creation(value: Any = None) -> str:
    requested = _normalize_uuid(value)
    return requested or _new_uuid()


def _user_id_for_save(con: Any, tenant_id: str, user_id: Any) -> str:
    requested = str(user_id or "").strip()
    if not requested:
        return _new_uuid()
    normalized = _normalize_uuid(requested)
    if normalized:
        return normalized
    aliased = _alias_target(con, "user", requested, tenant_id)
    if aliased:
        return aliased
    created = _new_uuid()
    _store_alias(con, entity_type="user", scope_id=tenant_id, old_id=requested, new_id=created)
    return created


def _campaign_id_for_save(con: Any, tenant_id: str, campaign_id: Any) -> str:
    requested = str(campaign_id or "").strip()
    if not requested:
        return _new_uuid()
    normalized = _normalize_uuid(requested)
    if normalized:
        return normalized
    aliased = _alias_target(con, "campaign", requested, tenant_id)
    if aliased:
        return aliased
    created = _new_uuid()
    _store_alias(con, entity_type="campaign", scope_id=tenant_id, old_id=requested, new_id=created)
    return created


def _migrate_existing_entity_ids_to_uuid(con: Any) -> None:
    if _migration_applied(con, 156):
        return

    con.execute("SET FOREIGN_KEY_CHECKS = 0")

    tenant_map: dict[str, str] = {}
    tenant_rows = con.execute("SELECT id FROM tenants ORDER BY id ASC").fetchall()
    for row in tenant_rows:
        old_id = str(row["id"] or "").strip()
        if not old_id:
            continue
        new_id = old_id.lower() if _is_uuid(old_id) else (_alias_target(con, "tenant", old_id) or _new_uuid())
        tenant_map[old_id] = new_id
        _store_alias(con, entity_type="tenant", old_id=old_id, new_id=new_id)

    tenant_tables = [
        "users",
        "user_credentials",
        "sync_snapshots",
        "campaign_states",
        "campaign_revisions",
        "authorized_devices",
        "device_enrollment_requests",
        "device_enrollment_codes",
        "api_sessions",
        "auth_attempts",
        "device_audit_log",
        "official_referentials",
        "official_referential_history",
        "sync_events",
        "backup_audit_log",
    ]
    for old_id, new_id in tenant_map.items():
        if old_id == new_id:
            continue
        con.execute("UPDATE tenants SET id = ? WHERE id = ?", (new_id, old_id))
        for table in tenant_tables:
            if _table_exists(con, table):
                columns = _table_columns(con, table)
                if "tenant_id" in columns:
                    con.execute(f"UPDATE {table} SET tenant_id = ? WHERE tenant_id = ?", (new_id, old_id))

    # Keep a permanent default alias even after the default tenant itself becomes a UUID.
    default_id = tenant_map.get(DEFAULT_TENANT_ID) or _alias_target(con, "tenant", DEFAULT_TENANT_ID)
    if default_id:
        con.execute("UPDATE tenants SET permanent = 1 WHERE id = ?", (default_id,))

    user_map: dict[str, str] = {}
    user_rows = con.execute(
        "SELECT tenant_id, user_id FROM users ORDER BY tenant_id ASC, user_id ASC"
    ).fetchall()
    for row in user_rows:
        tenant_id = str(row["tenant_id"] or "")
        old_user_id = str(row["user_id"] or "").strip()
        if not old_user_id:
            continue
        new_user_id = old_user_id.lower() if _is_uuid(old_user_id) else user_map.get(old_user_id) or _new_uuid()
        user_map[old_user_id] = new_user_id
        _store_alias(con, entity_type="user", scope_id=tenant_id, old_id=old_user_id, new_id=new_user_id)

    user_fk_tables = [
        ("users", "user_id"),
        ("user_credentials", "user_id"),
        ("api_sessions", "user_id"),
        ("auth_attempts", "user_id"),
        ("authorized_devices", "invited_by_user_id"),
        ("device_enrollment_requests", "decided_by_user_id"),
        ("device_enrollment_codes", "created_by_user_id"),
        ("official_referential_history", "triggered_by_user_id"),
        ("backup_audit_log", "triggered_by_user_id"),
    ]
    for old_user_id, new_user_id in user_map.items():
        if old_user_id == new_user_id:
            continue
        for table, column in user_fk_tables:
            columns = _table_columns(con, table) if _table_exists(con, table) else set()
            if column in columns:
                con.execute(f"UPDATE {table} SET {column} = ? WHERE {column} = ?", (new_user_id, old_user_id))

    campaign_map: dict[str, str] = {}
    campaign_rows = con.execute(
        """
        SELECT tenant_id, campaign_id FROM campaign_states
        UNION
        SELECT tenant_id, campaign_id FROM campaign_revisions
        ORDER BY tenant_id ASC, campaign_id ASC
        """
    ).fetchall()
    for row in campaign_rows:
        tenant_id = str(row["tenant_id"] or "")
        old_campaign_id = str(row["campaign_id"] or "").strip()
        if not old_campaign_id:
            continue
        new_campaign_id = old_campaign_id.lower() if _is_uuid(old_campaign_id) else campaign_map.get(old_campaign_id) or _new_uuid()
        campaign_map[old_campaign_id] = new_campaign_id
        _store_alias(con, entity_type="campaign", scope_id=tenant_id, old_id=old_campaign_id, new_id=new_campaign_id)

    for old_campaign_id, new_campaign_id in campaign_map.items():
        if old_campaign_id == new_campaign_id:
            continue
        for table in ["campaign_states", "campaign_revisions", "sync_events"]:
            columns = _table_columns(con, table) if _table_exists(con, table) else set()
            if "campaign_id" in columns:
                con.execute(
                    f"UPDATE {table} SET campaign_id = ? WHERE campaign_id = ?",
                    (new_campaign_id, old_campaign_id),
                )

    mappings: dict[str, str] = {}
    mappings.update({old: new for old, new in tenant_map.items() if old != new})
    mappings.update({old: new for old, new in user_map.items() if old != new})
    mappings.update({old: new for old, new in campaign_map.items() if old != new})

    for table, pk_columns, json_column in [
        ("users", ["tenant_id", "user_id"], "payload_json"),
        ("sync_snapshots", ["tenant_id", "server_sync_id"], "payload_json"),
        ("sync_snapshots", ["tenant_id", "server_sync_id"], "envelope_json"),
        ("campaign_states", ["tenant_id", "campaign_id"], "payload_json"),
        ("campaign_revisions", ["tenant_id", "campaign_id", "server_revision"], "payload_json"),
        ("device_enrollment_requests", ["tenant_id", "request_id"], "decision_note"),
        ("device_audit_log", ["id"], "payload_json"),
        ("official_referentials", ["tenant_id", "referential_id"], "payload_json"),
        ("official_referential_history", ["tenant_id", "history_id"], "payload_json"),
        ("sync_events", ["id"], "payload_json"),
        ("backup_audit_log", ["id"], "payload_json"),
    ]:
        if _table_exists(con, table):
            _rewrite_json_column(con, table, pk_columns, json_column, mappings)

    _record_migration(con, 156, "uuid_entity_ids_runtime_migration")
    con.execute("SET FOREIGN_KEY_CHECKS = 1")



def _delete_legacy_revoked_authorized_devices(con: Any) -> None:
    """Remove devices that were soft-revoked by older OpenIRN versions."""
    if _migration_applied(con, 161):
        return
    if not _table_exists(con, "authorized_devices"):
        _record_migration(con, 161, "delete_legacy_revoked_authorized_devices")
        return

    rows = con.execute(
        """
        SELECT tenant_id, device_id, name, platform, status, last_seen_at, revoked_at
        FROM authorized_devices
        WHERE status <> 'active' OR revoked_at IS NOT NULL
        """
    ).fetchall()
    for row in rows:
        tenant_id = str(row["tenant_id"] or "")
        device_id = str(row["device_id"] or "")
        if not tenant_id or not device_id:
            continue
        _record_device_audit(
            con,
            tenant_id,
            "device.deleted",
            device_id=device_id,
            payload={
                "reason": "legacy_revoked_device_cleanup",
                "name": row["name"],
                "platform": row["platform"],
                "previousStatus": row["status"],
                "lastSeenAt": row["last_seen_at"],
                "revokedAt": row["revoked_at"],
            },
        )
    con.execute(
        "DELETE FROM authorized_devices WHERE status <> 'active' OR revoked_at IS NOT NULL"
    )
    _record_migration(con, 161, "delete_legacy_revoked_authorized_devices")


def _index_exists(con: Any, table_name: str, index_name: str) -> bool:
    row = con.execute(
        """
        SELECT index_name
        FROM information_schema.statistics
        WHERE table_schema = DATABASE() AND table_name = ? AND index_name = ?
        LIMIT 1
        """,
        (table_name, index_name),
    ).fetchone()
    return row is not None


def _migrate_authorized_device_identity_schema(con: Any) -> None:
    if not _table_exists(con, "authorized_devices"):
        return
    if not _index_exists(con, "authorized_devices", "idx_authorized_devices_device_identity"):
        con.execute(
            """
            CREATE INDEX idx_authorized_devices_device_identity
            ON authorized_devices(device_id, status, last_seen_at)
            """
        )
    if not _migration_applied(con, 167):
        _record_migration(con, 167, "authorized_device_unique_identity_view")



def _migrate_terminal_identity_schema(con: Any) -> None:
    """Create and backfill the global terminal identity table.

    ``terminals`` stores the canonical identity of a physical terminal. The
    ``authorized_devices`` table remains the tenant-scoped authorization table.
    Enrollment may add a new authorization, but it must never rename an already
    known terminal.
    """
    if not _table_exists(con, "terminals"):
        con.execute(
            """
            CREATE TABLE terminals (
                device_id VARCHAR(160) NOT NULL,
                name VARCHAR(255) NOT NULL DEFAULT '',
                platform VARCHAR(64) NOT NULL DEFAULT '',
                created_at VARCHAR(40) NOT NULL,
                updated_at VARCHAR(40) NOT NULL,
                last_seen_at VARCHAR(40) NULL,
                PRIMARY KEY (device_id),
                KEY idx_terminals_updated (updated_at),
                KEY idx_terminals_last_seen (last_seen_at)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
            """
        )
    if _table_exists(con, "device_enrollment_requests"):
        request_columns = _table_columns(con, "device_enrollment_requests")
        if "device_id" not in request_columns:
            con.execute(
                """
                ALTER TABLE device_enrollment_requests
                ADD COLUMN device_id VARCHAR(160) NOT NULL DEFAULT '' AFTER request_id
                """
            )

    if _migration_applied(con, 168):
        return
    if not _table_exists(con, "authorized_devices"):
        _record_migration(con, 168, "terminal_identity_table")
        return

    rows = con.execute(
        """
        SELECT device_id, name, platform, status, created_at, last_seen_at
        FROM authorized_devices
        WHERE device_id IS NOT NULL AND device_id <> ''
        ORDER BY device_id ASC,
                 CASE WHEN status = 'active' THEN 0 ELSE 1 END ASC,
                 COALESCE(last_seen_at, created_at) DESC,
                 created_at ASC
        """
    ).fetchall()
    canonical: dict[str, dict[str, Any]] = {}
    for row in rows:
        device_id = str(row["device_id"] or "").strip()
        if not device_id or device_id in canonical:
            continue
        now = _utc_now().isoformat()
        canonical[device_id] = {
            "name": str(row["name"] or "").strip()[:120] or "Terminal OpenIRN",
            "platform": str(row["platform"] or "").strip()[:80],
            "created_at": str(row["created_at"] or now),
            "updated_at": str(row["last_seen_at"] or row["created_at"] or now),
            "last_seen_at": row["last_seen_at"],
        }

    for device_id, data in canonical.items():
        con.execute(
            """
            INSERT IGNORE INTO terminals(device_id, name, platform, created_at, updated_at, last_seen_at)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                device_id,
                data["name"],
                data["platform"],
                data["created_at"],
                data["updated_at"],
                data["last_seen_at"],
            ),
        )
        con.execute(
            """
            UPDATE authorized_devices
            SET name = ?, platform = ?
            WHERE device_id = ?
            """,
            (data["name"], data["platform"], device_id),
        )
    _record_migration(con, 168, "terminal_identity_table")


def _tenant_business_label_from_alias(value: str) -> str:
    raw = str(value or "").strip()
    if not raw:
        return "Espace de travail"
    if raw == DEFAULT_TENANT_ID:
        return "Défaut"
    label = re.sub(r"[_\-.]+", " ", raw).strip()
    if not label:
        return "Espace de travail"
    return " ".join(part[:1].upper() + part[1:] for part in label.split())


def _tenant_alias_label(con: Any, tenant_id: str) -> str:
    row = con.execute(
        """
        SELECT old_id
        FROM id_aliases
        WHERE entity_type = 'tenant' AND new_id = ?
        ORDER BY CASE WHEN old_id = ? THEN 0 ELSE 1 END, LENGTH(old_id) ASC, old_id ASC
        LIMIT 1
        """,
        (tenant_id, DEFAULT_TENANT_ID),
    ).fetchone()
    if row is None:
        return ""
    return _tenant_business_label_from_alias(str(row["old_id"] or ""))


def _is_generic_tenant_display_name(value: str, tenant_id: str = "") -> bool:
    label = str(value or "").strip()
    if not label:
        return True
    normalized = label.lower()
    return normalized == "espace de travail" or (tenant_id and label == tenant_id)


def _tenant_payload_from_row(con: Any, row: Any) -> dict[str, Any]:
    tenant_id = str(row["id"] or "")
    raw_display_name = str(_row_value(row, "display_name", "") or "").strip()
    display_name = raw_display_name
    if _is_generic_tenant_display_name(display_name, tenant_id):
        display_name = _tenant_alias_label(con, tenant_id) or display_name
    user_count = int(
        con.execute("SELECT COUNT(*) FROM users WHERE tenant_id = ?", (tenant_id,)).fetchone()[0]
    )
    active_user_count = int(
        con.execute(
            "SELECT COUNT(*) FROM users WHERE tenant_id = ? AND active = 1",
            (tenant_id,),
        ).fetchone()[0]
    )
    campaign_count = int(
        con.execute(
            "SELECT COUNT(*) FROM campaign_states WHERE tenant_id = ?",
            (tenant_id,),
        ).fetchone()[0]
    )
    pilot_count = int(
        con.execute(
            """
            SELECT COUNT(*) FROM users
            WHERE tenant_id = ? AND active = 1 AND role = 'campaign_manager'
            """,
            (tenant_id,),
        ).fetchone()[0]
    )
    admin_count = int(
        con.execute(
            """
            SELECT COUNT(*) FROM users
            WHERE tenant_id = ? AND active = 1 AND role = 'administrator'
            """,
            (tenant_id,),
        ).fetchone()[0]
    )
    return {
        "tenantId": tenant_id,
        "id": tenant_id,
        "displayName": display_name or "Espace de travail",
        "description": str(_row_value(row, "description", "") or ""),
        "permanent": bool(int(_row_value(row, "permanent", 0) or 0)),
        "isDefault": bool(int(_row_value(row, "permanent", 0) or 0)),
        "createdAt": row["created_at"],
        "updatedAt": row["updated_at"],
        "userCount": user_count,
        "activeUserCount": active_user_count,
        "pilotCount": pilot_count,
        "administratorCount": admin_count,
        "campaignCount": campaign_count,
    }




def _tenant_display_name(con: Any, tenant_id: str) -> str:
    safe_tenant_id = str(tenant_id or "").strip()
    if not safe_tenant_id:
        return "Espace de travail"
    row = con.execute(
        "SELECT display_name FROM tenants WHERE id = ?",
        (safe_tenant_id,),
    ).fetchone()
    if row is not None:
        display_name = str(_row_value(row, "display_name", "") or "").strip()
        if not _is_generic_tenant_display_name(display_name, safe_tenant_id):
            return display_name
    alias_label = _tenant_alias_label(con, safe_tenant_id)
    if alias_label:
        return alias_label
    if row is not None:
        display_name = str(_row_value(row, "display_name", "") or "").strip()
        if display_name and display_name != safe_tenant_id:
            return display_name
    return "Espace de travail"


def _user_display_name_from_parts(first_name: Any, last_name: Any, email: Any, fallback: str = "Utilisateur") -> str:
    parts = [str(first_name or "").strip(), str(last_name or "").strip()]
    full_name = " ".join(part for part in parts if part)
    email_value = str(email or "").strip()
    if full_name and email_value:
        return f"{full_name} <{email_value}>"
    if full_name:
        return full_name
    if email_value:
        return email_value
    return fallback

def _list_tenants(con: Any) -> list[dict[str, Any]]:
    _ensure_tenant(con, DEFAULT_TENANT_ID)
    rows = con.execute(
        """
        SELECT id, created_at, updated_at, display_name, description, permanent
        FROM tenants
        ORDER BY permanent DESC, display_name ASC, id ASC
        """,
    ).fetchall()
    return [_tenant_payload_from_row(con, row) for row in rows]


def _safe_tenant_id_for_creation(value: Any) -> str:
    tenant_id = _safe_segment(value, "").strip("._-")
    if not tenant_id:
        raise HTTPException(status_code=400, detail="Missing tenantId")
    if tenant_id != str(value or "").strip():
        raise HTTPException(
            status_code=400,
            detail="tenantId can only contain letters, digits, underscore, dot and dash",
        )
    if len(tenant_id) > 80:
        raise HTTPException(status_code=400, detail="tenantId is too long")
    return tenant_id



def _copy_user_to_tenant(
    con: Any,
    *,
    source_tenant_id: str,
    target_tenant_id: str,
    user_id: str,
) -> None:
    if not user_id:
        return
    source_user = con.execute(
        """
        SELECT user_id, first_name, last_name, email, role, active,
               created_at, updated_at, payload_json
        FROM users
        WHERE tenant_id = ? AND user_id = ? AND active = 1
        """,
        (source_tenant_id, user_id),
    ).fetchone()
    if source_user is None:
        return
    now = _utc_now().isoformat()
    payload = _parse_json(source_user["payload_json"], {})
    if not isinstance(payload, dict):
        payload = {}
    payload.update(
        {
            "id": source_user["user_id"],
            "firstName": source_user["first_name"],
            "lastName": source_user["last_name"],
            "email": source_user["email"],
            "role": source_user["role"],
            "active": bool(source_user["active"]),
            "createdAt": source_user["created_at"],
            "updatedAt": now,
        }
    )
    con.execute(
        """
        INSERT INTO users(
            tenant_id, user_id, first_name, last_name, email, role,
            active, created_at, updated_at, payload_json
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(tenant_id, user_id) DO NOTHING
        """,
        (
            target_tenant_id,
            source_user["user_id"],
            source_user["first_name"],
            source_user["last_name"],
            source_user["email"],
            source_user["role"],
            int(source_user["active"] or 0),
            source_user["created_at"],
            now,
            _canonical_json(payload),
        ),
    )
    credential = con.execute(
        """
        SELECT algorithm, iterations, salt, pin_hash, requires_change, updated_at
        FROM user_credentials
        WHERE tenant_id = ? AND user_id = ?
        """,
        (source_tenant_id, user_id),
    ).fetchone()
    if credential is not None:
        con.execute(
            """
            INSERT INTO user_credentials(
                tenant_id, user_id, algorithm, iterations, salt, pin_hash,
                requires_change, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(tenant_id, user_id) DO NOTHING
            """,
            (
                target_tenant_id,
                user_id,
                credential["algorithm"],
                credential["iterations"],
                credential["salt"],
                credential["pin_hash"],
                credential["requires_change"],
                credential["updated_at"],
            ),
        )


def _ensure_tenant(con: Any, tenant_id: str) -> None:
    now = _utc_now().isoformat()
    requested_tenant_id = _safe_segment(tenant_id, DEFAULT_TENANT_ID)
    resolved_tenant_id = _resolve_tenant_id(con, requested_tenant_id, DEFAULT_TENANT_ID)
    existing = con.execute("SELECT id FROM tenants WHERE id = ?", (resolved_tenant_id,)).fetchone()
    if existing is None:
        resolved_tenant_id = _tenant_id_for_creation(requested_tenant_id)
        display_name = "Défaut" if requested_tenant_id == DEFAULT_TENANT_ID else requested_tenant_id
        permanent = 1 if requested_tenant_id == DEFAULT_TENANT_ID else 0
        con.execute(
            """
            INSERT INTO tenants(id, created_at, updated_at, display_name, description, permanent)
            VALUES (?, ?, ?, ?, '', ?)
            """,
            (resolved_tenant_id, now, now, display_name, permanent),
        )
        _store_alias(con, entity_type="tenant", old_id=requested_tenant_id, new_id=resolved_tenant_id)
        return

    con.execute(
        """
        UPDATE tenants
        SET updated_at = CASE WHEN updated_at = '' THEN ? ELSE updated_at END,
            permanent = CASE WHEN ? = ? THEN 1 ELSE permanent END
        WHERE id = ?
        """,
        (now, requested_tenant_id, DEFAULT_TENANT_ID, resolved_tenant_id),
    )
    _store_alias(con, entity_type="tenant", old_id=requested_tenant_id, new_id=resolved_tenant_id)



def _backfill_default_tenant_display_name(con: Any) -> None:
    default_tenant_id = _resolve_tenant_id(con, DEFAULT_TENANT_ID, DEFAULT_TENANT_ID)
    if not default_tenant_id:
        return
    row = con.execute(
        "SELECT display_name FROM tenants WHERE id = ?",
        (default_tenant_id,),
    ).fetchone()
    if row is None:
        return
    display_name = str(_row_value(row, "display_name", "") or "").strip()
    if _is_generic_tenant_display_name(display_name, default_tenant_id):
        now = _utc_now().isoformat()
        con.execute(
            "UPDATE tenants SET display_name = ?, updated_at = ? WHERE id = ?",
            ("Défaut", now, default_tenant_id),
        )


def _sync_solution_administrators_to_all_tenants(con: Any) -> None:
    """Replicate solution administrator accounts to tenants.

    Terminal authorizations are intentionally not replicated here. A physical
    device may have a stable identity across the OpenIRN instance, but its
    enrollment remains tenant-scoped: being enrolled in one workspace never
    grants access to another workspace.
    """
    source_tenant_id = _resolve_tenant_id(con, SOLUTION_ADMIN_TENANT_ID, DEFAULT_TENANT_ID)
    if not source_tenant_id:
        return

    source_exists = con.execute(
        "SELECT 1 FROM tenants WHERE id = ?",
        (source_tenant_id,),
    ).fetchone()
    if source_exists is None:
        return

    administrators = con.execute(
        """
        SELECT user_id
        FROM users
        WHERE tenant_id = ? AND active = 1 AND role = 'administrator'
        ORDER BY user_id ASC
        """,
        (source_tenant_id,),
    ).fetchall()
    if not administrators:
        return

    tenants = con.execute(
        "SELECT id FROM tenants WHERE id <> ? ORDER BY id ASC",
        (source_tenant_id,),
    ).fetchall()
    if not tenants:
        return

    for tenant_row in tenants:
        target_tenant_id = str(tenant_row["id"] or "").strip()
        if not target_tenant_id:
            continue
        for administrator in administrators:
            _copy_user_to_tenant(
                con,
                source_tenant_id=source_tenant_id,
                target_tenant_id=target_tenant_id,
                user_id=str(administrator["user_id"] or ""),
            )



def _terminal_name_from_row(row: Any) -> str:
    return str(_row_value(row, "terminal_name", "") or _row_value(row, "name", "") or "").strip()[:120] or "Terminal OpenIRN"


def _terminal_platform_from_row(row: Any) -> str:
    return str(_row_value(row, "terminal_platform", "") or _row_value(row, "platform", "") or "").strip()[:80]


def _terminal_identity(con: Any, device_id: str) -> dict[str, Any] | None:
    normalized_device_id = _normalize_device_id(device_id)
    if not normalized_device_id or not _table_exists(con, "terminals"):
        return None
    row = con.execute(
        """
        SELECT device_id, name, platform, created_at, updated_at, last_seen_at
        FROM terminals
        WHERE device_id = ?
        """,
        (normalized_device_id,),
    ).fetchone()
    if row is None:
        return None
    return {
        "deviceId": str(row["device_id"] or ""),
        "name": str(row["name"] or "").strip()[:120] or "Terminal OpenIRN",
        "platform": str(row["platform"] or "").strip()[:80],
        "createdAt": row["created_at"],
        "updatedAt": row["updated_at"],
        "lastSeenAt": row["last_seen_at"],
    }


def _ensure_terminal_identity(
    con: Any,
    *,
    device_id: str,
    name: str,
    platform: str = "",
) -> tuple[dict[str, Any], bool]:
    """Ensure a global terminal identity exists.

    Returns ``(identity, already_known)``. Existing terminal names are never
    overwritten by enrollment: explicit renaming must go through the dedicated
    rename endpoint.
    """
    normalized_device_id = _normalize_device_id(device_id)
    if not normalized_device_id:
        raise HTTPException(status_code=400, detail="Invalid device id")
    clean_name = str(name or "").strip()[:120] or "Terminal OpenIRN"
    clean_platform = str(platform or "").strip()[:80]
    now = _utc_now().isoformat()
    existing = _terminal_identity(con, normalized_device_id)
    if existing is not None:
        canonical_platform = str(existing.get("platform") or "").strip()[:80]
        if not canonical_platform and clean_platform:
            con.execute(
                """
                UPDATE terminals
                SET platform = ?, updated_at = ?
                WHERE device_id = ?
                """,
                (clean_platform, now, normalized_device_id),
            )
            existing = _terminal_identity(con, normalized_device_id) or existing
        return existing, True
    con.execute(
        """
        INSERT INTO terminals(device_id, name, platform, created_at, updated_at, last_seen_at)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (normalized_device_id, clean_name, clean_platform, now, now, now),
    )
    return {
        "deviceId": normalized_device_id,
        "name": clean_name,
        "platform": clean_platform,
        "createdAt": now,
        "updatedAt": now,
        "lastSeenAt": now,
    }, False


def _touch_terminal(con: Any, device_id: str, when: str | None = None) -> None:
    normalized_device_id = _normalize_device_id(device_id)
    if not normalized_device_id:
        return
    seen_at = when or _utc_now().isoformat()
    con.execute(
        """
        UPDATE terminals
        SET last_seen_at = ?, updated_at = ?
        WHERE device_id = ?
        """,
        (seen_at, seen_at, normalized_device_id),
    )


def _device_tenant_membership_from_row(row: Any) -> dict[str, Any]:
    return {
        "tenantId": row["tenant_id"],
        "tenantDisplayName": str(_row_value(row, "tenant_display_name", "") or "").strip() or "Espace de travail",
        "status": row["status"],
        "createdAt": row["created_at"],
        "lastSeenAt": row["last_seen_at"],
        "revokedAt": row["revoked_at"],
        "invitedByUserId": row["invited_by_user_id"],
        "invitedByUserDisplayName": _user_display_name_from_parts(
            _row_value(row, "invited_by_first_name", ""),
            _row_value(row, "invited_by_last_name", ""),
            _row_value(row, "invited_by_email", ""),
            fallback="Utilisateur",
        ),
        "enrollmentId": row["enrollment_id"],
    }


def _device_from_row(row: Any) -> dict[str, Any]:
    membership = _device_tenant_membership_from_row(row)
    return {
        "tenantId": row["tenant_id"],
        "tenantDisplayName": membership["tenantDisplayName"],
        "deviceId": row["device_id"],
        "name": _terminal_name_from_row(row),
        "platform": _terminal_platform_from_row(row),
        "status": row["status"],
        "createdAt": row["created_at"],
        "lastSeenAt": row["last_seen_at"],
        "revokedAt": row["revoked_at"],
        "invitedByUserId": row["invited_by_user_id"],
        "invitedByUserDisplayName": membership["invitedByUserDisplayName"],
        "enrollmentId": row["enrollment_id"],
        "tenantIds": [membership["tenantId"]],
        "tenantLabels": [membership["tenantDisplayName"]],
        "tenantCount": 1,
        "tenants": [membership],
    }


def _is_newer_device_row(candidate: dict[str, Any], current: dict[str, Any]) -> bool:
    candidate_time = _parse_datetime(candidate.get("lastSeenAt") or candidate.get("createdAt"))
    current_time = _parse_datetime(current.get("lastSeenAt") or current.get("createdAt"))
    return candidate_time > current_time


def _merge_device_identity(existing: dict[str, Any], candidate: dict[str, Any]) -> dict[str, Any]:
    existing_tenants = list(existing.get("tenants") or [])
    candidate_tenants = list(candidate.get("tenants") or [])
    for membership in candidate_tenants:
        tenant_id = str(membership.get("tenantId") or "").strip()
        if not tenant_id:
            continue
        already = any(str(item.get("tenantId") or "").strip() == tenant_id for item in existing_tenants)
        if not already:
            existing_tenants.append(membership)

    representative = candidate if _is_newer_device_row(candidate, existing) else existing
    active_memberships = [item for item in existing_tenants if str(item.get("status") or "").lower() == "active" and not str(item.get("revokedAt") or "").strip()]
    status = "active" if active_memberships else "revoked"
    last_seen_values = [item.get("lastSeenAt") for item in existing_tenants if str(item.get("lastSeenAt") or "").strip()]
    created_values = [item.get("createdAt") for item in existing_tenants if str(item.get("createdAt") or "").strip()]
    revoked_values = [item.get("revokedAt") for item in existing_tenants if str(item.get("revokedAt") or "").strip()]
    last_seen_at = max(last_seen_values, key=_parse_datetime) if last_seen_values else representative.get("lastSeenAt")
    created_at = min(created_values, key=_parse_datetime) if created_values else representative.get("createdAt")
    revoked_at = "" if status == "active" else (max(revoked_values, key=_parse_datetime) if revoked_values else representative.get("revokedAt"))

    tenant_ids = [str(item.get("tenantId") or "").strip() for item in existing_tenants if str(item.get("tenantId") or "").strip()]
    tenant_labels = [str(item.get("tenantDisplayName") or "").strip() or "Espace de travail" for item in existing_tenants]
    existing.update(
        {
            "tenantId": representative.get("tenantId") or existing.get("tenantId"),
            "tenantDisplayName": representative.get("tenantDisplayName") or existing.get("tenantDisplayName"),
            "name": representative.get("name") or existing.get("name"),
            "platform": representative.get("platform") or existing.get("platform"),
            "status": status,
            "createdAt": created_at,
            "lastSeenAt": last_seen_at,
            "revokedAt": revoked_at,
            "invitedByUserId": representative.get("invitedByUserId") or existing.get("invitedByUserId"),
            "invitedByUserDisplayName": representative.get("invitedByUserDisplayName") or existing.get("invitedByUserDisplayName"),
            "enrollmentId": representative.get("enrollmentId") or existing.get("enrollmentId"),
            "tenantIds": tenant_ids,
            "tenantLabels": tenant_labels,
            "tenantCount": len(tenant_ids),
            "tenants": existing_tenants,
        }
    )
    return existing


def _group_devices_by_identity(devices: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = {}
    for device in devices:
        device_id = str(device.get("deviceId") or "").strip()
        if not device_id:
            continue
        if device_id not in grouped:
            grouped[device_id] = dict(device)
            continue
        grouped[device_id] = _merge_device_identity(grouped[device_id], device)
    return sorted(
        grouped.values(),
        key=lambda item: (
            1 if str(item.get("status") or "").lower() == "active" else 0,
            _parse_datetime(item.get("lastSeenAt") or item.get("createdAt")),
            str(item.get("name") or "").lower(),
        ),
        reverse=True,
    )


def _record_device_audit(
    con: Any,
    tenant_id: str,
    event_type: str,
    *,
    device_id: str | None = None,
    payload: dict[str, Any] | None = None,
) -> None:
    con.execute(
        """
        INSERT INTO device_audit_log(tenant_id, device_id, event_type, created_at, payload_json)
        VALUES (?, ?, ?, ?, ?)
        """,
        (
            tenant_id,
            device_id,
            event_type,
            _utc_now().isoformat(),
            _canonical_json(payload or {}),
        ),
    )


def _create_device(
    con: Any,
    tenant_id: str,
    *,
    name: str,
    platform: str = "",
    invited_by_user_id: str = "",
    enrollment_id: str = "",
    device_id: str = "",
) -> tuple[dict[str, Any], str]:
    requested_device_id = _normalize_device_id(device_id)
    effective_device_id = requested_device_id or f"device_{uuid.uuid4().hex}"
    terminal_identity, terminal_already_known = _ensure_terminal_identity(
        con,
        device_id=effective_device_id,
        name=name,
        platform=platform,
    )
    token = _new_device_token()
    now = _utc_now().isoformat()
    canonical_name = str(terminal_identity.get("name") or "").strip()[:120] or "Terminal OpenIRN"
    canonical_platform = str(terminal_identity.get("platform") or platform or "").strip()[:80]
    existing = con.execute(
        """
        SELECT 1
        FROM authorized_devices
        WHERE tenant_id = ? AND device_id = ?
        """,
        (tenant_id, effective_device_id),
    ).fetchone()
    if existing is None:
        con.execute(
            """
            INSERT INTO authorized_devices(
                tenant_id, device_id, name, platform, token_hash, status,
                created_at, last_seen_at, revoked_at, invited_by_user_id, enrollment_id
            ) VALUES (?, ?, ?, ?, ?, 'active', ?, ?, NULL, ?, ?)
            """,
            (
                tenant_id,
                effective_device_id,
                canonical_name,
                canonical_platform,
                _secret_hash(token),
                now,
                now,
                str(invited_by_user_id or "").strip(),
                str(enrollment_id or "").strip(),
            ),
        )
        event_type = "device.created"
    else:
        con.execute(
            """
            UPDATE authorized_devices
            SET name = ?, platform = ?, token_hash = ?, status = 'active',
                last_seen_at = ?, revoked_at = NULL, invited_by_user_id = ?, enrollment_id = ?
            WHERE tenant_id = ? AND device_id = ?
            """,
            (
                canonical_name,
                canonical_platform,
                _secret_hash(token),
                now,
                str(invited_by_user_id or "").strip(),
                str(enrollment_id or "").strip(),
                tenant_id,
                effective_device_id,
            ),
        )
        event_type = "device.reenrolled"
    _touch_terminal(con, effective_device_id, now)
    _record_device_audit(
        con,
        tenant_id,
        event_type,
        device_id=effective_device_id,
        payload={
            "name": canonical_name,
            "platform": canonical_platform,
            "invitedByUserId": invited_by_user_id or "",
            "enrollmentId": enrollment_id or "",
            "clientProvidedDeviceId": bool(requested_device_id),
            "terminalAlreadyKnown": terminal_already_known,
            "submittedNameIgnored": bool(terminal_already_known and str(name or "").strip() and str(name or "").strip()[:120] != canonical_name),
        },
    )
    row = con.execute(
        """
        SELECT d.tenant_id, COALESCE(NULLIF(t.display_name, ''), 'Espace de travail') AS tenant_display_name,
               d.device_id,
               COALESCE(NULLIF(tr.name, ''), d.name) AS name,
               COALESCE(NULLIF(tr.platform, ''), d.platform) AS platform,
               tr.name AS terminal_name,
               tr.platform AS terminal_platform,
               d.status, d.created_at,
               d.last_seen_at, d.revoked_at, d.invited_by_user_id, d.enrollment_id,
               u.first_name AS invited_by_first_name, u.last_name AS invited_by_last_name,
               u.email AS invited_by_email
        FROM authorized_devices d
        LEFT JOIN terminals tr ON tr.device_id = d.device_id
        LEFT JOIN tenants t ON t.id = d.tenant_id
        LEFT JOIN users u ON u.tenant_id = d.tenant_id AND u.user_id = d.invited_by_user_id
        WHERE d.tenant_id = ? AND d.device_id = ?
        """,
        (tenant_id, effective_device_id),
    ).fetchone()
    if row is None:
        raise HTTPException(status_code=500, detail="Device creation failed")
    return (_device_from_row(row), token)

def _list_devices(con: Any, tenant_id: str) -> list[dict[str, Any]]:
    rows = con.execute(
        """
        SELECT d.tenant_id, COALESCE(NULLIF(t.display_name, ''), 'Espace de travail') AS tenant_display_name,
               d.device_id,
               COALESCE(NULLIF(tr.name, ''), d.name) AS name,
               COALESCE(NULLIF(tr.platform, ''), d.platform) AS platform,
               tr.name AS terminal_name,
               tr.platform AS terminal_platform,
               d.status, d.created_at,
               d.last_seen_at, d.revoked_at, d.invited_by_user_id, d.enrollment_id,
               u.first_name AS invited_by_first_name, u.last_name AS invited_by_last_name,
               u.email AS invited_by_email
        FROM authorized_devices d
        LEFT JOIN terminals tr ON tr.device_id = d.device_id
        LEFT JOIN tenants t ON t.id = d.tenant_id
        LEFT JOIN users u ON u.tenant_id = d.tenant_id AND u.user_id = d.invited_by_user_id
        WHERE d.tenant_id = ?
        ORDER BY d.status ASC, COALESCE(d.last_seen_at, d.created_at) DESC, d.name ASC
        """,
        (tenant_id,),
    ).fetchall()
    return [_device_from_row(row) for row in rows]


def _list_all_devices(con: Any) -> list[dict[str, Any]]:
    rows = con.execute(
        """
        SELECT d.tenant_id, COALESCE(NULLIF(t.display_name, ''), 'Espace de travail') AS tenant_display_name,
               d.device_id,
               COALESCE(NULLIF(tr.name, ''), d.name) AS name,
               COALESCE(NULLIF(tr.platform, ''), d.platform) AS platform,
               tr.name AS terminal_name,
               tr.platform AS terminal_platform,
               d.status, d.created_at,
               d.last_seen_at, d.revoked_at, d.invited_by_user_id, d.enrollment_id,
               u.first_name AS invited_by_first_name, u.last_name AS invited_by_last_name,
               u.email AS invited_by_email
        FROM authorized_devices d
        LEFT JOIN terminals tr ON tr.device_id = d.device_id
        LEFT JOIN tenants t ON t.id = d.tenant_id
        LEFT JOIN users u ON u.tenant_id = d.tenant_id AND u.user_id = d.invited_by_user_id
        ORDER BY d.device_id ASC, d.status ASC, COALESCE(d.last_seen_at, d.created_at) DESC, d.name ASC
        """
    ).fetchall()
    return _group_devices_by_identity([_device_from_row(row) for row in rows])


def _enrollment_request_from_row(row: Any) -> dict[str, Any]:
    return {
        "tenantId": row["tenant_id"],
        "tenantDisplayName": str(_row_value(row, "tenant_display_name", "") or "").strip() or "Espace de travail",
        "requestId": row["request_id"],
        "deviceId": str(_row_value(row, "device_id", "") or ""),
        "deviceName": row["device_name"],
        "platform": row["platform"],
        "requesterNote": row["requester_note"],
        "status": row["status"],
        "requestedAt": row["requested_at"],
        "decidedAt": row["decided_at"],
        "decidedByUserId": row["decided_by_user_id"],
        "decidedByUserDisplayName": _user_display_name_from_parts(
            _row_value(row, "decided_by_first_name", ""),
            _row_value(row, "decided_by_last_name", ""),
            _row_value(row, "decided_by_email", ""),
            fallback="Utilisateur",
        ),
        "decisionNote": row["decision_note"],
        "enrollmentId": row["enrollment_id"],
    }


def _list_device_enrollment_requests(
    con: Any,
    tenant_id: str,
    *,
    include_all_tenants: bool = False,
    limit: int = 100,
) -> list[dict[str, Any]]:
    if include_all_tenants:
        rows = con.execute(
            """
            SELECT r.tenant_id, COALESCE(NULLIF(t.display_name, ''), 'Espace de travail') AS tenant_display_name,
                   r.request_id, r.device_id, r.device_name, r.platform, r.requester_note,
                   r.status, r.requested_at, r.decided_at, r.decided_by_user_id,
                   u.first_name AS decided_by_first_name, u.last_name AS decided_by_last_name,
                   u.email AS decided_by_email,
                   r.decision_note, r.enrollment_id
            FROM device_enrollment_requests r
            LEFT JOIN tenants t ON t.id = r.tenant_id
            LEFT JOIN users u ON u.tenant_id = r.tenant_id AND u.user_id = r.decided_by_user_id
            ORDER BY CASE WHEN r.status = 'pending' THEN 0 ELSE 1 END,
                     r.requested_at DESC
            LIMIT ?
            """,
            (max(1, min(limit, 500)),),
        ).fetchall()
    else:
        rows = con.execute(
            """
            SELECT r.tenant_id, COALESCE(NULLIF(t.display_name, ''), 'Espace de travail') AS tenant_display_name,
                   r.request_id, r.device_id, r.device_name, r.platform, r.requester_note,
                   r.status, r.requested_at, r.decided_at, r.decided_by_user_id,
                   u.first_name AS decided_by_first_name, u.last_name AS decided_by_last_name,
                   u.email AS decided_by_email,
                   r.decision_note, r.enrollment_id
            FROM device_enrollment_requests r
            LEFT JOIN tenants t ON t.id = r.tenant_id
            LEFT JOIN users u ON u.tenant_id = r.tenant_id AND u.user_id = r.decided_by_user_id
            WHERE r.tenant_id = ?
            ORDER BY CASE WHEN r.status = 'pending' THEN 0 ELSE 1 END,
                     r.requested_at DESC
            LIMIT ?
            """,
            (tenant_id, max(1, min(limit, 200))),
        ).fetchall()
    return [_enrollment_request_from_row(row) for row in rows]


ADRI_PILLAR_RE = re.compile(r"^RES-[1-8]$")
ADRI_CRITERION_RE = re.compile(r"^(RES-[1-8])([.-])([0-9]+)$")
ADRI_VERSION_RE = re.compile(r"(?:Questionnaire_IRN|Référentiel\s+IRN)_v\.?([0-9]+(?:\.[0-9]+)*)\.xlsx$", re.IGNORECASE)

ADRI_DEFAULT_PILLAR_LABELS = {
    "RES-1": "Résilience stratégique",
    "RES-2": "Résilience économique et juridique",
    "RES-3": "Résilience Data & IA",
    "RES-4": "Résilience opérationnelle",
    "RES-5": "Résilience Supply-Chain",
    "RES-6": "Résilience Technologique",
    "RES-7": "Résilience Sécurité",
    "RES-8": "Résilience Environnementale",
}

ADRI_EXPECTED_HEADERS = {
    "Dimension": "pillarId",
    "ID": "sourceCode",
    "Intitulé du critère": "label",
    "Critère": "shortLabel",
    "Description (objectif)": "description",
    "Portée du critère": "sourceScope",
    "Références réglementaires (TBD)": "regulatoryReferences",
    "Recommandations": "recommendations",
}

OPENIRN_RNR_SCORING_METADATA = {
    "method": "Moyenne des niveaux IRN : NR=10, Intention=25, Moyen=50, Résultat=95",
    "methodLabel": "Score OpenIRN",
    "methodStatus": "irn_scale_unweighted_v1",
    "notAnsweredPolicy": "not_answered_excluded_nc_excluded_from_score_included_in_completion",
    "criteriaWeightPolicy": "uniform_per_scored_criterion",
    "globalAggregationPolicy": "average_numeric_scored_criteria",
    "weightedOfficialMethodImplemented": False,
    "officialWeightedMethodStatus": "implemented_public_level_scale_without_additional_weighting",
    "disclaimer": (
        "OpenIRN applique la grille de notation IRN : N.C. exclu du score, "
        "Non résilient = 10/100, Intention = 25/100, Moyen = 50/100, "
        "Résultat = 95/100. Aucune pondération additionnelle n’est appliquée."
    ),
}


def _adri_version_from_filename(name: str) -> str:
    match = ADRI_VERSION_RE.search(str(name or ""))
    return f"v{match.group(1)}" if match else "unknown"


def _adri_version_key(version: str) -> tuple[int, ...]:
    raw = str(version or "").lower().lstrip("v")
    parts = []
    for item in raw.split("."):
        try:
            parts.append(int(item))
        except ValueError:
            parts.append(0)
    return tuple(parts or [0])


def _adri_safe_filename(value: str) -> str:
    cleaned = str(value or "").lower().replace(".", "_").replace("-", "_")
    cleaned = re.sub(r"[^a-z0-9_]+", "_", cleaned)
    return re.sub(r"_+", "_", cleaned).strip("_") or "adri_irn"


def _adri_norm(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _adri_clean_multiline(value: Any) -> str:
    return re.sub(r"\s+", " ", _adri_norm(value)).strip()


def _adri_normalize_code(value: Any) -> str:
    raw = _adri_norm(value)
    match = ADRI_CRITERION_RE.match(raw)
    if not match:
        return raw
    return f"{match.group(1)}.{match.group(3)}"


def _adri_scope_from_label(value: Any) -> str:
    raw = _adri_norm(value).lower()
    if "actif" in raw:
        return "asset"
    if "système" in raw or "systeme" in raw:
        return "criticalSystem"
    if "fonction" in raw:
        return "organization"
    if "organisation" in raw:
        return "organization"
    return "unknown"


def _adri_sort_key(code: str) -> tuple[int, int]:
    criterion_match = ADRI_CRITERION_RE.match(str(code or ""))
    if criterion_match:
        return int(criterion_match.group(1).split("-")[1]), int(criterion_match.group(3))
    pillar_match = re.match(r"^RES-([1-8])$", str(code or ""))
    if pillar_match:
        return int(pillar_match.group(1)), 0
    return 99, 99


def _adri_find_header_row(ws: Any) -> tuple[int, dict[int, str]]:
    for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
        values = [_adri_norm(value) for value in row]
        if "Dimension" in values and "ID" in values and "Intitulé du critère" in values:
            return row_index, {index: value for index, value in enumerate(values) if value}
    raise HTTPException(status_code=422, detail=f"Ligne d'en-tête introuvable dans l'onglet {ws.title!r}")


def _adri_parse_pillar_labels_from_grid(wb: Any) -> dict[str, str]:
    labels = dict(ADRI_DEFAULT_PILLAR_LABELS)
    if "Grille V1" not in wb.sheetnames:
        return labels
    ws = wb["Grille V1"]
    for row in ws.iter_rows(values_only=True):
        first_cell = _adri_norm(row[0] if row else None)
        if not first_cell or "RES-" not in first_cell:
            continue
        match = re.search(r"(RES-[1-8])", first_cell)
        if not match:
            continue
        code = match.group(1)
        label = re.sub(r"\(?\s*RES-[1-8]\s*\)?", "", first_cell)
        label = _adri_clean_multiline(label)
        if label:
            labels[code] = label
    return labels


def _adri_parse_workbook(raw_xlsx: bytes, *, version: str, remote: dict[str, Any]) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise HTTPException(status_code=503, detail="La dépendance Python openpyxl est requise pour importer le référentiel officiel") from exc

    wb = load_workbook(BytesIO(raw_xlsx), data_only=True, read_only=True)
    if "Référentiel v1" not in wb.sheetnames:
        raise HTTPException(status_code=422, detail="Onglet 'Référentiel v1' introuvable dans le fichier officiel")

    warnings: list[str] = []
    pillar_labels = _adri_parse_pillar_labels_from_grid(wb)
    ws = wb["Référentiel v1"]
    header_row, header_by_index = _adri_find_header_row(ws)
    criteria_by_code: dict[str, dict[str, Any]] = {}
    seen_source_codes: set[str] = set()

    for row_number, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        raw: dict[str, str] = {}
        for index, header in header_by_index.items():
            if header in ADRI_EXPECTED_HEADERS:
                raw[ADRI_EXPECTED_HEADERS[header]] = _adri_norm(row[index] if index < len(row) else None)

        source_code = raw.get("sourceCode", "")
        if not ADRI_CRITERION_RE.match(source_code):
            continue

        code = _adri_normalize_code(source_code)
        pillar_id = raw.get("pillarId") or code.split(".")[0]
        if not ADRI_PILLAR_RE.match(pillar_id):
            warnings.append(f"Ligne {row_number}: dimension invalide {pillar_id!r} pour {source_code}")
            pillar_id = code.split(".")[0]

        if source_code != code:
            warnings.append(f"Ligne {row_number}: identifiant normalisé {source_code!r} -> {code!r}")
        if source_code in seen_source_codes:
            warnings.append(f"Ligne {row_number}: doublon sourceCode {source_code!r}")
        seen_source_codes.add(source_code)

        label = _adri_clean_multiline(raw.get("label", "")) or _adri_clean_multiline(raw.get("shortLabel", ""))
        if not label:
            warnings.append(f"Ligne {row_number}: critère {code} sans libellé")

        criteria_by_code[code] = {
            "id": code,
            "code": code,
            "sourceCode": source_code,
            "pillarId": pillar_id,
            "label": label,
            "shortLabel": _adri_clean_multiline(raw.get("shortLabel", "")),
            "description": _adri_clean_multiline(raw.get("description", "")),
            "scope": _adri_scope_from_label(raw.get("sourceScope", "")),
            "sourceScope": _adri_clean_multiline(raw.get("sourceScope", "")),
            "answerMode": "R_NR",
            "regulatoryReferences": _adri_clean_multiline(raw.get("regulatoryReferences", "")),
            "recommendations": _adri_clean_multiline(raw.get("recommendations", "")),
            "active": True,
            "source": {"sheet": ws.title, "row": row_number},
        }

    pillars = [
        {"id": code, "code": code, "label": pillar_labels.get(code, code)}
        for code in sorted(pillar_labels.keys(), key=_adri_sort_key)
        if ADRI_PILLAR_RE.match(code)
    ]
    criteria = sorted(criteria_by_code.values(), key=lambda criterion: _adri_sort_key(criterion["code"]))

    if len(pillars) != 8:
        warnings.append(f"Nombre de piliers inattendu: {len(pillars)} au lieu de 8")
    if not criteria:
        warnings.append("Aucun critère extrait: vérifier la structure du fichier officiel et le mapping des colonnes")

    return {
        "id": f"adri-irn-{version}",
        "version": version,
        "importedAt": _utc_now().isoformat(),
        "source": {
            "type": "gitlab",
            "url": OFFICIAL_ADRI_SOURCE_URL,
            "projectPath": OFFICIAL_ADRI_PROJECT_PATH,
            "defaultBranch": remote.get("defaultBranch") or OFFICIAL_ADRI_DEFAULT_BRANCH,
            "filePath": remote.get("filePath") or "",
            "commitSha": remote.get("commitSha") or "",
            "blobId": remote.get("blobId") or "",
            "checksumSha256": hashlib.sha256(raw_xlsx).hexdigest(),
            "license": OFFICIAL_ADRI_LICENSE,
        },
        "scoring": dict(OPENIRN_RNR_SCORING_METADATA),
        "pillars": pillars,
        "criteria": criteria,
        "importWarnings": [
            "Référentiel officiel téléchargé depuis le dépôt public aDRI IRN.",
            "Le serveur conserve une copie canonique JSON pour OpenIRN.",
            *warnings,
        ],
    }


def _adri_validation_report(referential: dict[str, Any]) -> dict[str, Any]:
    errors: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = []
    pillars = referential.get("pillars") if isinstance(referential.get("pillars"), list) else []
    criteria = referential.get("criteria") if isinstance(referential.get("criteria"), list) else []
    if len(pillars) != 8:
        errors.append({"code": "unexpected_pillar_count", "message": "Le référentiel doit contenir 8 piliers.", "found": len(pillars)})
    if not criteria:
        errors.append({"code": "no_criteria", "message": "Aucun critère n'a été extrait du référentiel."})

    pillar_ids = {str(pillar.get("id") or pillar.get("code") or "") for pillar in pillars if isinstance(pillar, dict)}
    criterion_codes: set[str] = set()
    for criterion in criteria:
        if not isinstance(criterion, dict):
            continue
        code = str(criterion.get("code") or criterion.get("id") or "")
        if not re.match(r"^RES-[1-8]\.[0-9]+$", code):
            errors.append({"code": "invalid_criterion_code", "message": "Code critère invalide.", "criterion": code})
            continue
        if code in criterion_codes:
            errors.append({"code": "duplicate_criterion_code", "message": "Code critère dupliqué.", "criterion": code})
        criterion_codes.add(code)
        pillar_id = str(criterion.get("pillarId") or "")
        if pillar_id not in pillar_ids:
            errors.append({"code": "unknown_pillar", "message": "Le critère référence un pilier inconnu.", "criterion": code, "pillarId": pillar_id})
        if not str(criterion.get("label") or "").strip():
            warnings.append({"code": "missing_label", "message": "Libellé critère absent.", "criterion": code})

    status = "failed" if errors else ("passed_with_warnings" if warnings else "passed")
    return {
        "status": status,
        "generatedAt": _utc_now().isoformat(),
        "errors": errors,
        "warnings": warnings,
        "summary": {
            "referentialId": referential.get("id"),
            "version": referential.get("version"),
            "pillars": len(pillars),
            "criteria": len(criteria),
            "scoringMethodStatus": OPENIRN_RNR_SCORING_METADATA["methodStatus"],
            "weightedOfficialMethodImplemented": OPENIRN_RNR_SCORING_METADATA["weightedOfficialMethodImplemented"],
        },
    }


def _gitlab_quote(value: str) -> str:
    return urllib.parse.quote(str(value), safe="")


def _gitlab_request_json(path: str, query: dict[str, str] | None = None) -> Any:
    query_string = f"?{urllib.parse.urlencode(query or {})}" if query else ""
    url = f"{OFFICIAL_ADRI_GITLAB_API}{path}{query_string}"
    request = urllib.request.Request(url, headers={"Accept": "application/json", "User-Agent": "OpenIRN"})
    try:
        with urllib.request.urlopen(request, timeout=20) as response:
            raw = response.read()
    except urllib.error.HTTPError as exc:
        raise HTTPException(status_code=502, detail=f"GitLab a répondu HTTP {exc.code} pour {url}") from exc
    except urllib.error.URLError as exc:
        raise HTTPException(status_code=502, detail=f"GitLab est injoignable: {exc.reason}") from exc
    try:
        return json.loads(raw.decode("utf-8"))
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=502, detail="Réponse GitLab JSON invalide") from exc


def _gitlab_request_bytes(path: str, query: dict[str, str] | None = None) -> bytes:
    query_string = f"?{urllib.parse.urlencode(query or {})}" if query else ""
    url = f"{OFFICIAL_ADRI_GITLAB_API}{path}{query_string}"
    request = urllib.request.Request(url, headers={"Accept": "application/octet-stream", "User-Agent": "OpenIRN"})
    try:
        with urllib.request.urlopen(request, timeout=45) as response:
            return response.read()
    except urllib.error.HTTPError as exc:
        raise HTTPException(status_code=502, detail=f"Téléchargement GitLab refusé HTTP {exc.code} pour {url}") from exc
    except urllib.error.URLError as exc:
        raise HTTPException(status_code=502, detail=f"GitLab est injoignable: {exc.reason}") from exc


def _gitlab_repository_tree(
    project_id: str,
    *,
    tree_path: str,
    ref: str,
    recursive: bool = False,
) -> list[dict[str, Any]]:
    query = {
        "ref": ref,
        "per_page": "100",
    }
    if tree_path:
        query["path"] = tree_path
    if recursive:
        query["recursive"] = "true"

    tree = _gitlab_request_json(
        f"/projects/{project_id}/repository/tree",
        query,
    )
    if not isinstance(tree, list):
        raise HTTPException(status_code=502, detail="Réponse GitLab repository/tree invalide")
    return [item for item in tree if isinstance(item, dict)]


def _gitlab_latest_commit_for_file(project_id: str, file_path: str, ref: str) -> str:
    if not file_path:
        return ""
    try:
        commits = _gitlab_request_json(
            f"/projects/{project_id}/repository/commits",
            {"ref_name": ref, "path": file_path, "per_page": "1"},
        )
    except HTTPException:
        # Commit SHA is an audit enrichment only. The blob id remains the
        # stable content marker used for update detection.
        return ""
    if isinstance(commits, list) and commits:
        first = commits[0]
        if isinstance(first, dict):
            return str(first.get("id") or first.get("short_id") or "")
    return ""


def _official_remote_latest() -> dict[str, Any]:
    project_id = _gitlab_quote(OFFICIAL_ADRI_PROJECT_PATH)
    project = _gitlab_request_json(f"/projects/{project_id}")
    default_branch = str(project.get("default_branch") or OFFICIAL_ADRI_DEFAULT_BRANCH)

    candidate_tree_paths = [
        OFFICIAL_ADRI_TREE_PATH,
        "Grille d'évaluation IRN (FR)/xlsx",
        "Grille d'évaluation IRN (FR)",
        "Référentiel d'évaluation IRN (FR)",
        "",
    ]

    seen_paths: set[str] = set()
    candidates: list[dict[str, Any]] = []
    last_error: HTTPException | None = None

    for tree_path in candidate_tree_paths:
        if tree_path in seen_paths:
            continue
        seen_paths.add(tree_path)

        try:
            tree = _gitlab_repository_tree(
                project_id,
                tree_path=tree_path,
                ref=default_branch,
                recursive=True,
            )
        except HTTPException as exc:
            last_error = exc
            detail = str(exc.detail)
            if "HTTP 404" in detail or exc.status_code == 404:
                continue
            raise

        for item in tree:
            if item.get("type") != "blob":
                continue
            name = str(item.get("name") or "")
            if not ADRI_VERSION_RE.search(name):
                continue
            version = _adri_version_from_filename(name)
            file_path = str(item.get("path") or "")
            commit_sha = _gitlab_latest_commit_for_file(project_id, file_path, default_branch)
            candidates.append({
                "version": version,
                "fileName": name,
                "filePath": file_path,
                "blobId": str(item.get("id") or ""),
                "commitSha": commit_sha,
                "defaultBranch": default_branch,
                "projectPath": OFFICIAL_ADRI_PROJECT_PATH,
                "sourceUrl": OFFICIAL_ADRI_SOURCE_URL,
                "webUrl": f"{OFFICIAL_ADRI_SOURCE_URL}/-/blob/{urllib.parse.quote(default_branch, safe='')}/{urllib.parse.quote(file_path)}",
            })

        if candidates:
            break

    if not candidates:
        if last_error is not None:
            detail = str(last_error.detail)
            if "HTTP 404" in detail:
                detail = (
                    "Le chemin GitLab configuré pour le référentiel aDRI est introuvable. "
                    "Le dépôt officiel utilise maintenant le dossier "
                    "Grille d'évaluation IRN (FR)/xlsx. "
                    "Vérifiez OPENIRN_ADRI_TREE_PATH si cette variable est définie."
                )
                raise HTTPException(status_code=404, detail=detail) from last_error
        raise HTTPException(
            status_code=404,
            detail="Aucun fichier Référentiel IRN_v*.xlsx ou Questionnaire_IRN_v*.xlsx trouvé dans le dépôt aDRI",
        )
    candidates.sort(key=lambda item: _adri_version_key(str(item.get("version") or "")), reverse=True)
    return candidates[0]


def _download_official_adri_xlsx(remote: dict[str, Any]) -> bytes:
    project_id = _gitlab_quote(OFFICIAL_ADRI_PROJECT_PATH)
    file_path = _gitlab_quote(str(remote.get("filePath") or ""))
    raw = _gitlab_request_bytes(
        f"/projects/{project_id}/repository/files/{file_path}/raw",
        {"ref": str(remote.get("defaultBranch") or OFFICIAL_ADRI_DEFAULT_BRANCH)},
    )
    if len(raw) < 1024 or not raw.startswith(b"PK"):
        raise HTTPException(status_code=502, detail="Le fichier téléchargé depuis GitLab ne ressemble pas à un fichier XLSX valide")
    return raw



def _row_value(row: Any, key: str, fallback: Any = "") -> Any:
    return row[key] if key in row.keys() else fallback


def _official_referential_web_url(source_url: Any, default_branch: Any, file_path: Any) -> str:
    base = str(source_url or "").strip().rstrip("/")
    branch = str(default_branch or "").strip()
    path = str(file_path or "").strip()
    if not base or not branch or not path:
        return ""
    return f"{base}/-/blob/{urllib.parse.quote(branch, safe='')}/{urllib.parse.quote(path)}"


def _official_referential_summary_from_row(row: Any | None) -> dict[str, Any] | None:
    if row is None:
        return None
    report = _parse_json(row["validation_report_json"], {})
    payload = _parse_json(_row_value(row, "payload_json", "{}"), {})
    source_metadata = payload.get("source") if isinstance(payload, dict) else {}
    if not isinstance(source_metadata, dict):
        source_metadata = {}
    scoring_metadata = payload.get("scoring") if isinstance(payload, dict) else {}
    if not isinstance(scoring_metadata, dict):
        scoring_metadata = dict(OPENIRN_RNR_SCORING_METADATA)
    source_url = row["source_url"]
    default_branch = row["default_branch"]
    file_path = row["file_path"]
    source_blob_id = str(source_metadata.get("blobId") or row["source_blob_id"] or "")
    source_commit_sha = str(source_metadata.get("commitSha") or "")
    return {
        "historyId": _row_value(row, "history_id", ""),
        "referentialId": row["referential_id"],
        "version": row["version"],
        "active": bool(row["active"]),
        "sourceUrl": source_url,
        "projectPath": row["project_path"],
        "defaultBranch": default_branch,
        "filePath": file_path,
        "sourceBlobId": source_blob_id,
        "sourceCommitSha": source_commit_sha,
        "sourceSha256": row["source_sha256"],
        "canonicalSha256": row["canonical_sha256"],
        "downloadedAt": row["downloaded_at"],
        "importedAt": row["imported_at"],
        "pillarCount": row["pillar_count"],
        "criterionCount": row["criterion_count"],
        "triggeredByUserId": _row_value(row, "triggered_by_user_id", ""),
        "validationStatus": report.get("status") if isinstance(report, dict) else "unknown",
        "scoringMethod": scoring_metadata,
        "webUrl": _official_referential_web_url(source_url, default_branch, file_path),
    }


def _load_global_official_referential(
    con: Any,
    *,
    preferred_tenant_id: str = "",
) -> Any | None:
    """Return the active official referential shared by the OpenIRN instance.

    Users, campaigns, sessions and devices remain tenant-scoped, but the
    official aDRI/IRN referential is an instance-level resource.  Older
    versions of OpenIRN stored it with a tenant_id; this helper keeps backwards
    compatibility by selecting the best active row across tenants.
    """
    preferred = _resolve_tenant_id(con, preferred_tenant_id, "") if preferred_tenant_id else ""
    solution_tenant_id = _resolve_tenant_id(con, SOLUTION_ADMIN_TENANT_ID, DEFAULT_TENANT_ID)
    default_tenant_id = _default_tenant_id(con)
    rows = con.execute(
        """
        SELECT tenant_id, referential_id, version, active, source_url, project_path,
               default_branch, file_path, source_blob_id, source_sha256,
               canonical_sha256, downloaded_at, imported_at, pillar_count,
               criterion_count, import_warnings_json, validation_report_json, payload_json
        FROM official_referentials
        WHERE active = 1
        ORDER BY
            CASE
                WHEN tenant_id = ? THEN 0
                WHEN tenant_id = ? THEN 1
                WHEN tenant_id = ? THEN 2
                ELSE 3
            END,
            imported_at DESC
        LIMIT 1
        """,
        (preferred, solution_tenant_id, default_tenant_id),
    ).fetchone()
    return rows


def _load_current_official_referential(con: Any, tenant_id: str) -> Any | None:
    requested_tenant_id = _resolve_tenant_id(con, tenant_id, DEFAULT_TENANT_ID)
    local = con.execute(
        """
        SELECT tenant_id, referential_id, version, active, source_url, project_path,
               default_branch, file_path, source_blob_id, source_sha256,
               canonical_sha256, downloaded_at, imported_at, pillar_count,
               criterion_count, import_warnings_json, validation_report_json, payload_json
        FROM official_referentials
        WHERE tenant_id = ? AND active = 1
        ORDER BY imported_at DESC
        LIMIT 1
        """,
        (requested_tenant_id,),
    ).fetchone()
    if local is not None:
        return local
    return _load_global_official_referential(
        con,
        preferred_tenant_id=requested_tenant_id,
    )


def _official_referential_history_id(referential_id: str, canonical_sha256: str, imported_at: str) -> str:
    normalized_imported_at = re.sub(r"[^0-9A-Za-z]+", "", imported_at)[:24]
    digest = str(canonical_sha256 or "")[:12] or uuid.uuid4().hex[:12]
    safe_referential_id = _adri_safe_filename(referential_id or "adri-irn")
    return f"{safe_referential_id}-{normalized_imported_at}-{digest}"[:180]


def _backfill_official_referential_history(con: Any) -> None:
    rows = con.execute(
        """
        SELECT tenant_id, referential_id, version, active, source_url, project_path,
               default_branch, file_path, source_blob_id, source_sha256,
               canonical_sha256, downloaded_at, imported_at, pillar_count,
               criterion_count, import_warnings_json, validation_report_json, payload_json
        FROM official_referentials
        """
    ).fetchall()
    for row in rows:
        existing_history = con.execute(
            """
            SELECT 1
            FROM official_referential_history
            WHERE tenant_id = ?
              AND referential_id = ?
              AND canonical_sha256 = ?
              AND source_sha256 = ?
            LIMIT 1
            """,
            (
                row["tenant_id"],
                row["referential_id"],
                row["canonical_sha256"],
                row["source_sha256"],
            ),
        ).fetchone()
        if existing_history is not None:
            continue

        history_id = _official_referential_history_id(
            str(row["referential_id"] or "adri-irn"),
            str(row["canonical_sha256"] or row["source_sha256"] or ""),
            str(row["imported_at"] or row["downloaded_at"] or _utc_now().isoformat()),
        )
        con.execute(
            """
            INSERT OR IGNORE INTO official_referential_history(
                tenant_id, history_id, referential_id, version, active,
                source_url, project_path, default_branch, file_path,
                source_blob_id, source_sha256, canonical_sha256,
                downloaded_at, imported_at, pillar_count, criterion_count,
                triggered_by_user_id, import_warnings_json,
                validation_report_json, payload_json
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, '', ?, ?, ?)
            """,
            (
                row["tenant_id"],
                history_id,
                row["referential_id"],
                row["version"],
                row["active"],
                row["source_url"],
                row["project_path"],
                row["default_branch"],
                row["file_path"],
                row["source_blob_id"],
                row["source_sha256"],
                row["canonical_sha256"],
                row["downloaded_at"],
                row["imported_at"],
                row["pillar_count"],
                row["criterion_count"],
                row["import_warnings_json"],
                row["validation_report_json"],
                row["payload_json"],
            ),
        )



def _seed_official_referential_from_default(
    con: Any,
    tenant_id: str,
    *,
    source_tenant_id: str = DEFAULT_TENANT_ID,
) -> bool:
    """Seed a tenant with the default official referential when none is active.

    Tenants remain isolated: this only creates tenant-local copies of the
    official referential metadata/payload. Campaigns, users, sessions and device
    authorizations are never copied by this helper.
    """
    target_tenant_id = _resolve_tenant_id(con, tenant_id, DEFAULT_TENANT_ID)
    source_tenant_id = _resolve_tenant_id(con, source_tenant_id, DEFAULT_TENANT_ID)
    if target_tenant_id == source_tenant_id:
        return False

    existing = con.execute(
        """
        SELECT 1
        FROM official_referentials
        WHERE tenant_id = ? AND active = 1
        LIMIT 1
        """,
        (target_tenant_id,),
    ).fetchone()
    if existing is not None:
        return False

    source = con.execute(
        """
        SELECT tenant_id, referential_id, version, active, source_url, project_path,
               default_branch, file_path, source_blob_id, source_sha256,
               canonical_sha256, downloaded_at, imported_at, pillar_count,
               criterion_count, import_warnings_json, validation_report_json, payload_json
        FROM official_referentials
        WHERE tenant_id = ? AND active = 1
        ORDER BY imported_at DESC
        LIMIT 1
        """,
        (source_tenant_id,),
    ).fetchone()
    if source is None:
        return False

    _ensure_tenant(con, target_tenant_id)
    con.execute(
        """
        INSERT INTO official_referentials(
            tenant_id, referential_id, version, active, source_url, project_path,
            default_branch, file_path, source_blob_id, source_sha256,
            canonical_sha256, downloaded_at, imported_at, pillar_count,
            criterion_count, import_warnings_json, validation_report_json, payload_json
        ) VALUES (?, ?, ?, 1, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(tenant_id, referential_id) DO UPDATE SET
            active = 1,
            version = excluded.version,
            source_url = excluded.source_url,
            project_path = excluded.project_path,
            default_branch = excluded.default_branch,
            file_path = excluded.file_path,
            source_blob_id = excluded.source_blob_id,
            source_sha256 = excluded.source_sha256,
            canonical_sha256 = excluded.canonical_sha256,
            downloaded_at = excluded.downloaded_at,
            imported_at = excluded.imported_at,
            pillar_count = excluded.pillar_count,
            criterion_count = excluded.criterion_count,
            import_warnings_json = excluded.import_warnings_json,
            validation_report_json = excluded.validation_report_json,
            payload_json = excluded.payload_json
        """,
        (
            target_tenant_id,
            source["referential_id"],
            source["version"],
            source["source_url"],
            source["project_path"],
            source["default_branch"],
            source["file_path"],
            source["source_blob_id"],
            source["source_sha256"],
            source["canonical_sha256"],
            source["downloaded_at"],
            source["imported_at"],
            source["pillar_count"],
            source["criterion_count"],
            source["import_warnings_json"],
            source["validation_report_json"],
            source["payload_json"],
        ),
    )

    _backfill_official_referential_history(con)
    source_history_rows = con.execute(
        """
        SELECT history_id, referential_id, version, active, source_url, project_path,
               default_branch, file_path, source_blob_id, source_sha256,
               canonical_sha256, downloaded_at, imported_at, pillar_count,
               criterion_count, triggered_by_user_id, import_warnings_json,
               validation_report_json, payload_json
        FROM official_referential_history
        WHERE tenant_id = ?
        ORDER BY imported_at ASC, history_id ASC
        """,
        (source_tenant_id,),
    ).fetchall()
    for row in source_history_rows:
        con.execute(
            """
            INSERT OR IGNORE INTO official_referential_history(
                tenant_id, history_id, referential_id, version, active,
                source_url, project_path, default_branch, file_path,
                source_blob_id, source_sha256, canonical_sha256,
                downloaded_at, imported_at, pillar_count, criterion_count,
                triggered_by_user_id, import_warnings_json,
                validation_report_json, payload_json
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                target_tenant_id,
                row["history_id"],
                row["referential_id"],
                row["version"],
                row["active"],
                row["source_url"],
                row["project_path"],
                row["default_branch"],
                row["file_path"],
                row["source_blob_id"],
                row["source_sha256"],
                row["canonical_sha256"],
                row["downloaded_at"],
                row["imported_at"],
                row["pillar_count"],
                row["criterion_count"],
                row["triggered_by_user_id"],
                row["import_warnings_json"],
                row["validation_report_json"],
                row["payload_json"],
            ),
        )
    return True

def _list_official_referential_history(
    con: Any,
    tenant_id: str,
    limit: int = 50,
) -> list[dict[str, Any]]:
    _backfill_official_referential_history(con)
    rows = con.execute(
        """
        SELECT tenant_id, history_id, referential_id, version, active,
               source_url, project_path, default_branch, file_path,
               source_blob_id, source_sha256, canonical_sha256,
               downloaded_at, imported_at, pillar_count, criterion_count,
               triggered_by_user_id, import_warnings_json,
               validation_report_json, payload_json
        FROM official_referential_history
        WHERE tenant_id = ?
        ORDER BY imported_at DESC, history_id DESC
        LIMIT ?
        """,
        (tenant_id, max(1, min(int(limit), 200))),
    ).fetchall()
    return [
        summary
        for summary in (_official_referential_summary_from_row(row) for row in rows)
        if summary is not None
    ]


def _store_official_referential(
    con: Any,
    tenant_id: str,
    referential: dict[str, Any],
    remote: dict[str, Any],
    report: dict[str, Any],
    raw_xlsx: bytes,
    triggered_by_user_id: str = "",
) -> dict[str, Any]:
    referential_id = str(referential.get("id") or "adri-irn")
    source = referential.get("source") if isinstance(referential.get("source"), dict) else {}
    source_sha256 = hashlib.sha256(raw_xlsx).hexdigest()
    canonical_sha256 = _json_sha256(referential)
    downloaded_at = _utc_now().isoformat()
    imported_at = str(referential.get("importedAt") or downloaded_at)
    pillars = referential.get("pillars") if isinstance(referential.get("pillars"), list) else []
    criteria = referential.get("criteria") if isinstance(referential.get("criteria"), list) else []

    output_dir = OFFICIAL_REFERENTIAL_DIR / _safe_segment(tenant_id, "default") / _adri_safe_filename(referential_id)
    output_dir.mkdir(parents=True, exist_ok=True)
    raw_path = output_dir / str(remote.get("fileName") or "Questionnaire_IRN.xlsx")
    json_path = output_dir / f"{_adri_safe_filename(referential_id)}.json"
    report_path = output_dir / f"{_adri_safe_filename(referential_id)}_validation.json"
    raw_path.write_bytes(raw_xlsx)
    json_path.write_text(_pretty_json(referential) + "\n", encoding="utf-8")
    report_path.write_text(_pretty_json(report) + "\n", encoding="utf-8")

    source_url = str(source.get("url") or OFFICIAL_ADRI_SOURCE_URL)
    project_path = str(source.get("projectPath") or OFFICIAL_ADRI_PROJECT_PATH)
    default_branch = str(source.get("defaultBranch") or remote.get("defaultBranch") or OFFICIAL_ADRI_DEFAULT_BRANCH)
    file_path = str(source.get("filePath") or remote.get("filePath") or "")
    source_blob_id = str(source.get("blobId") or remote.get("blobId") or "")
    import_warnings_json = _canonical_json(
        referential.get("importWarnings") if isinstance(referential.get("importWarnings"), list) else []
    )
    validation_report_json = _canonical_json(report)
    payload_json = _canonical_json(referential)
    history_id = _official_referential_history_id(referential_id, canonical_sha256, downloaded_at)

    con.execute("UPDATE official_referentials SET active = 0 WHERE tenant_id = ?", (tenant_id,))
    con.execute("UPDATE official_referential_history SET active = 0 WHERE tenant_id = ?", (tenant_id,))
    con.execute(
        """
        INSERT INTO official_referentials(
            tenant_id, referential_id, version, active, source_url, project_path,
            default_branch, file_path, source_blob_id, source_sha256,
            canonical_sha256, downloaded_at, imported_at, pillar_count,
            criterion_count, import_warnings_json, validation_report_json, payload_json
        ) VALUES (?, ?, ?, 1, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(tenant_id, referential_id) DO UPDATE SET
            version = excluded.version,
            active = 1,
            source_url = excluded.source_url,
            project_path = excluded.project_path,
            default_branch = excluded.default_branch,
            file_path = excluded.file_path,
            source_blob_id = excluded.source_blob_id,
            source_sha256 = excluded.source_sha256,
            canonical_sha256 = excluded.canonical_sha256,
            downloaded_at = excluded.downloaded_at,
            imported_at = excluded.imported_at,
            pillar_count = excluded.pillar_count,
            criterion_count = excluded.criterion_count,
            import_warnings_json = excluded.import_warnings_json,
            validation_report_json = excluded.validation_report_json,
            payload_json = excluded.payload_json
        """,
        (
            tenant_id,
            referential_id,
            str(referential.get("version") or "unknown"),
            source_url,
            project_path,
            default_branch,
            file_path,
            source_blob_id,
            source_sha256,
            canonical_sha256,
            downloaded_at,
            imported_at,
            len(pillars),
            len(criteria),
            import_warnings_json,
            validation_report_json,
            payload_json,
        ),
    )
    con.execute(
        """
        INSERT INTO official_referential_history(
            tenant_id, history_id, referential_id, version, active,
            source_url, project_path, default_branch, file_path,
            source_blob_id, source_sha256, canonical_sha256,
            downloaded_at, imported_at, pillar_count, criterion_count,
            triggered_by_user_id, import_warnings_json,
            validation_report_json, payload_json
        ) VALUES (?, ?, ?, ?, 1, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(tenant_id, history_id) DO UPDATE SET
            active = 1,
            triggered_by_user_id = excluded.triggered_by_user_id,
            import_warnings_json = excluded.import_warnings_json,
            validation_report_json = excluded.validation_report_json,
            payload_json = excluded.payload_json
        """,
        (
            tenant_id,
            history_id,
            referential_id,
            str(referential.get("version") or "unknown"),
            source_url,
            project_path,
            default_branch,
            file_path,
            source_blob_id,
            source_sha256,
            canonical_sha256,
            downloaded_at,
            imported_at,
            len(pillars),
            len(criteria),
            str(triggered_by_user_id or "")[:120],
            import_warnings_json,
            validation_report_json,
            payload_json,
        ),
    )
    return {
        "historyId": history_id,
        "referentialId": referential_id,
        "version": referential.get("version") or "unknown",
        "sourceSha256": source_sha256,
        "canonicalSha256": canonical_sha256,
        "downloadedAt": downloaded_at,
        "importedAt": imported_at,
        "pillarCount": len(pillars),
        "criterionCount": len(criteria),
        "validationStatus": report.get("status"),
        "storedFiles": {
            "xlsx": str(raw_path),
            "json": str(json_path),
            "validation": str(report_path),
        },
    }


def _role_normalize(value: Any) -> str:
    role = str(value or "reader").strip().lower()
    aliases = {
        "admin": "administrator",
        "administrateur": "administrator",
        "administrator": "administrator",
        "pilote": "campaign_manager",
        "pilote_irn": "campaign_manager",
        "campaignmanager": "campaign_manager",
        "campaign_manager": "campaign_manager",
        "evaluateur": "evaluator",
        "évaluateur": "evaluator",
        "evaluator": "evaluator",
        "validateur": "reviewer",
        "validator": "reviewer",
        "reviewer": "reviewer",
        "lecteur": "reader",
        "reader": "reader",
    }
    return aliases.get(role, "reader")


def _sanitize_user(raw_user: Any) -> dict[str, Any] | None:
    if not isinstance(raw_user, dict):
        return None

    user_id = str(raw_user.get("id") or "").strip()
    if not user_id:
        return None

    created_at = str(raw_user.get("createdAt") or _utc_now().isoformat())
    updated_at = str(raw_user.get("updatedAt") or created_at)
    active = raw_user.get("active")

    return {
        "id": user_id,
        "firstName": str(raw_user.get("firstName") or "").strip(),
        "lastName": str(raw_user.get("lastName") or "").strip(),
        "email": str(raw_user.get("email") or "").strip().lower(),
        "role": _role_normalize(raw_user.get("role") or raw_user.get("roleLabel")),
        "active": active if isinstance(active, bool) else True,
        "createdAt": created_at,
        "updatedAt": updated_at,
    }


def _sort_users(users: list[dict[str, Any]]) -> list[dict[str, Any]]:
    def key(user: dict[str, Any]) -> tuple[int, str]:
        active_weight = 0 if user.get("active") is True else 1
        display = " ".join(
            part
            for part in [
                str(user.get("firstName") or "").strip(),
                str(user.get("lastName") or "").strip(),
                str(user.get("email") or "").strip(),
                str(user.get("id") or "").strip(),
            ]
            if part
        ).lower()
        return (active_weight, display)

    return sorted(users, key=key)


def _row_to_user(row: Any) -> dict[str, Any]:
    payload = _parse_json(row["payload_json"], {})
    if not isinstance(payload, dict):
        payload = {}
    payload.update(
        {
            "tenantId": row["tenant_id"],
            "tenantDisplayName": str(_row_value(row, "tenant_display_name", "") or "").strip() or "Espace de travail",
            "id": row["user_id"],
            "firstName": row["first_name"],
            "lastName": row["last_name"],
            "email": row["email"],
            "role": row["role"],
            "active": bool(row["active"]),
            "createdAt": row["created_at"],
            "updatedAt": row["updated_at"],
        }
    )
    return payload


def _load_central_users(con: Any, tenant_id: str) -> list[dict[str, Any]]:
    rows = con.execute(
        """
        SELECT u.tenant_id, COALESCE(NULLIF(t.display_name, ''), 'Espace de travail') AS tenant_display_name,
               u.user_id, u.first_name, u.last_name, u.email, u.role,
               u.active, u.created_at, u.updated_at, u.payload_json
        FROM users u
        LEFT JOIN tenants t ON t.id = u.tenant_id
        WHERE u.tenant_id = ?
        """,
        (tenant_id,),
    ).fetchall()
    tenant_label = _tenant_display_name(con, tenant_id)
    users = [_row_to_user(row) for row in rows]
    for user in users:
        user["tenantDisplayName"] = tenant_label
    return _sort_users(users)


def _save_user(con: Any, tenant_id: str, user: dict[str, Any]) -> None:
    con.execute(
        """
        INSERT INTO users(
            tenant_id, user_id, first_name, last_name, email, role,
            active, created_at, updated_at, payload_json
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(tenant_id, user_id) DO UPDATE SET
            first_name = excluded.first_name,
            last_name = excluded.last_name,
            email = excluded.email,
            role = excluded.role,
            active = excluded.active,
            updated_at = excluded.updated_at,
            payload_json = excluded.payload_json
        """,
        (
            tenant_id,
            user["id"],
            user["firstName"],
            user["lastName"],
            user["email"],
            user["role"],
            1 if user["active"] else 0,
            user["createdAt"],
            user["updatedAt"],
            _canonical_json(user),
        ),
    )


def _ensure_user_credentials(con: Any, tenant_id: str, users: list[dict[str, Any]]) -> None:
    for user in users:
        user_id = str(user.get("id") or "").strip()
        if not user_id:
            continue
        exists = con.execute(
            "SELECT 1 FROM user_credentials WHERE tenant_id = ? AND user_id = ?",
            (tenant_id, user_id),
        ).fetchone()
        if exists:
            continue
        salt = uuid.uuid4().hex
        con.execute(
            """
            INSERT INTO user_credentials(
                tenant_id, user_id, algorithm, iterations, salt, pin_hash,
                requires_change, updated_at
            ) VALUES (?, ?, 'pbkdf2_sha256', ?, ?, ?, 1, ?)
            """,
            (tenant_id, user_id, PIN_ITERATIONS, salt, _pin_hash(PIN_DEFAULT, salt, PIN_ITERATIONS), _utc_now().isoformat()),
        )


def _merge_central_users(con: Any, tenant_id: str, raw_users: Any) -> int:
    if not isinstance(raw_users, list):
        return 0

    current = {user["id"]: user for user in _load_central_users(con, tenant_id)}
    changed = 0
    for raw_user in raw_users:
        user = _sanitize_user(raw_user)
        if not user:
            continue
        existing = current.get(user["id"])
        if existing is None or _parse_datetime(user.get("updatedAt")) >= _parse_datetime(existing.get("updatedAt")):
            _save_user(con, tenant_id, user)
            current[user["id"]] = user
            changed += 1

    if changed:
        _ensure_user_credentials(con, tenant_id, list(current.values()))
    return changed


def _set_user_pin(con: Any, tenant_id: str, user_id: str, pin: str, *, requires_change: bool) -> None:
    cleaned_pin = str(pin or "").strip()
    if len(cleaned_pin) < 4 or len(cleaned_pin) > 32:
        raise HTTPException(status_code=400, detail="Le code personnel doit contenir entre 4 et 32 caractères")

    user_exists = con.execute(
        "SELECT 1 FROM users WHERE tenant_id = ? AND user_id = ?",
        (tenant_id, user_id),
    ).fetchone()
    if not user_exists:
        raise HTTPException(status_code=404, detail="Unknown user")

    salt = uuid.uuid4().hex
    con.execute(
        """
        INSERT INTO user_credentials(
            tenant_id, user_id, algorithm, iterations, salt, pin_hash,
            requires_change, updated_at
        ) VALUES (?, ?, 'pbkdf2_sha256', ?, ?, ?, ?, ?)
        ON CONFLICT(tenant_id, user_id) DO UPDATE SET
            algorithm = excluded.algorithm,
            iterations = excluded.iterations,
            salt = excluded.salt,
            pin_hash = excluded.pin_hash,
            requires_change = excluded.requires_change,
            updated_at = excluded.updated_at
        """,
        (
            tenant_id,
            user_id,
            PIN_ITERATIONS,
            salt,
            _pin_hash(cleaned_pin, salt, PIN_ITERATIONS),
            1 if requires_change else 0,
            _utc_now().isoformat(),
        ),
    )


def _verify_user_pin(con: Any, tenant_id: str, user_id: str, pin: str) -> tuple[bool, bool]:
    credential = con.execute(
        """
        SELECT salt, pin_hash, iterations, requires_change
        FROM user_credentials
        WHERE tenant_id = ? AND user_id = ?
        """,
        (tenant_id, user_id),
    ).fetchone()
    if credential is None:
        users = _load_central_users(con, tenant_id)
        _ensure_user_credentials(con, tenant_id, users)
        credential = con.execute(
            """
            SELECT salt, pin_hash, iterations, requires_change
            FROM user_credentials
            WHERE tenant_id = ? AND user_id = ?
            """,
            (tenant_id, user_id),
        ).fetchone()
    if credential is None:
        return (False, False)

    salt = str(credential["salt"] or "")
    expected_hash = str(credential["pin_hash"] or "")
    iterations = int(credential["iterations"] or PIN_ITERATIONS)
    if not salt or not expected_hash:
        return (False, False)

    provided_hash = _pin_hash(str(pin or "").strip(), salt, iterations)
    accepted = hmac.compare_digest(provided_hash, expected_hash)
    return (accepted, bool(credential["requires_change"]))


def _campaign_record(raw_campaign: dict[str, Any]) -> dict[str, Any]:
    nested = raw_campaign.get("campaign")
    if isinstance(nested, dict):
        return nested
    return raw_campaign


def _campaign_id(raw_campaign: dict[str, Any]) -> str | None:
    campaign = _campaign_record(raw_campaign)
    for source in (campaign, raw_campaign):
        for key in ("id", "campaignId"):
            value = str(source.get(key) or "").strip()
            if value:
                return value
    return None


def _campaign_updated_at(raw_campaign: dict[str, Any], payload: dict[str, Any], received_at: str) -> str:
    campaign = _campaign_record(raw_campaign)
    for source in (campaign, raw_campaign, payload):
        for key in ("updatedAt", "lastUpdatedAt", "modifiedAt", "createdAt", "generatedAt"):
            value = source.get(key)
            if value:
                return str(value)
    return received_at


def _looks_like_campaign_snapshot(item: Any) -> bool:
    if not isinstance(item, dict):
        return False
    if _campaign_id(item):
        return True
    nested = item.get("campaign")
    return isinstance(nested, dict) and _campaign_id(nested) is not None


def _extract_campaigns(payload: dict[str, Any]) -> list[dict[str, Any]]:
    direct = payload.get("campaigns")
    if isinstance(direct, list):
        return [item for item in direct if isinstance(item, dict)]

    found: list[dict[str, Any]] = []

    def walk(value: Any) -> None:
        if isinstance(value, dict):
            for key, child in value.items():
                if key == "campaigns" and isinstance(child, list):
                    for item in child:
                        if _looks_like_campaign_snapshot(item):
                            found.append(item)
                else:
                    walk(child)
        elif isinstance(value, list):
            for item in value:
                walk(item)

    walk(payload)
    return found


def _record_campaign_revisions(
    con: Any,
    tenant_id: str,
    server_sync_id: str,
    device_id: str,
    received_at: str,
    payload: dict[str, Any],
) -> dict[str, int]:
    campaigns = _extract_campaigns(payload)
    revision_count = 0
    conflict_count = 0
    skipped_without_id = 0
    deleted_count = 0
    received_campaign_ids: set[str] = set()

    for raw_campaign in campaigns:
        original_cid = _campaign_id(raw_campaign)
        if not original_cid:
            skipped_without_id += 1
            continue
        cid = _campaign_id_for_save(con, tenant_id, original_cid)
        if cid != original_cid:
            raw_campaign = _rewrite_json_ids(raw_campaign, {original_cid: cid})
        received_campaign_ids.add(cid)

        campaign_payload_sha256 = _json_sha256(raw_campaign)
        updated_at = _campaign_updated_at(raw_campaign, payload, received_at)
        existing = con.execute(
            """
            SELECT server_revision, payload_sha256, device_id, received_at
            FROM campaign_states
            WHERE tenant_id = ? AND campaign_id = ?
            """,
            (tenant_id, cid),
        ).fetchone()

        if existing and str(existing["payload_sha256"]) == campaign_payload_sha256:
            continue

        next_revision = 1
        conflict_detected = 0
        conflict_reason = None
        if existing:
            next_revision = int(existing["server_revision"] or 0) + 1
            if str(existing["device_id"] or "") != device_id:
                conflict_detected = 1
                conflict_reason = "last_write_wins_over_previous_device_revision"
            elif _parse_datetime(existing["received_at"]) > _parse_datetime(received_at):
                conflict_detected = 1
                conflict_reason = "older_received_at_after_newer_state"

        if conflict_detected:
            conflict_count += 1

        con.execute(
            """
            INSERT INTO campaign_revisions(
                tenant_id, campaign_id, server_revision, server_sync_id,
                device_id, updated_at, received_at, payload_sha256,
                payload_json, conflict_policy, conflict_detected, conflict_reason
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'last_write_wins', ?, ?)
            ON CONFLICT(tenant_id, campaign_id, server_revision) DO NOTHING
            """,
            (
                tenant_id,
                cid,
                next_revision,
                server_sync_id,
                device_id,
                updated_at,
                received_at,
                campaign_payload_sha256,
                _canonical_json(raw_campaign),
                conflict_detected,
                conflict_reason,
            ),
        )

        con.execute(
            """
            INSERT INTO campaign_states(
                tenant_id, campaign_id, server_revision, server_sync_id,
                device_id, updated_at, received_at, payload_sha256,
                payload_json, conflict_policy
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'last_write_wins')
            ON CONFLICT(tenant_id, campaign_id) DO UPDATE SET
                server_revision = excluded.server_revision,
                server_sync_id = excluded.server_sync_id,
                device_id = excluded.device_id,
                updated_at = excluded.updated_at,
                received_at = excluded.received_at,
                payload_sha256 = excluded.payload_sha256,
                payload_json = excluded.payload_json,
                conflict_policy = excluded.conflict_policy
            """,
            (
                tenant_id,
                cid,
                next_revision,
                server_sync_id,
                device_id,
                updated_at,
                received_at,
                campaign_payload_sha256,
                _canonical_json(raw_campaign),
            ),
        )

        con.execute(
            """
            INSERT INTO sync_events(
                tenant_id, event_type, server_sync_id, campaign_id,
                device_id, created_at, payload_json
            ) VALUES (?, 'campaign_revision_received', ?, ?, ?, ?, ?)
            """,
            (
                tenant_id,
                server_sync_id,
                cid,
                device_id,
                _utc_now().isoformat(),
                _canonical_json(
                    {
                        "campaignId": cid,
                        "serverRevision": next_revision,
                        "conflictDetected": conflict_detected == 1,
                        "conflictReason": conflict_reason,
                    }
                ),
            ),
        )
        revision_count += 1

    # OpenIRN clients publish the complete campaign set for a tenant.
    # A campaign absent from a new snapshot is therefore considered deleted.
    current_rows = con.execute(
        "SELECT campaign_id FROM campaign_states WHERE tenant_id = ?",
        (tenant_id,),
    ).fetchall()
    deleted_campaign_ids = [
        str(row["campaign_id"] or "")
        for row in current_rows
        if str(row["campaign_id"] or "") and str(row["campaign_id"] or "") not in received_campaign_ids
    ]

    for deleted_campaign_id in deleted_campaign_ids:
        con.execute(
            "DELETE FROM campaign_states WHERE tenant_id = ? AND campaign_id = ?",
            (tenant_id, deleted_campaign_id),
        )
        con.execute(
            "DELETE FROM campaign_revisions WHERE tenant_id = ? AND campaign_id = ?",
            (tenant_id, deleted_campaign_id),
        )
        con.execute(
            "DELETE FROM sync_events WHERE tenant_id = ? AND campaign_id = ?",
            (tenant_id, deleted_campaign_id),
        )
        con.execute(
            """
            INSERT INTO sync_events(
                tenant_id, event_type, server_sync_id, campaign_id,
                device_id, created_at, payload_json
            ) VALUES (?, 'campaign_deleted_by_snapshot_absence', ?, ?, ?, ?, ?)
            """,
            (
                tenant_id,
                server_sync_id,
                deleted_campaign_id,
                device_id,
                _utc_now().isoformat(),
                _canonical_json({"campaignId": deleted_campaign_id}),
            ),
        )
        deleted_count += 1

    return {
        "campaignCount": len(campaigns),
        "revisionCount": revision_count,
        "conflictCount": conflict_count,
        "deletedCount": deleted_count,
        "skippedWithoutId": skipped_without_id,
    }


def _public_snapshot_from_row(row: Any) -> dict[str, Any]:
    payload = _parse_json(row["payload_json"], {})
    return {
        "serverSyncId": row["server_sync_id"],
        "receivedAt": row["received_at"],
        "tenantId": row["tenant_id"],
        "deviceId": row["device_id"],
        "payloadSha256": row["payload_sha256"],
        "campaignCount": int(row["campaign_count"] or 0),
        "payload": payload if isinstance(payload, dict) else None,
    }


def _snapshot_summary_from_row(row: Any | None) -> dict[str, Any] | None:
    if row is None:
        return None
    public = _public_snapshot_from_row(row)
    public.pop("payload", None)
    return public


def _campaign_title_from_payload(payload: Any, fallback: str) -> str:
    if not isinstance(payload, dict):
        return fallback
    campaign = _campaign_record(payload)
    for key in ("name", "title", "label"):
        value = str(campaign.get(key) or "").strip()
        if value:
            return value
    return fallback


def _public_campaign_state_from_row(row: Any) -> dict[str, Any]:
    payload = _parse_json(row["payload_json"], {})
    campaign_id = str(row["campaign_id"] or "")
    return {
        "tenantId": row["tenant_id"],
        "tenantDisplayName": str(_row_value(row, "tenant_display_name", "") or "").strip() or "Espace de travail",
        "campaignId": campaign_id,
        "campaignName": _campaign_title_from_payload(payload, campaign_id),
        "serverRevision": int(row["server_revision"] or 0),
        "serverSyncId": row["server_sync_id"],
        "deviceId": row["device_id"],
        "updatedAt": row["updated_at"],
        "receivedAt": row["received_at"],
        "payloadSha256": row["payload_sha256"],
        "conflictPolicy": row["conflict_policy"],
    }


def _public_campaign_revision_from_row(row: Any, *, include_payload: bool = False) -> dict[str, Any]:
    payload = _parse_json(row["payload_json"], {})
    campaign_id = str(row["campaign_id"] or "")
    public = {
        "tenantId": row["tenant_id"],
        "tenantDisplayName": str(_row_value(row, "tenant_display_name", "") or "").strip() or "Espace de travail",
        "campaignId": campaign_id,
        "campaignName": _campaign_title_from_payload(payload, campaign_id),
        "serverRevision": int(row["server_revision"] or 0),
        "serverSyncId": row["server_sync_id"],
        "deviceId": row["device_id"],
        "updatedAt": row["updated_at"],
        "receivedAt": row["received_at"],
        "payloadSha256": row["payload_sha256"],
        "conflictPolicy": row["conflict_policy"],
        "conflictDetected": bool(row["conflict_detected"]),
        "conflictReason": row["conflict_reason"],
    }
    if include_payload:
        public["payload"] = payload if isinstance(payload, dict) else None
    return public


def _file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _table_counts(con: Any) -> dict[str, int | None]:
    counts: dict[str, int | None] = {}
    for table in [
        "tenants",
        "users",
        "sync_snapshots",
        "campaign_states",
        "campaign_revisions",
        "authorized_devices",
        "device_enrollment_requests",
        "device_enrollment_codes",
        "device_audit_log",
        "official_referentials",
        "official_referential_history",
        "sync_events",
        "backup_audit_log",
    ]:
        try:
            counts[table] = int(con.execute(f"select count(*) from {table}").fetchone()[0])
        except DbError:
            counts[table] = None
    return counts


def _write_private_text(path: Path, content: str) -> None:
    path.write_text(content, encoding="utf-8")
    _chmod_private(path, 0o600)


def _backup_metadata_from_file(path: Path) -> dict[str, Any]:
    meta_path = path.with_suffix(path.suffix + ".json")
    metadata = _parse_json(meta_path.read_text(encoding="utf-8") if meta_path.exists() else None, {})
    if not isinstance(metadata, dict):
        metadata = {}

    created_at = metadata.get("createdAt")
    if not created_at:
        created_at = datetime.fromtimestamp(path.stat().st_mtime, timezone.utc).isoformat()

    sha256 = str(metadata.get("sha256") or "").strip()
    if not sha256:
        sha_path = path.with_suffix(path.suffix + ".sha256")
        if sha_path.exists():
            sha_parts = sha_path.read_text(encoding="utf-8").split()
            sha256 = sha_parts[0] if sha_parts else ""

    signature_status = _backup_signature_status(metadata) if metadata else "unsigned"
    return {
        "name": path.name,
        "path": str(path),
        "createdAt": created_at,
        "sizeBytes": path.stat().st_size,
        "sha256": sha256,
        "integrityCheck": str(metadata.get("integrityCheck") or "unknown"),
        "reason": str(metadata.get("reason") or ""),
        "automatic": bool(metadata.get("automatic") is True),
        "triggeredByUserId": str(metadata.get("triggeredByUserId") or ""),
        "signatureStatus": signature_status,
        "signed": signature_status == "valid",
        "counts": metadata.get("counts") if isinstance(metadata.get("counts"), dict) else {},
    }


def _backup_files() -> list[Path]:
    if not BACKUP_DIR.exists():
        return []
    candidates = list(BACKUP_DIR.glob("openirn-*.mariadb.sql"))
    return sorted(candidates, key=lambda item: item.stat().st_mtime, reverse=True)


def _list_backups(limit: int = 10) -> list[dict[str, Any]]:
    backups = _backup_files()
    return [_backup_metadata_from_file(path) for path in backups[: max(1, min(limit, 100))]]


def _backup_audit_events(con: Any, tenant_id: str, limit: int = 20) -> list[dict[str, Any]]:
    rows = con.execute(
        """
        SELECT backup_name, event_type, reason, triggered_by_user_id,
               created_at, sha256, size_bytes, payload_json
        FROM backup_audit_log
        WHERE tenant_id = ?
        ORDER BY created_at DESC, id DESC
        LIMIT ?
        """,
        (tenant_id, max(1, min(limit, 100))),
    ).fetchall()
    events: list[dict[str, Any]] = []
    for row in rows:
        events.append(
            {
                "backupName": row["backup_name"],
                "eventType": row["event_type"],
                "reason": row["reason"],
                "triggeredByUserId": row["triggered_by_user_id"],
                "createdAt": row["created_at"],
                "sha256": row["sha256"],
                "sizeBytes": int(row["size_bytes"] or 0),
                "payload": _parse_json(row["payload_json"], {}),
            }
        )
    return events


def _record_backup_audit(
    con: Any,
    tenant_id: str,
    event_type: str,
    *,
    backup_name: str = "",
    reason: str = "",
    triggered_by_user_id: str | None = None,
    sha256: str = "",
    size_bytes: int = 0,
    payload: dict[str, Any] | None = None,
) -> None:
    con.execute(
        """
        INSERT INTO backup_audit_log(
            tenant_id, backup_name, event_type, reason, triggered_by_user_id,
            created_at, sha256, size_bytes, payload_json
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            tenant_id,
            backup_name[:240],
            event_type[:120],
            reason[:160],
            (triggered_by_user_id or "")[:160],
            _utc_now().isoformat(),
            sha256,
            int(size_bytes or 0),
            _canonical_json(payload or {}),
        ),
    )


def _last_protective_backup_at(con: Any, tenant_id: str, reason: str) -> datetime | None:
    row = con.execute(
        """
        SELECT created_at
        FROM backup_audit_log
        WHERE tenant_id = ?
          AND event_type = 'backup.created'
          AND reason = ?
        ORDER BY created_at DESC, id DESC
        LIMIT 1
        """,
        (tenant_id, reason),
    ).fetchone()
    if row is None:
        return None
    return _parse_datetime(row["created_at"])


def _cleanup_old_backups(protected_names: set[str] | None = None) -> list[str]:
    if BACKUP_KEEP <= 0 or not BACKUP_DIR.exists():
        return []
    protected_names = protected_names or set()
    backups = _backup_files()
    removed: list[str] = []
    for old in backups[BACKUP_KEEP:]:
        if old.name in protected_names:
            continue
        for companion in [old, old.with_suffix(old.suffix + ".sha256"), old.with_suffix(old.suffix + ".json")]:
            if companion.exists():
                companion.unlink()
        removed.append(old.name)
    return removed


def _mariadb_dump_binary() -> str:
    if MARIADB_DUMP_BIN:
        return MARIADB_DUMP_BIN
    return shutil.which("mariadb-dump") or shutil.which("mysqldump") or ""


def _create_mariadb_backup(
    triggered_by_user_id: str | None = None,
    protected_names: set[str] | None = None,
    *,
    tenant_id: str = "default",
    reason: str = "manual",
    automatic: bool = False,
) -> dict[str, Any]:
    config = _parse_mysql_url()
    dump_binary = _mariadb_dump_binary()
    if not dump_binary:
        raise HTTPException(status_code=503, detail="mariadb-dump/mysqldump is not available on the server")

    _ensure_private_backup_dir()
    stamp = _utc_now().strftime("%Y%m%dT%H%M%SZ")
    backup_path = BACKUP_DIR / f"openirn-{stamp}.mariadb.sql"
    suffix = 1
    while backup_path.exists():
        backup_path = BACKUP_DIR / f"openirn-{stamp}-{suffix}.mariadb.sql"
        suffix += 1

    defaults_path = BACKUP_DIR / f".openirn-mariadb-client-{uuid.uuid4().hex}.cnf"
    defaults_content = "\n".join(
        [
            "[client]",
            f"host={config['host']}",
            f"port={config['port']}",
            f"user={config['user']}",
            f"password={config['password']}",
            f"default-character-set={config.get('charset') or 'utf8mb4'}",
            "",
        ]
    )
    _write_private_text(defaults_path, defaults_content)
    cmd = [
        dump_binary,
        f"--defaults-extra-file={defaults_path}",
        "--single-transaction",
        "--quick",
        "--skip-comments",
        "--hex-blob",
        config["database"],
    ]
    try:
        with backup_path.open("wb") as output:
            proc = subprocess.run(cmd, stdout=output, stderr=subprocess.PIPE, check=False)
        if proc.returncode != 0:
            stderr = proc.stderr.decode("utf-8", errors="replace") if isinstance(proc.stderr, bytes) else str(proc.stderr or "")
            if backup_path.exists():
                backup_path.unlink()
            raise HTTPException(status_code=500, detail=f"MariaDB dump failed: {stderr[-600:]}")
    finally:
        try:
            defaults_path.unlink()
        except OSError:
            pass

    _chmod_private(backup_path, 0o600)
    digest = _file_sha256(backup_path)
    sha_path = backup_path.with_suffix(backup_path.suffix + ".sha256")
    _write_private_text(sha_path, f"{digest}  {backup_path.name}\n")

    with _db() as con:
        counts = _table_counts(con)
        con.commit()

    metadata = {
        "type": "openirn.mariadbDumpBackup",
        "formatVersion": 1,
        "backend": "mariadb",
        "createdAt": _utc_now().isoformat(),
        "sourceDb": _mariadb_target_label(),
        "backupDb": str(backup_path),
        "backupName": backup_path.name,
        "tenantId": tenant_id,
        "reason": reason,
        "automatic": automatic,
        "triggeredByUserId": triggered_by_user_id or "",
        "sha256": digest,
        "sizeBytes": backup_path.stat().st_size,
        "integrityCheck": "logical_dump_created",
        "retentionKeep": BACKUP_KEEP,
        "counts": counts,
    }
    signature = _backup_metadata_signature(metadata)
    if signature:
        metadata["signatureAlgorithm"] = "hmac-sha256-canonical-json-v1"
        metadata["signature"] = signature
        metadata["signatureStatus"] = "valid"
    else:
        metadata["signatureAlgorithm"] = ""
        metadata["signature"] = ""
        metadata["signatureStatus"] = "unsigned"

    meta_path = backup_path.with_suffix(backup_path.suffix + ".json")
    _write_private_text(meta_path, _pretty_json(metadata) + "\n")
    removed = _cleanup_old_backups(protected_names=protected_names)

    with _db() as con:
        with con:
            _ensure_tenant(con, tenant_id)
            _record_backup_audit(
                con,
                tenant_id,
                "backup.created",
                backup_name=backup_path.name,
                reason=reason,
                triggered_by_user_id=triggered_by_user_id,
                sha256=digest,
                size_bytes=backup_path.stat().st_size,
                payload={
                    "backend": "mariadb",
                    "automatic": automatic,
                    "signatureStatus": metadata["signatureStatus"],
                    "removedOldBackups": removed,
                },
            )

    return {**metadata, "name": backup_path.name, "removedOldBackups": removed}


def _create_database_backup(
    triggered_by_user_id: str | None = None,
    protected_names: set[str] | None = None,
    *,
    tenant_id: str = "default",
    reason: str = "manual",
    automatic: bool = False,
) -> dict[str, Any]:
    return _create_mariadb_backup(
        tenant_id=tenant_id,
        triggered_by_user_id=triggered_by_user_id,
        protected_names=protected_names,
        reason=reason,
        automatic=automatic,
    )


def _create_protective_backup(
    tenant_id: str,
    *,
    reason: str,
    triggered_by_user_id: str | None = None,
    context: dict[str, Any] | None = None,
) -> dict[str, Any]:
    if not BACKUP_PROTECTIVE_ENABLED:
        return {"status": "disabled", "reason": reason}

    min_interval = max(0, BACKUP_PROTECTIVE_MIN_INTERVAL_MINUTES)
    with _db() as con:
        _ensure_tenant(con, tenant_id)
        last_created_at = _last_protective_backup_at(con, tenant_id, reason)
        con.commit()

    if last_created_at is not None and min_interval > 0:
        elapsed = _utc_now() - last_created_at
        if elapsed < timedelta(minutes=min_interval):
            return {
                "status": "skipped_recent",
                "reason": reason,
                "lastCreatedAt": last_created_at.isoformat(),
                "minIntervalMinutes": min_interval,
            }

    backup = _create_database_backup(
        tenant_id=tenant_id,
        triggered_by_user_id=triggered_by_user_id,
        reason=reason,
        automatic=True,
    )
    if context:
        with _db() as con:
            with con:
                _ensure_tenant(con, tenant_id)
                _record_backup_audit(
                    con,
                    tenant_id,
                    "backup.protective_context",
                    backup_name=str(backup.get("name") or ""),
                    reason=reason,
                    triggered_by_user_id=triggered_by_user_id,
                    sha256=str(backup.get("sha256") or ""),
                    size_bytes=int(backup.get("sizeBytes") or 0),
                    payload=context,
                )
    return {"status": "created", "backup": backup, "reason": reason}


def _backup_path_from_name(backup_name: str) -> Path:
    raw_name = str(backup_name or "").strip()
    safe_name = Path(raw_name).name
    valid_extension = safe_name.endswith(".mariadb.sql")
    if raw_name != safe_name or not safe_name.startswith("openirn-") or not valid_extension:
        raise HTTPException(status_code=400, detail="Invalid backup name")

    backup_dir = BACKUP_DIR.resolve()
    backup_path = (BACKUP_DIR / safe_name).resolve()
    try:
        backup_path.relative_to(backup_dir)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid backup path") from exc

    if not backup_path.exists() or not backup_path.is_file():
        raise HTTPException(status_code=404, detail="Backup not found")
    return backup_path


def _verify_backup_file(backup_path: Path) -> dict[str, Any]:
    digest = _file_sha256(backup_path)
    sha_path = backup_path.with_suffix(backup_path.suffix + ".sha256")
    if sha_path.exists():
        expected_parts = sha_path.read_text(encoding="utf-8").split()
        expected = expected_parts[0].strip() if expected_parts else ""
        if expected and not hmac.compare_digest(digest, expected):
            raise HTTPException(status_code=500, detail="Backup SHA-256 checksum mismatch")

    meta_path = backup_path.with_suffix(backup_path.suffix + ".json")
    metadata = _parse_json(meta_path.read_text(encoding="utf-8") if meta_path.exists() else None, {})
    if not isinstance(metadata, dict):
        metadata = {}
    signature_status = _backup_signature_status(metadata) if metadata else "unsigned"
    if signature_status == "invalid":
        raise HTTPException(status_code=500, detail="Backup manifest HMAC signature mismatch")
    return {"sha256": digest, "integrityCheck": "logical_dump", "signatureStatus": signature_status}


def _delete_database_backup(
    backup_name: str,
    *,
    tenant_id: str = "default",
    triggered_by_user_id: str | None = None,
) -> dict[str, Any]:
    backup_path = _backup_path_from_name(backup_name)
    metadata = _backup_metadata_from_file(backup_path)
    deleted: list[str] = []
    for companion in [
        backup_path,
        backup_path.with_suffix(backup_path.suffix + ".sha256"),
        backup_path.with_suffix(backup_path.suffix + ".json"),
    ]:
        if companion.exists():
            companion.unlink()
            deleted.append(companion.name)
    with _db() as con:
        with con:
            _ensure_tenant(con, tenant_id)
            _record_backup_audit(
                con,
                tenant_id,
                "backup.deleted",
                backup_name=backup_path.name,
                reason="manual_delete",
                triggered_by_user_id=triggered_by_user_id,
                sha256=str(metadata.get("sha256") or ""),
                size_bytes=int(metadata.get("sizeBytes") or 0),
                payload={"deletedFiles": deleted},
            )
    return {
        "status": "ok",
        "type": "openirn.databaseBackupDeleted",
        "deletedAt": _utc_now().isoformat(),
        "backupName": backup_path.name,
        "deletedFiles": deleted,
    }


def _database_status_payload(counts: dict[str, int | None]) -> dict[str, Any]:
    version = "unknown"
    try:
        with _db() as con:
            row = con.execute("SELECT VERSION()").fetchone()
            version = str(row[0]) if row else "unknown"
            con.commit()
    except Exception:
        version = "unavailable"
    return {
        "backend": "mariadb",
        "target": _mariadb_target_label(),
        "serverVersion": version,
        "charset": "utf8mb4",
        "counts": counts,
    }


def _maintenance_status(limit: int = 10, tenant_id: str = "default") -> dict[str, Any]:
    with _db() as con:
        _ensure_tenant(con, tenant_id)
        counts = _table_counts(con)
        audit_events = _backup_audit_events(con, tenant_id, limit=limit)
        con.commit()

    backups = _list_backups(limit=limit)
    signature_secret_configured = bool(_backup_signing_secret())
    unsigned_count = sum(1 for backup in backups if backup.get("signatureStatus") != "valid")
    return {
        "status": "ok",
        "type": "openirn.maintenanceStatus",
        "application": "OpenIRN API",
        "version": APP_VERSION,
        "serverTime": _utc_now().isoformat(),
        "database": _database_status_payload(counts),
        "backup": {
            "directory": str(BACKUP_DIR),
            "keep": BACKUP_KEEP,
            "count": len(_backup_files()),
            "latest": backups[0] if backups else None,
            "backups": backups,
            "auditEvents": audit_events,
            "security": {
                "autoEnabled": BACKUP_AUTO_ENABLED,
                "protectiveEnabled": BACKUP_PROTECTIVE_ENABLED,
                "protectiveMinIntervalMinutes": BACKUP_PROTECTIVE_MIN_INTERVAL_MINUTES,
                "privateDirectory": True,
                "fileMode": "0600",
                "directoryMode": "0700",
                "signatureAlgorithm": "hmac-sha256-canonical-json-v1",
                "signatureSecretConfigured": signature_secret_configured,
                "unsignedVisibleBackups": unsigned_count,
            },
        },
    }



@app.get("/tenants")
def tenants(
    request: Request,
    tenantId: str = Query(default=DEFAULT_TENANT_ID, min_length=1, max_length=80),
) -> dict[str, Any]:
    requester_tenant_id = _resolve_tenant_id_for_request(tenantId, DEFAULT_TENANT_ID)
    # Tenant discovery is intentionally public: the Flutter client must be able
    # to start without a selected tenant or session, then let the user choose
    # the workspace before checking terminal enrollment.
    with _db() as con:
        _ensure_tenant(con, requester_tenant_id)
        _sync_solution_administrators_to_all_tenants(con)
        items = _list_tenants(con)
        con.commit()
    solution_administrator = _request_has_solution_admin_authorization(request)
    return {
        "status": "ok",
        "type": "openirn.tenants",
        "application": "OpenIRN API",
        "version": APP_VERSION,
        "tenantId": requester_tenant_id,
        "defaultTenantId": DEFAULT_TENANT_ID,
        "solutionAdminTenantId": SOLUTION_ADMIN_TENANT_ID,
        "solutionAdministrator": solution_administrator,
        "tenantCount": len(items),
        "tenants": items,
        "serverTime": _utc_now().isoformat(),
    }


@app.post("/tenants")
async def tenant_create(request: Request) -> dict[str, Any]:
    try:
        payload = await request.json()
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail="Invalid JSON payload") from exc
    if not isinstance(payload, dict):
        raise HTTPException(status_code=400, detail="Payload must be a JSON object")

    requester_tenant_id = _resolve_tenant_id_for_request(payload.get("requesterTenantId") or payload.get("adminTenantId") or payload.get("sourceTenantId"), DEFAULT_TENANT_ID)
    auth_context = _require_admin_authorization(request, requester_tenant_id)
    tenant_id = _tenant_id_for_creation(payload.get("tenantId"))

    display_name = str(payload.get("displayName") or payload.get("name") or "Nouvel espace de travail").strip()[:160]
    description = str(payload.get("description") or "").strip()[:500]
    pilot_payload = payload.get("pilot")
    if not isinstance(pilot_payload, dict):
        raise HTTPException(status_code=400, detail="Missing pilot user")
    pilot_email = str(pilot_payload.get("email") or "").strip().lower()
    pilot_first_name = str(pilot_payload.get("firstName") or "").strip()
    pilot_last_name = str(pilot_payload.get("lastName") or "").strip()
    pilot_pin = str(pilot_payload.get("pin") or payload.get("pilotPin") or "").strip()
    if not pilot_email:
        raise HTTPException(status_code=400, detail="Pilot email is required")
    if len(pilot_pin) < 4 or len(pilot_pin) > 32:
        raise HTTPException(status_code=400, detail="Pilot Le code personnel doit contenir entre 4 et 32 caractères")

    now = _utc_now().isoformat()
    pilot_user = {
        "id": _normalize_uuid(pilot_payload.get("id")) or _new_uuid(),
        "firstName": pilot_first_name,
        "lastName": pilot_last_name,
        "email": pilot_email,
        "role": "campaign_manager",
        "active": True,
        "createdAt": now,
        "updatedAt": now,
    }
    device_id = str(auth_context.get("deviceId") or _request_device_id(request, payload) or "").strip()[:160]
    actor_user_id = str(auth_context.get("userId") or "server").strip()[:120]

    with _db() as con:
        with con:
            _ensure_tenant(con, DEFAULT_TENANT_ID)
            existing = con.execute("SELECT 1 FROM tenants WHERE id = ?", (tenant_id,)).fetchone()
            if existing is not None:
                raise HTTPException(status_code=409, detail="Cet espace de travail existe déjà")
            con.execute(
                """
                INSERT INTO tenants(id, created_at, updated_at, display_name, description, permanent)
                VALUES (?, ?, ?, ?, ?, 0)
                """,
                (tenant_id, now, now, display_name or tenant_id, description),
            )
            _copy_user_to_tenant(
                con,
                source_tenant_id=requester_tenant_id,
                target_tenant_id=tenant_id,
                user_id=actor_user_id,
            )
            _save_user(con, tenant_id, pilot_user)
            _set_user_pin(con, tenant_id, pilot_user["id"], pilot_pin, requires_change=False)
            _seed_official_referential_from_default(
                con,
                tenant_id,
                source_tenant_id=requester_tenant_id,
            )
            _record_device_audit(
                con,
                requester_tenant_id,
                "tenant.created",
                device_id=device_id,
                payload={
                    "tenantId": tenant_id,
                    "displayName": display_name or "Espace de travail",
                    "pilotUserId": pilot_user["id"],
                    "actorUserId": actor_user_id,
                    "deviceEnrollmentIsolation": "tenant-scoped",
                },
            )
            _sync_solution_administrators_to_all_tenants(con)
            items = _list_tenants(con)

    return {
        "status": "created",
        "type": "openirn.tenantCreated",
        "application": "OpenIRN API",
        "version": APP_VERSION,
        "tenantId": tenant_id,
        "requesterTenantId": requester_tenant_id,
        "defaultTenantId": DEFAULT_TENANT_ID,
        "solutionAdminTenantId": SOLUTION_ADMIN_TENANT_ID,
        "solutionAdministrator": _is_solution_admin_context(auth_context),
        "serverTime": _utc_now().isoformat(),
        "tenant": next((item for item in items if item.get("tenantId") == tenant_id), None),
        "pilot": {key: value for key, value in pilot_user.items() if key != "pin"},
        "tenants": items,
        "message": "Espace de travail créé avec un Pilote IRN initial.",
    }


@app.patch("/tenants/{tenant_id}")
async def tenant_update(tenant_id: str, request: Request) -> dict[str, Any]:
    try:
        payload = await request.json()
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail="Invalid JSON payload") from exc

    if not isinstance(payload, dict):
        raise HTTPException(status_code=400, detail="Payload must be a JSON object")

    target_tenant_id = _resolve_tenant_id_for_request(tenant_id, DEFAULT_TENANT_ID)
    auth_context = _require_admin_authorization(request, target_tenant_id)
    display_name = str(payload.get("displayName") or payload.get("name") or "").strip()
    if not display_name:
        raise HTTPException(status_code=400, detail="Le nom affiché de l’espace de travail est obligatoire")
    if len(display_name) > 160:
        raise HTTPException(status_code=400, detail="Le nom affiché ne doit pas dépasser 160 caractères")

    now = _utc_now().isoformat()
    actor_user_id = str(auth_context.get("userId") or "server").strip()[:120]
    device_id = str(auth_context.get("deviceId") or _request_device_id(request, payload) or "").strip()[:160]

    with _db() as con:
        with con:
            _ensure_tenant(con, DEFAULT_TENANT_ID)
            row = con.execute(
                "SELECT id, display_name FROM tenants WHERE id = ?",
                (target_tenant_id,),
            ).fetchone()
            if row is None:
                raise HTTPException(status_code=404, detail="Espace de travail introuvable")
            previous_display_name = str(row["display_name"] or target_tenant_id)
            con.execute(
                """
                UPDATE tenants
                SET display_name = ?, updated_at = ?
                WHERE id = ?
                """,
                (display_name, now, target_tenant_id),
            )
            _record_device_audit(
                con,
                target_tenant_id,
                "tenant.renamed",
                device_id=device_id,
                payload={
                    "actorUserId": actor_user_id,
                    "previousDisplayName": previous_display_name,
                    "displayName": display_name,
                },
            )
            _sync_solution_administrators_to_all_tenants(con)
            items = _list_tenants(con)

    return {
        "status": "accepted",
        "type": "openirn.tenantUpdated",
        "application": "OpenIRN API",
        "version": APP_VERSION,
        "tenantId": target_tenant_id,
        "defaultTenantId": DEFAULT_TENANT_ID,
        "solutionAdminTenantId": SOLUTION_ADMIN_TENANT_ID,
        "solutionAdministrator": _is_solution_admin_context(auth_context),
        "serverTime": _utc_now().isoformat(),
        "tenant": next((item for item in items if item.get("tenantId") == target_tenant_id), None),
        "tenants": items,
        "message": "Nom de l’espace de travail mis à jour.",
    }


@app.delete("/tenants/{tenant_id}")
def tenant_delete(tenant_id: str, request: Request) -> dict[str, Any]:
    target_tenant_id = _resolve_tenant_id_for_request(tenant_id, DEFAULT_TENANT_ID)
    auth_context = _require_admin_authorization(request, target_tenant_id)
    actor_user_id = str(auth_context.get("userId") or "server").strip()[:120]
    device_id = str(auth_context.get("deviceId") or "").strip()[:160]
    auth_tenant_id = str(auth_context.get("tenantId") or "").strip()

    with _db() as con:
        with con:
            _ensure_tenant(con, DEFAULT_TENANT_ID)
            default_tenant_id = _resolve_tenant_id(con, DEFAULT_TENANT_ID, DEFAULT_TENANT_ID)
            row = con.execute(
                "SELECT id, display_name, permanent FROM tenants WHERE id = ?",
                (target_tenant_id,),
            ).fetchone()
            if row is None:
                raise HTTPException(status_code=404, detail="Espace de travail introuvable")

            display_name = _tenant_display_name(con, target_tenant_id)
            is_permanent = bool(int(row["permanent"] or 0))
            if is_permanent or target_tenant_id == default_tenant_id:
                raise HTTPException(
                    status_code=403,
                    detail="L’espace de travail par défaut ne peut pas être supprimé",
                )

            counts = {
                "users": int(con.execute(
                    "SELECT COUNT(*) FROM users WHERE tenant_id = ?",
                    (target_tenant_id,),
                ).fetchone()[0]),
                "campaigns": int(con.execute(
                    "SELECT COUNT(*) FROM campaign_states WHERE tenant_id = ?",
                    (target_tenant_id,),
                ).fetchone()[0]),
                "devices": int(con.execute(
                    "SELECT COUNT(*) FROM authorized_devices WHERE tenant_id = ?",
                    (target_tenant_id,),
                ).fetchone()[0]),
                "enrollmentRequests": int(con.execute(
                    "SELECT COUNT(*) FROM device_enrollment_requests WHERE tenant_id = ?",
                    (target_tenant_id,),
                ).fetchone()[0]),
            }

            audit_tenant_id = auth_tenant_id if auth_tenant_id and auth_tenant_id != target_tenant_id else default_tenant_id
            if audit_tenant_id and audit_tenant_id != target_tenant_id:
                _record_device_audit(
                    con,
                    audit_tenant_id,
                    "tenant.deleted",
                    device_id=device_id,
                    payload={
                        "tenantId": target_tenant_id,
                        "displayName": display_name,
                        "actorUserId": actor_user_id,
                        "deletedCounts": counts,
                    },
                )

            con.execute(
                """
                DELETE FROM id_aliases
                WHERE (entity_type = 'tenant' AND new_id = ?)
                   OR scope_id = ?
                   OR new_id = ?
                """,
                (target_tenant_id, target_tenant_id, target_tenant_id),
            )
            con.execute("DELETE FROM tenants WHERE id = ?", (target_tenant_id,))
            _sync_solution_administrators_to_all_tenants(con)
            items = _list_tenants(con)

    return {
        "status": "deleted",
        "type": "openirn.tenantDeleted",
        "application": "OpenIRN API",
        "version": APP_VERSION,
        "tenantId": target_tenant_id,
        "tenantDisplayName": display_name,
        "defaultTenantId": DEFAULT_TENANT_ID,
        "solutionAdminTenantId": SOLUTION_ADMIN_TENANT_ID,
        "solutionAdministrator": _is_solution_admin_context(auth_context),
        "deletedCounts": counts,
        "tenants": items,
        "serverTime": _utc_now().isoformat(),
        "message": f"L’espace de travail « {display_name} » a été supprimé avec ses utilisateurs et ses campagnes.",
    }


@app.get("/devices")
def devices(
    request: Request,
    tenantId: str = Query(default="default", min_length=1, max_length=80),
    allTenants: bool = Query(default=False),
) -> dict[str, Any]:
    tenant_id = _resolve_tenant_id_for_request(tenantId, DEFAULT_TENANT_ID)
    auth_context = _require_campaign_manager_authorization(request, tenant_id)
    include_all_tenants = bool(allTenants and _is_administrator_context(auth_context))
    with _db() as con:
        _ensure_tenant(con, tenant_id)
        devices_list = _list_all_devices(con) if include_all_tenants else _list_devices(con, tenant_id)
        requests_list = _list_device_enrollment_requests(
            con,
            tenant_id,
            include_all_tenants=include_all_tenants,
        )
        tenant_display_name = _tenant_display_name(con, tenant_id)
        con.commit()
    return {
        "status": "ok",
        "type": "openirn.devices",
        "tenantId": tenant_id,
        "tenantDisplayName": tenant_display_name,
        "scope": "all_tenants" if include_all_tenants else "tenant",
        "deviceCount": len(devices_list),
        "enrollmentRequestCount": len(requests_list),
        "devices": devices_list,
        "enrollmentRequests": requests_list,
        "serverTime": _utc_now().isoformat(),
    }


@app.post("/devices/enrollment/request")
async def devices_enrollment_request(request: Request) -> dict[str, Any]:
    try:
        payload = await request.json()
    except json.JSONDecodeError:
        payload = {}
    if not isinstance(payload, dict):
        payload = {}

    tenant_id = _resolve_tenant_id_for_request(payload.get("tenantId"), DEFAULT_TENANT_ID)
    device_name = str(payload.get("deviceName") or "").strip()[:120] or "Terminal OpenIRN"
    platform = str(payload.get("platform") or "").strip()[:80]
    requested_device_id = _normalize_device_id(payload.get("deviceId"))
    requester_note = str(payload.get("note") or payload.get("requesterNote") or "").strip()[:500]
    request_id = f"enrollment_request_{uuid.uuid4().hex}"
    now = _utc_now().isoformat()

    with _db() as con:
        tenant_exists = con.execute("SELECT 1 FROM tenants WHERE id = ?", (tenant_id,)).fetchone()
        if tenant_exists is None:
            raise HTTPException(status_code=404, detail="Espace de travail introuvable")
        known_terminal = _terminal_identity(con, requested_device_id) if requested_device_id else None
        if known_terminal is not None:
            device_name = str(known_terminal.get("name") or "").strip()[:120] or device_name
            platform = str(known_terminal.get("platform") or "").strip()[:80] or platform
        con.execute(
            """
            INSERT INTO device_enrollment_requests(
                tenant_id, request_id, device_id, device_name, platform, requester_note,
                requester_ip, status, requested_at, decided_at, decided_by_user_id,
                decision_note, enrollment_id
            ) VALUES (?, ?, ?, ?, ?, ?, ?, 'pending', ?, NULL, NULL, '', NULL)
            """,
            (
                tenant_id,
                request_id,
                requested_device_id,
                device_name,
                platform,
                requester_note,
                _request_client_ip(request),
                now,
            ),
        )
        _record_device_audit(
            con,
            tenant_id,
            "enrollment_request.created",
            payload={
                "requestId": request_id,
                "deviceId": requested_device_id,
                "deviceName": device_name,
                "platform": platform,
                "knownTerminal": known_terminal is not None,
            },
        )
        con.commit()

    return {
        "status": "pending",
        "type": "openirn.deviceEnrollmentRequest",
        "tenantId": tenant_id,
        "requestId": request_id,
        "deviceId": requested_device_id,
        "deviceName": device_name,
        "platform": platform,
        "serverTime": _utc_now().isoformat(),
        "message": "Demande d’autorisation envoyée. Un Pilote IRN ou un administrateur doit maintenant la traiter.",
    }


@app.post("/devices/enrollment")
async def devices_enrollment(request: Request) -> dict[str, Any]:
    try:
        payload = await request.json()
    except json.JSONDecodeError:
        payload = {}
    if not isinstance(payload, dict):
        payload = {}

    tenant_id = _resolve_tenant_id_for_request(payload.get("tenantId"), DEFAULT_TENANT_ID)
    _require_campaign_manager_authorization(request, tenant_id)
    created_by_user_id = str(payload.get("createdByUserId") or "").strip()[:120]
    label = str(payload.get("label") or "").strip()[:120]
    allowed_expiration_minutes = {5, 10, 15}
    try:
        expires_in_minutes = int(payload.get("expiresInMinutes") or 10)
    except (TypeError, ValueError):
        expires_in_minutes = 10
    if expires_in_minutes not in allowed_expiration_minutes:
        expires_in_minutes = 10

    raw_code = _new_enrollment_code()
    display_code = _format_enrollment_code(raw_code)
    normalized_code = _normalize_enrollment_code(raw_code)
    enrollment_id = f"enrollment_{uuid.uuid4().hex}"
    now = _utc_now()
    expires_at = now + timedelta(minutes=expires_in_minutes)

    with _db() as con:
        _ensure_tenant(con, tenant_id)
        con.execute(
            """
            INSERT INTO device_enrollment_codes(
                tenant_id, enrollment_id, code_hash, created_by_user_id, label,
                expires_at, consumed_at, consumed_by_device_id, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, NULL, NULL, ?)
            """,
            (
                tenant_id,
                enrollment_id,
                _enrollment_code_hash(tenant_id, normalized_code),
                created_by_user_id,
                label,
                expires_at.isoformat(),
                now.isoformat(),
            ),
        )
        _record_device_audit(
            con,
            tenant_id,
            "enrollment.created",
            payload={
                "enrollmentId": enrollment_id,
                "createdByUserId": created_by_user_id,
                "label": label,
                "expiresAt": expires_at.isoformat(),
            },
        )
        con.commit()

    qr_payload = {
        "type": "openirn.deviceEnrollment",
        "tenantId": tenant_id,
        "code": display_code,
        "enrollmentId": enrollment_id,
        "expiresAt": expires_at.isoformat(),
    }
    return {
        "status": "ok",
        "type": "openirn.deviceEnrollment",
        "tenantId": tenant_id,
        "enrollmentId": enrollment_id,
        "code": display_code,
        "expiresAt": expires_at.isoformat(),
        "expiresInMinutes": expires_in_minutes,
        "qrPayload": qr_payload,
        "qrPayloadText": _canonical_json(qr_payload),
        "serverTime": _utc_now().isoformat(),
    }


@app.post("/devices/enrollment/requests/{request_id}/approve")
async def devices_enrollment_request_approve(request_id: str, request: Request) -> dict[str, Any]:
    try:
        payload = await request.json()
    except json.JSONDecodeError:
        payload = {}
    if not isinstance(payload, dict):
        payload = {}

    tenant_id = _resolve_tenant_id_for_request(payload.get("tenantId"), DEFAULT_TENANT_ID)
    auth_context = _require_campaign_manager_authorization(request, tenant_id)
    actor_user_id = str(auth_context.get("userId") or "server").strip()[:120]
    allowed_expiration_minutes = {5, 10, 15}
    try:
        expires_in_minutes = int(payload.get("expiresInMinutes") or 15)
    except (TypeError, ValueError):
        expires_in_minutes = 15
    if expires_in_minutes not in allowed_expiration_minutes:
        expires_in_minutes = 15

    raw_code = _new_enrollment_code()
    display_code = _format_enrollment_code(raw_code)
    normalized_code = _normalize_enrollment_code(raw_code)
    enrollment_id = f"enrollment_{uuid.uuid4().hex}"
    now = _utc_now()
    expires_at = now + timedelta(minutes=expires_in_minutes)

    with _db() as con:
        with con:
            row = con.execute(
                """
                SELECT tenant_id, request_id, device_name, platform, status
                FROM device_enrollment_requests
                WHERE tenant_id = ? AND request_id = ?
                """,
                (tenant_id, request_id),
            ).fetchone()
            if row is None:
                raise HTTPException(status_code=404, detail="Demande d’enrôlement introuvable")
            if str(row["status"] or "") != "pending":
                raise HTTPException(status_code=409, detail="Cette demande a déjà été traitée")
            label = f"Demande {request_id} — {row['device_name']}"[:120]
            con.execute(
                """
                INSERT INTO device_enrollment_codes(
                    tenant_id, enrollment_id, code_hash, created_by_user_id, label,
                    expires_at, consumed_at, consumed_by_device_id, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, NULL, NULL, ?)
                """,
                (
                    tenant_id,
                    enrollment_id,
                    _enrollment_code_hash(tenant_id, normalized_code),
                    actor_user_id,
                    label,
                    expires_at.isoformat(),
                    now.isoformat(),
                ),
            )
            con.execute(
                """
                UPDATE device_enrollment_requests
                SET status = 'approved', decided_at = ?, decided_by_user_id = ?,
                    decision_note = ?, enrollment_id = ?
                WHERE tenant_id = ? AND request_id = ?
                """,
                (
                    now.isoformat(),
                    actor_user_id,
                    str(payload.get("decisionNote") or "").strip()[:500],
                    enrollment_id,
                    tenant_id,
                    request_id,
                ),
            )
            _record_device_audit(
                con,
                tenant_id,
                "enrollment_request.approved",
                payload={
                    "requestId": request_id,
                    "enrollmentId": enrollment_id,
                    "actorUserId": actor_user_id,
                    "expiresAt": expires_at.isoformat(),
                },
            )
            requests_list = _list_device_enrollment_requests(con, tenant_id)
            devices_list = _list_devices(con, tenant_id)

    qr_payload = {
        "type": "openirn.deviceEnrollment",
        "tenantId": tenant_id,
        "code": display_code,
        "enrollmentId": enrollment_id,
        "expiresAt": expires_at.isoformat(),
    }
    return {
        "status": "approved",
        "type": "openirn.deviceEnrollmentRequestApproved",
        "tenantId": tenant_id,
        "requestId": request_id,
        "enrollmentId": enrollment_id,
        "code": display_code,
        "expiresAt": expires_at.isoformat(),
        "expiresInMinutes": expires_in_minutes,
        "qrPayload": qr_payload,
        "qrPayloadText": _canonical_json(qr_payload),
        "devices": devices_list,
        "enrollmentRequests": requests_list,
        "serverTime": _utc_now().isoformat(),
        "message": "Demande approuvée. Le code d’appairage peut être transmis au demandeur.",
    }


@app.post("/devices/enrollment/requests/{request_id}/reject")
async def devices_enrollment_request_reject(request_id: str, request: Request) -> dict[str, Any]:
    try:
        payload = await request.json()
    except json.JSONDecodeError:
        payload = {}
    if not isinstance(payload, dict):
        payload = {}

    tenant_id = _resolve_tenant_id_for_request(payload.get("tenantId"), DEFAULT_TENANT_ID)
    auth_context = _require_campaign_manager_authorization(request, tenant_id)
    actor_user_id = str(auth_context.get("userId") or "server").strip()[:120]
    now = _utc_now().isoformat()

    with _db() as con:
        with con:
            row = con.execute(
                """
                SELECT status FROM device_enrollment_requests
                WHERE tenant_id = ? AND request_id = ?
                """,
                (tenant_id, request_id),
            ).fetchone()
            if row is None:
                raise HTTPException(status_code=404, detail="Demande d’enrôlement introuvable")
            if str(row["status"] or "") != "pending":
                raise HTTPException(status_code=409, detail="Cette demande a déjà été traitée")
            con.execute(
                """
                UPDATE device_enrollment_requests
                SET status = 'rejected', decided_at = ?, decided_by_user_id = ?,
                    decision_note = ?
                WHERE tenant_id = ? AND request_id = ?
                """,
                (
                    now,
                    actor_user_id,
                    str(payload.get("decisionNote") or "").strip()[:500],
                    tenant_id,
                    request_id,
                ),
            )
            _record_device_audit(
                con,
                tenant_id,
                "enrollment_request.rejected",
                payload={"requestId": request_id, "actorUserId": actor_user_id},
            )
            requests_list = _list_device_enrollment_requests(con, tenant_id)
            devices_list = _list_devices(con, tenant_id)

    return {
        "status": "rejected",
        "type": "openirn.deviceEnrollmentRequestRejected",
        "tenantId": tenant_id,
        "requestId": request_id,
        "devices": devices_list,
        "enrollmentRequests": requests_list,
        "serverTime": _utc_now().isoformat(),
        "message": "Demande refusée.",
    }


@app.post("/devices/enrollment/consume")
async def devices_enrollment_consume(request: Request) -> dict[str, Any]:
    try:
        payload = await request.json()
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail="Invalid JSON payload") from exc
    if not isinstance(payload, dict):
        raise HTTPException(status_code=400, detail="Invalid JSON payload")

    tenant_id = _resolve_tenant_id_for_request(payload.get("tenantId"), DEFAULT_TENANT_ID)
    code = _normalize_enrollment_code(payload.get("code"))
    if len(code) < 8:
        raise HTTPException(status_code=400, detail="Invalid enrollment code")

    device_name = str(payload.get("deviceName") or "").strip()[:120] or "Terminal OpenIRN"
    platform = str(payload.get("platform") or "").strip()[:80]
    requested_device_id = _normalize_device_id(payload.get("deviceId"))
    now = _utc_now()
    code_hashes = _enrollment_code_hash_candidates(tenant_id, code)
    placeholders = ", ".join("?" for _ in code_hashes)

    with _db() as con:
        _ensure_tenant(con, tenant_id)
        enrollment = con.execute(
            f"""
            SELECT tenant_id, enrollment_id, created_by_user_id, label,
                   expires_at, consumed_at, consumed_by_device_id, created_at
            FROM device_enrollment_codes
            WHERE tenant_id = ? AND code_hash IN ({placeholders})
            """,
            (tenant_id, *code_hashes),
        ).fetchone()
        if enrollment is None:
            raise HTTPException(status_code=404, detail="Unknown enrollment code")
        if enrollment["consumed_at"]:
            raise HTTPException(status_code=409, detail="Enrollment code has already been consumed")
        if _parse_datetime(enrollment["expires_at"]) < now:
            raise HTTPException(status_code=410, detail="Enrollment code has expired")

        device, token = _create_device(
            con,
            tenant_id,
            name=device_name,
            platform=platform,
            invited_by_user_id=str(enrollment["created_by_user_id"] or ""),
            enrollment_id=str(enrollment["enrollment_id"] or ""),
            device_id=requested_device_id,
        )
        con.execute(
            """
            UPDATE device_enrollment_codes
            SET consumed_at = ?, consumed_by_device_id = ?
            WHERE tenant_id = ? AND enrollment_id = ?
            """,
            (now.isoformat(), device["deviceId"], tenant_id, enrollment["enrollment_id"]),
        )
        con.execute(
            """
            UPDATE device_enrollment_requests
            SET status = 'consumed'
            WHERE tenant_id = ? AND enrollment_id = ? AND status = 'approved'
            """,
            (tenant_id, enrollment["enrollment_id"]),
        )
        _record_device_audit(
            con,
            tenant_id,
            "enrollment.consumed",
            device_id=device["deviceId"],
            payload={
                "enrollmentId": enrollment["enrollment_id"],
                "deviceName": device_name,
                "platform": platform,
                "requestedDeviceId": requested_device_id,
            },
        )
        con.commit()

    return {
        "status": "ok",
        "type": "openirn.deviceEnrollmentConsumed",
        "tenantId": tenant_id,
        "apiToken": "",
        "device": device,
        "serverTime": _utc_now().isoformat(),
    }


@app.post("/devices/{device_id}/rename")
async def device_rename(device_id: str, request: Request) -> dict[str, Any]:
    try:
        payload = await request.json()
    except json.JSONDecodeError:
        payload = {}
    if not isinstance(payload, dict):
        payload = {}
    tenant_id = _resolve_tenant_id_for_request(payload.get("tenantId"), DEFAULT_TENANT_ID)
    _require_campaign_manager_authorization(request, tenant_id)
    name = str(payload.get("name") or "").strip()[:120]
    if not name:
        raise HTTPException(status_code=400, detail="Missing device name")

    with _db() as con:
        row = con.execute(
            "SELECT 1 FROM authorized_devices WHERE tenant_id = ? AND device_id = ?",
            (tenant_id, device_id),
        ).fetchone()
        if row is None:
            raise HTTPException(status_code=404, detail="Device not found")
        normalized_device_id = _normalize_device_id(device_id)
        now = _utc_now().isoformat()
        if _terminal_identity(con, normalized_device_id) is None:
            _ensure_terminal_identity(con, device_id=normalized_device_id, name=name, platform="")
        else:
            con.execute(
                "UPDATE terminals SET name = ?, updated_at = ? WHERE device_id = ?",
                (name, now, normalized_device_id),
            )
        con.execute(
            "UPDATE authorized_devices SET name = ? WHERE device_id = ?",
            (name, normalized_device_id),
        )
        _record_device_audit(
            con,
            tenant_id,
            "device.renamed",
            device_id=device_id,
            payload={"name": name},
        )
        con.commit()
        devices_list = _list_devices(con, tenant_id)
    return {
        "status": "ok",
        "type": "openirn.deviceRenamed",
        "tenantId": tenant_id,
        "deviceId": device_id,
        "devices": devices_list,
        "serverTime": _utc_now().isoformat(),
    }


@app.delete("/devices/{device_id}")
def device_revoke(
    device_id: str,
    request: Request,
    tenantId: str = Query(default="default", min_length=1, max_length=80),
) -> dict[str, Any]:
    tenant_id = _resolve_tenant_id_for_request(tenantId, DEFAULT_TENANT_ID)
    _require_campaign_manager_authorization(request, tenant_id)
    now = _utc_now().isoformat()
    with _db() as con:
        row = con.execute(
            """
            SELECT name, platform, status, last_seen_at, revoked_at, invited_by_user_id, enrollment_id
            FROM authorized_devices
            WHERE tenant_id = ? AND device_id = ?
            """,
            (tenant_id, device_id),
        ).fetchone()
        if row is None:
            raise HTTPException(status_code=404, detail="Device not found")
        con.execute(
            "DELETE FROM authorized_devices WHERE tenant_id = ? AND device_id = ?",
            (tenant_id, device_id),
        )
        _record_device_audit(
            con,
            tenant_id,
            "device.revoked",
            device_id=device_id,
            payload={
                "deleted": True,
                "deletedAt": now,
                "name": row["name"],
                "platform": row["platform"],
                "previousStatus": row["status"],
                "lastSeenAt": row["last_seen_at"],
                "previousRevokedAt": row["revoked_at"],
                "invitedByUserId": row["invited_by_user_id"],
                "enrollmentId": row["enrollment_id"],
            },
        )
        con.commit()
        devices_list = _list_devices(con, tenant_id)
    return {
        "status": "ok",
        "type": "openirn.deviceRevoked",
        "tenantId": tenant_id,
        "deviceId": device_id,
        "devices": devices_list,
        "serverTime": _utc_now().isoformat(),
    }


@app.get("/referential/official/status")
def official_referential_status(
    request: Request,
    tenantId: str = Query(default="default", min_length=1, max_length=80),
) -> dict[str, Any]:
    tenant_id = _resolve_tenant_id_for_request(tenantId, DEFAULT_TENANT_ID)
    _require_admin_authorization(request, tenant_id)
    remote = _official_remote_latest()
    with _db() as con:
        _ensure_tenant(con, tenant_id)
        _seed_official_referential_from_default(con, tenant_id)
        current_row = _load_current_official_referential(con, tenant_id)
        current = _official_referential_summary_from_row(current_row)
        con.commit()

    update_available = current is None or str(current.get("sourceBlobId") or "") != str(remote.get("blobId") or "")
    if current is not None and str(current.get("version") or "") != str(remote.get("version") or ""):
        update_available = True

    return {
        "status": "ok",
        "type": "openirn.officialReferentialStatus",
        "application": "OpenIRN API",
        "version": APP_VERSION,
        "tenantId": tenant_id,
        "referentialTenantId": str(current_row["tenant_id"] or "") if current_row is not None else tenant_id,
        "sharedAcrossTenants": bool(current_row is not None and str(current_row["tenant_id"] or "") != tenant_id),
        "serverTime": _utc_now().isoformat(),
        "source": {
            "provider": "gitlab",
            "projectPath": OFFICIAL_ADRI_PROJECT_PATH,
            "treePath": OFFICIAL_ADRI_TREE_PATH,
            "sourceUrl": OFFICIAL_ADRI_SOURCE_URL,
        },
        "scoringMethod": dict(OPENIRN_RNR_SCORING_METADATA),
        "current": current,
        "remote": remote,
        "updateAvailable": update_available,
    }


@app.get("/referential/official/current")
def official_referential_current(
    request: Request,
    tenantId: str = Query(default="default", min_length=1, max_length=80),
) -> dict[str, Any]:
    tenant_id = _resolve_tenant_id_for_request(tenantId, DEFAULT_TENANT_ID)
    if not _request_has_api_authorization(request):
        _require_active_device(request, tenant_id)
    with _db() as con:
        current_row = _load_current_official_referential(con, tenant_id)
        if current_row is None:
            raise HTTPException(status_code=404, detail="Aucun référentiel officiel n'est installé sur l'instance OpenIRN")
        payload = _parse_json(current_row["payload_json"], {})
    if not isinstance(payload, dict):
        raise HTTPException(status_code=500, detail="Référentiel officiel serveur invalide")
    summary = _official_referential_summary_from_row(current_row)
    source_tenant_id = str(current_row["tenant_id"] or "") if current_row is not None else tenant_id
    return {
        "status": "ok",
        "type": "openirn.officialReferentialCurrent",
        "application": "OpenIRN API",
        "version": APP_VERSION,
        "tenantId": tenant_id,
        "referentialTenantId": source_tenant_id,
        "sharedAcrossTenants": source_tenant_id != tenant_id,
        "serverTime": _utc_now().isoformat(),
        "referential": payload,
        "scoringMethod": payload.get("scoring") if isinstance(payload.get("scoring"), dict) else dict(OPENIRN_RNR_SCORING_METADATA),
        "summary": summary,
    }


@app.get("/referential/official/history")
def official_referential_history(
    request: Request,
    tenantId: str = Query(default="default", min_length=1, max_length=80),
    limit: int = Query(default=50, ge=1, le=200),
) -> dict[str, Any]:
    tenant_id = _resolve_tenant_id_for_request(tenantId, DEFAULT_TENANT_ID)
    _require_admin_authorization(request, tenant_id)
    with _db() as con:
        _ensure_tenant(con, tenant_id)
        _seed_official_referential_from_default(con, tenant_id)
        history = _list_official_referential_history(con, tenant_id, limit=limit)
        con.commit()
    return {
        "status": "ok",
        "type": "openirn.officialReferentialHistory",
        "application": "OpenIRN API",
        "version": APP_VERSION,
        "tenantId": tenant_id,
        "serverTime": _utc_now().isoformat(),
        "limit": limit,
        "count": len(history),
        "history": history,
    }


@app.post("/referential/official/update")
async def official_referential_update(request: Request) -> dict[str, Any]:
    try:
        payload = await request.json()
    except json.JSONDecodeError:
        payload = {}
    if not isinstance(payload, dict):
        payload = {}

    tenant_id = _resolve_tenant_id_for_request(payload.get("tenantId"), DEFAULT_TENANT_ID)
    auth_context = _require_admin_authorization(request, tenant_id)
    force = payload.get("force") is True
    triggered_by_user_id = (
        str(payload.get("triggeredByUserId") or "").strip()
        or str(auth_context.get("userId") or "").strip()
    )[:120]

    remote = _official_remote_latest()
    with _db() as con:
        _ensure_tenant(con, tenant_id)
        _seed_official_referential_from_default(con, tenant_id)
        current_row = _load_current_official_referential(con, tenant_id)
        current = _official_referential_summary_from_row(current_row)
        con.commit()

    update_available = current is None or str(current.get("sourceBlobId") or "") != str(remote.get("blobId") or "")
    if current is not None and str(current.get("version") or "") != str(remote.get("version") or ""):
        update_available = True

    if current is not None and not update_available and not force:
        return {
            "status": "up_to_date",
            "type": "openirn.officialReferentialUpdate",
            "application": "OpenIRN API",
            "version": APP_VERSION,
            "tenantId": tenant_id,
            "serverTime": _utc_now().isoformat(),
            "message": "Le référentiel officiel serveur est déjà aligné avec le dernier fichier GitLab détecté.",
            "current": current,
            "remote": remote,
            "updateAvailable": False,
        }

    protective_backup = _create_protective_backup(
        tenant_id,
        reason="pre_official_referential_update",
        triggered_by_user_id=triggered_by_user_id,
        context={
            "force": force,
            "currentVersion": current.get("version") if isinstance(current, dict) else None,
            "remoteVersion": remote.get("version"),
            "remoteBlobId": remote.get("blobId"),
            "remoteCommitSha": remote.get("commitSha"),
        },
    )

    raw_xlsx = _download_official_adri_xlsx(remote)
    referential = _adri_parse_workbook(raw_xlsx, version=str(remote.get("version") or "unknown"), remote=remote)
    validation = _adri_validation_report(referential)
    if validation.get("status") == "failed":
        raise HTTPException(status_code=422, detail={"message": "Le référentiel téléchargé n'a pas passé la validation OpenIRN", "validation": validation})

    with _db() as con:
        with con:
            _ensure_tenant(con, tenant_id)
            stored = _store_official_referential(
                con,
                tenant_id,
                referential,
                remote,
                validation,
                raw_xlsx,
                triggered_by_user_id=triggered_by_user_id,
            )
            con.execute(
                """
                INSERT INTO sync_events(tenant_id, event_type, server_sync_id, campaign_id, device_id, created_at, payload_json)
                VALUES (?, 'official_referential_updated', NULL, NULL, ?, ?, ?)
                """,
                (
                    tenant_id,
                    triggered_by_user_id or "server",
                    _utc_now().isoformat(),
                    _canonical_json({
                        "referentialId": stored["referentialId"],
                        "version": stored["version"],
                        "sourceBlobId": remote.get("blobId") or "",
                        "sourceCommitSha": remote.get("commitSha") or "",
                        "criterionCount": stored["criterionCount"],
                        "scoringMethodStatus": OPENIRN_RNR_SCORING_METADATA["methodStatus"],
                    }),
                ),
            )
            current_row = _load_current_official_referential(con, tenant_id)
            current = _official_referential_summary_from_row(current_row)

    return {
        "status": "updated",
        "type": "openirn.officialReferentialUpdate",
        "application": "OpenIRN API",
        "version": APP_VERSION,
        "tenantId": tenant_id,
        "serverTime": _utc_now().isoformat(),
        "message": "Référentiel officiel aDRI téléchargé, validé et installé sur le serveur.",
        "current": current,
        "remote": remote,
        "validation": validation,
        "stored": stored,
        "protectiveBackup": protective_backup,
        "updateAvailable": False,
    }

@app.get("/health")
def health() -> dict[str, Any]:
    """Public, compact health check for monitoring."""
    status = "ok"
    tenant_number = 0
    try:
        with _db() as con:
            row = con.execute("SELECT COUNT(*) FROM tenants").fetchone()
            tenant_number = int(row[0]) if row and row[0] is not None else 0
            con.commit()
    except Exception:
        status = "degraded"
        tenant_number = 0

    return {
        "status": status,
        "application": "OpenIRN API",
        "version": APP_VERSION,
        "storage": _db_backend(),
        "tenantNumber": tenant_number,
        "authRequired": True,
        "authMode": "server_session_with_role_policy",
        "serverTime": _utc_now().isoformat(),
    }


@app.post("/sync/push")
async def sync_push(request: Request) -> dict[str, Any]:
    try:
        payload = await request.json()
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail="Invalid JSON payload") from exc

    if not isinstance(payload, dict):
        raise HTTPException(status_code=400, detail="Payload must be a JSON object")
    if payload.get("type") != "openirn.syncPush":
        raise HTTPException(status_code=400, detail="Unsupported payload type")

    sync_context = payload.get("sync")
    if not isinstance(sync_context, dict):
        raise HTTPException(status_code=400, detail="Missing sync context")

    tenant_id = _resolve_tenant_id_for_request(sync_context.get("tenantId"), DEFAULT_TENANT_ID)
    _require_write_authorization(request, tenant_id)
    device_id = _safe_segment(sync_context.get("deviceId"), "unknown-device")
    campaigns = _extract_campaigns(payload)

    received_at = _utc_now().isoformat()
    server_sync_id = f"sync_{_utc_now().strftime('%Y%m%dT%H%M%SZ')}_{uuid.uuid4().hex[:12]}"
    payload_sha256 = _json_sha256(payload)
    envelope = {
        "serverSyncId": server_sync_id,
        "receivedAt": received_at,
        "tenantId": tenant_id,
        "deviceId": device_id,
        "payloadSha256": payload_sha256,
        "payload": payload,
    }

    with _db() as con:
        with con:
            _ensure_tenant(con, tenant_id)
            con.execute(
                """
                INSERT INTO sync_snapshots(
                    tenant_id, server_sync_id, device_id, received_at,
                    payload_sha256, campaign_count, payload_json, envelope_json
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    tenant_id,
                    server_sync_id,
                    device_id,
                    received_at,
                    payload_sha256,
                    len(campaigns),
                    _canonical_json(payload),
                    _canonical_json(envelope),
                ),
            )
            central_user_count = _merge_central_users(con, tenant_id, payload.get("users"))
            revision_stats = _record_campaign_revisions(con, tenant_id, server_sync_id, device_id, received_at, payload)
            con.execute(
                """
                INSERT INTO sync_events(
                    tenant_id, event_type, server_sync_id, device_id,
                    created_at, payload_json
                ) VALUES (?, 'snapshot_accepted', ?, ?, ?, ?)
                """,
                (
                    tenant_id,
                    server_sync_id,
                    device_id,
                    _utc_now().isoformat(),
                    _canonical_json(
                        {
                            "serverSyncId": server_sync_id,
                            "campaignCount": len(campaigns),
                            "revisionCount": revision_stats["revisionCount"],
                            "conflictCount": revision_stats["conflictCount"],
                            "deletedCount": revision_stats.get("deletedCount", 0),
                        }
                    ),
                ),
            )

    return {
        "status": "accepted",
        "application": "OpenIRN API",
        "version": APP_VERSION,
        "storage": "mariadb",
        "serverSyncId": server_sync_id,
        "receivedAt": received_at,
        "tenantId": tenant_id,
        "deviceId": device_id,
        "payloadSha256": payload_sha256,
        "stored": True,
        "campaignCount": len(campaigns),
        "centralUserCount": central_user_count,
        "campaignRevisionCount": revision_stats["revisionCount"],
        "conflictCount": revision_stats["conflictCount"],
        "deletedCount": revision_stats.get("deletedCount", 0),
        "conflictPolicy": "last_write_wins",
    }


@app.post("/auth/verify")
async def auth_verify(request: Request) -> dict[str, Any]:
    try:
        payload = await request.json()
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail="Invalid JSON payload") from exc

    if not isinstance(payload, dict):
        raise HTTPException(status_code=400, detail="Payload must be a JSON object")

    tenant_id = _resolve_tenant_id_for_request(payload.get("tenantId"), DEFAULT_TENANT_ID)
    user_id = str(payload.get("userId") or "").strip()
    pin = str(payload.get("pin") or "")
    device_id = _require_active_device(request, tenant_id, payload)
    ip_address = _request_client_ip(request)
    if not user_id:
        raise HTTPException(status_code=400, detail="Missing userId")
    if not pin.strip():
        raise HTTPException(status_code=400, detail="Code personnel manquant")

    with _db() as con:
        _ensure_tenant(con, tenant_id)
        _enforce_auth_rate_limit(
            con,
            tenant_id,
            device_id=device_id,
            user_id=user_id,
            ip_address=ip_address,
        )
        users = _load_central_users(con, tenant_id)
        _ensure_user_credentials(con, tenant_id, users)
        con.commit()
        user = next((candidate for candidate in users if candidate.get("id") == user_id), None)
        if user is None:
            _record_auth_attempt(
                con,
                tenant_id,
                device_id=device_id,
                user_id=user_id,
                ip_address=ip_address,
                successful=False,
                reason="unknown_user",
            )
            _record_device_audit(
                con,
                tenant_id,
                "auth.failed",
                device_id=device_id,
                payload={"userId": user_id, "reason": "unknown_user", "ipAddress": ip_address},
            )
            con.commit()
            raise HTTPException(status_code=404, detail="Unknown user")
        if user.get("active") is not True:
            _record_auth_attempt(
                con,
                tenant_id,
                device_id=device_id,
                user_id=user_id,
                ip_address=ip_address,
                successful=False,
                reason="inactive_user",
            )
            _record_device_audit(
                con,
                tenant_id,
                "auth.failed",
                device_id=device_id,
                payload={"userId": user_id, "reason": "inactive_user", "ipAddress": ip_address},
            )
            con.commit()
            raise HTTPException(status_code=403, detail="Inactive user")
        accepted, requires_change = _verify_user_pin(con, tenant_id, user_id, pin)
        if not accepted:
            _record_auth_attempt(
                con,
                tenant_id,
                device_id=device_id,
                user_id=user_id,
                ip_address=ip_address,
                successful=False,
                reason="invalid_pin",
            )
            _record_device_audit(
                con,
                tenant_id,
                "auth.failed",
                device_id=device_id,
                payload={"userId": user_id, "reason": "invalid_pin", "ipAddress": ip_address},
            )
            con.commit()
            raise HTTPException(status_code=403, detail="Invalid user code")
        session_id, session_token, expires_at = _create_api_session(
            con,
            tenant_id,
            device_id,
            user_id,
        )
        _record_auth_attempt(
            con,
            tenant_id,
            device_id=device_id,
            user_id=user_id,
            ip_address=ip_address,
            successful=True,
            reason="accepted",
        )
        _record_device_audit(
            con,
            tenant_id,
            "session.created",
            device_id=device_id,
            payload={"userId": user_id, "sessionId": session_id, "ipAddress": ip_address},
        )
        con.commit()

    return {
        "status": "accepted",
        "type": "openirn.userAuthentication",
        "application": "OpenIRN API",
        "version": APP_VERSION,
        "tenantId": tenant_id,
        "serverTime": _utc_now().isoformat(),
        "userId": user_id,
        "mustChangePin": requires_change,
        "user": user,
        "sessionId": session_id,
        "apiToken": session_token,
        "expiresAt": expires_at.isoformat(),
        "sessionTtlMinutes": max(5, SESSION_TTL_MINUTES),
        "idleTimeoutMinutes": max(1, SESSION_IDLE_TIMEOUT_MINUTES),
    }



@app.post("/auth/change-pin")
async def auth_change_pin(request: Request) -> dict[str, Any]:
    try:
        payload = await request.json()
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail="Invalid JSON payload") from exc

    if not isinstance(payload, dict):
        raise HTTPException(status_code=400, detail="Payload must be a JSON object")

    auth_context = _request_auth_context(request)
    if auth_context is None:
        raise _authorization_unavailable_exception()
    if str(auth_context.get("authMode") or "") != "session":
        raise HTTPException(
            status_code=403,
            detail="Cette opération exige une session utilisateur active",
        )

    tenant_id = str(auth_context.get("tenantId") or "").strip()
    user_id = str(auth_context.get("userId") or "").strip()
    user_role = _role_normalize(auth_context.get("userRole"))
    device_id = str(auth_context.get("deviceId") or _request_device_id(request, payload) or "").strip()[:160]
    if not tenant_id or not user_id:
        raise HTTPException(status_code=403, detail="Session utilisateur incomplète")
    if user_role not in API_ROLE_READ:
        raise HTTPException(status_code=403, detail="Votre profil ne permet pas de changer ce code")

    requested_tenant_id = str(payload.get("tenantId") or tenant_id).strip()
    if requested_tenant_id and _safe_segment(requested_tenant_id, tenant_id) != tenant_id:
        raise HTTPException(
            status_code=403,
            detail="La session ne correspond pas à l’espace de travail demandé",
        )

    current_pin = str(payload.get("currentPin") or payload.get("oldPin") or "")
    new_pin = str(payload.get("newPin") or payload.get("pin") or "").strip()
    if not current_pin.strip():
        raise HTTPException(status_code=400, detail="Code actuel manquant")
    if len(new_pin) < 4 or len(new_pin) > 32:
        raise HTTPException(status_code=400, detail="Le nouveau code doit contenir entre 4 et 32 caractères")

    ip_address = _request_client_ip(request)
    with _db() as con:
        _ensure_tenant(con, tenant_id)
        _enforce_auth_rate_limit(
            con,
            tenant_id,
            device_id=device_id,
            user_id=user_id,
            ip_address=ip_address,
        )
        users = _load_central_users(con, tenant_id)
        user = next((candidate for candidate in users if candidate.get("id") == user_id), None)
        if user is None or user.get("active") is not True:
            raise HTTPException(status_code=403, detail="Utilisateur inactif ou introuvable")
        accepted, _requires_change = _verify_user_pin(con, tenant_id, user_id, current_pin)
        if not accepted:
            _record_auth_attempt(
                con,
                tenant_id,
                device_id=device_id,
                user_id=user_id,
                ip_address=ip_address,
                successful=False,
                reason="invalid_current_pin",
            )
            _record_device_audit(
                con,
                tenant_id,
                "user.pin_change_failed",
                device_id=device_id,
                payload={"userId": user_id, "reason": "invalid_current_pin", "ipAddress": ip_address},
            )
            con.commit()
            raise HTTPException(status_code=403, detail="Le code actuel est incorrect")
        con.commit()

    protective_backup = _create_protective_backup(
        tenant_id,
        reason="pre_self_pin_change",
        triggered_by_user_id=user_id,
        context={"userId": user_id},
    )

    with _db() as con:
        with con:
            _ensure_tenant(con, tenant_id)
            _set_user_pin(con, tenant_id, user_id, new_pin, requires_change=False)
            _record_auth_attempt(
                con,
                tenant_id,
                device_id=device_id,
                user_id=user_id,
                ip_address=ip_address,
                successful=True,
                reason="pin_changed",
            )
            _record_device_audit(
                con,
                tenant_id,
                "user.pin_changed",
                device_id=device_id,
                payload={"userId": user_id, "ipAddress": ip_address},
            )

    return {
        "status": "accepted",
        "type": "openirn.selfPinUpdate",
        "application": "OpenIRN API",
        "version": APP_VERSION,
        "tenantId": tenant_id,
        "serverTime": _utc_now().isoformat(),
        "userId": user_id,
        "protectiveBackup": protective_backup,
        "message": "Code d’accès mis à jour.",
    }


def _session_row_to_payload(row: Any, *, current_token_hash: str = '') -> dict[str, Any]:
    now = _utc_now()
    expires_at = _parse_datetime(row["expires_at"])
    revoked_at_raw = row["revoked_at"]
    revoked_at = _parse_datetime(revoked_at_raw) if revoked_at_raw else None
    if revoked_at is not None:
        status = "revoked"
    elif expires_at < now:
        status = "expired"
    else:
        status = "active"

    user_payload = _parse_json(row["user_payload_json"], {})
    first_name = str(user_payload.get("firstName") or user_payload.get("first_name") or "").strip()
    last_name = str(user_payload.get("lastName") or user_payload.get("last_name") or "").strip()
    full_name = " ".join(part for part in (first_name, last_name) if part).strip()
    email = str(user_payload.get("email") or "").strip()
    role = str(user_payload.get("role") or row["user_role"] or "").strip()

    return {
        "sessionId": row["session_id"],
        "tenantId": row["tenant_id"],
        "deviceId": row["device_id"],
        "deviceName": row["device_name"] or row["device_id"],
        "devicePlatform": row["device_platform"] or "",
        "userId": row["user_id"],
        "userDisplayName": full_name or email or row["user_id"],
        "userEmail": email,
        "userRole": role,
        "status": status,
        "isCurrentSession": bool(current_token_hash and hmac.compare_digest(row["token_hash"], current_token_hash)),
        "createdAt": row["created_at"],
        "expiresAt": row["expires_at"],
        "lastSeenAt": row["last_seen_at"],
        "revokedAt": row["revoked_at"],
    }



@app.get("/security/audit")
def security_audit(
    request: Request,
    tenantId: str = Query(default="default", min_length=1, max_length=80),
    limit: int = Query(default=100, ge=25, le=500),
    includeAuthAttempts: bool = Query(default=True),
    includeDeviceAudit: bool = Query(default=True),
) -> dict[str, Any]:
    tenant_id = _resolve_tenant_id_for_request(tenantId, DEFAULT_TENANT_ID)
    _require_admin_authorization(request, tenant_id)
    safe_limit = max(25, min(int(limit), 500))
    events: list[dict[str, Any]] = []

    with _db() as con:
        _ensure_tenant(con, tenant_id)
        if includeDeviceAudit:
            audit_rows = con.execute(
                """
                SELECT id, tenant_id, device_id, event_type, created_at, payload_json
                FROM device_audit_log
                WHERE tenant_id = ?
                ORDER BY created_at DESC, id DESC
                LIMIT ?
                """,
                (tenant_id, safe_limit),
            ).fetchall()
            for row in audit_rows:
                events.append(
                    {
                        "source": "deviceAudit",
                        "eventId": f"audit-{row['id']}",
                        "tenantId": row["tenant_id"],
                        "deviceId": row["device_id"] or "",
                        "eventType": row["event_type"] or "",
                        "createdAt": row["created_at"],
                        "payload": _parse_json(row["payload_json"], {}),
                    }
                )

        if includeAuthAttempts:
            attempt_rows = con.execute(
                """
                SELECT
                    tenant_id, attempt_id, device_id, user_id, ip_address,
                    successful, reason, created_at
                FROM auth_attempts
                WHERE tenant_id = ?
                ORDER BY created_at DESC, attempt_id DESC
                LIMIT ?
                """,
                (tenant_id, safe_limit),
            ).fetchall()
            for row in attempt_rows:
                events.append(
                    {
                        "source": "authAttempt",
                        "eventId": row["attempt_id"],
                        "tenantId": row["tenant_id"],
                        "deviceId": row["device_id"] or "",
                        "eventType": "auth.success" if bool(row["successful"]) else "auth.failed",
                        "createdAt": row["created_at"],
                        "userId": row["user_id"] or "",
                        "ipAddress": row["ip_address"] or "",
                        "successful": bool(row["successful"]),
                        "reason": row["reason"] or "",
                        "payload": {
                            "reason": row["reason"] or "",
                            "successful": bool(row["successful"]),
                        },
                    }
                )

    events.sort(key=lambda item: str(item.get("createdAt") or ""), reverse=True)
    events = events[:safe_limit]
    auth_count = sum(1 for event in events if event.get("source") == "authAttempt")
    device_count = sum(1 for event in events if event.get("source") == "deviceAudit")
    failure_count = sum(1 for event in events if event.get("successful") is False)

    return {
        "status": "ok",
        "type": "openirn.securityAudit",
        "application": "OpenIRN API",
        "version": APP_VERSION,
        "tenantId": tenant_id,
        "serverTime": _utc_now().isoformat(),
        "eventCount": len(events),
        "authAttemptCount": auth_count,
        "deviceAuditCount": device_count,
        "failureCount": failure_count,
        "events": events,
    }


@app.delete("/auth/session/current")
def revoke_current_auth_session(
    request: Request,
    tenantId: str = Query(default="default", min_length=1, max_length=80),
) -> dict[str, Any]:
    tenant_id = _resolve_tenant_id_for_request(tenantId, DEFAULT_TENANT_ID)
    provided_token = _extract_bearer_token(request)
    if not provided_token or not provided_token.startswith("ost_"):
        raise HTTPException(status_code=401, detail="Session courante requise")
    token_hash = _secret_hash(provided_token)
    now = _utc_now()

    with _db() as con:
        _ensure_tenant(con, tenant_id)
        row = con.execute(
            """
            SELECT tenant_id, session_id, device_id, user_id, revoked_at
            FROM api_sessions
            WHERE tenant_id = ? AND token_hash = ?
            """,
            (tenant_id, token_hash),
        ).fetchone()
        if row is None:
            row = con.execute(
                """
                SELECT tenant_id, session_id, device_id, user_id, revoked_at
                FROM api_sessions
                WHERE token_hash = ?
                """,
                (token_hash,),
            ).fetchone()
        if row is None:
            raise HTTPException(status_code=404, detail="Session inconnue")
        session_tenant_id = str(row["tenant_id"] or tenant_id)
        if row["revoked_at"] is None:
            con.execute(
                """
                UPDATE api_sessions
                SET revoked_at = ?
                WHERE tenant_id = ? AND session_id = ?
                """,
                (now.isoformat(), session_tenant_id, row["session_id"]),
            )
            _record_device_audit(
                con,
                session_tenant_id,
                "session.locked",
                device_id=row["device_id"],
                payload={
                    "sessionId": row["session_id"],
                    "userId": row["user_id"],
                    "requestedTenantId": tenant_id,
                },
            )
            con.commit()

    return {
        "status": "ok",
        "type": "openirn.currentApiSessionRevoked",
        "application": "OpenIRN API",
        "version": APP_VERSION,
        "tenantId": session_tenant_id,
        "requestedTenantId": tenant_id,
        "serverTime": _utc_now().isoformat(),
        "sessionId": row["session_id"],
        "message": "Session courante verrouillée.",
    }


@app.get("/auth/sessions")
def auth_sessions(
    request: Request,
    tenantId: str = Query(default="default", min_length=1, max_length=80),
    includeInactive: bool = Query(default=True),
) -> dict[str, Any]:
    tenant_id = _resolve_tenant_id_for_request(tenantId, DEFAULT_TENANT_ID)
    _require_admin_authorization(request, tenant_id)
    provided_token = _extract_bearer_token(request)
    current_token_hash = _secret_hash(provided_token) if provided_token.startswith("ost_") else ""
    now_iso = _utc_now().isoformat()

    where_inactive = "" if includeInactive else "AND s.revoked_at IS NULL AND s.expires_at >= ?"
    params: tuple[Any, ...]
    if includeInactive:
        params = (tenant_id,)
    else:
        params = (tenant_id, now_iso)

    with _db() as con:
        _ensure_tenant(con, tenant_id)
        rows = con.execute(
            f"""
            SELECT
                s.tenant_id,
                s.session_id,
                s.token_hash,
                s.device_id,
                s.user_id,
                s.created_at,
                s.expires_at,
                s.last_seen_at,
                s.revoked_at,
                d.name AS device_name,
                d.platform AS device_platform,
                u.role AS user_role,
                u.payload_json AS user_payload_json
            FROM api_sessions s
            LEFT JOIN authorized_devices d
              ON d.tenant_id = s.tenant_id AND d.device_id = s.device_id
            LEFT JOIN users u
              ON u.tenant_id = s.tenant_id AND u.user_id = s.user_id
            WHERE s.tenant_id = ?
              {where_inactive}
            ORDER BY
              CASE WHEN s.revoked_at IS NULL AND s.expires_at >= ? THEN 0 ELSE 1 END,
              s.last_seen_at DESC,
              s.created_at DESC
            LIMIT 250
            """,
            params + (now_iso,),
        ).fetchall()

    sessions = [
        _session_row_to_payload(row, current_token_hash=current_token_hash)
        for row in rows
    ]
    active_count = sum(1 for session in sessions if session["status"] == "active")
    expired_count = sum(1 for session in sessions if session["status"] == "expired")
    revoked_count = sum(1 for session in sessions if session["status"] == "revoked")

    return {
        "status": "ok",
        "type": "openirn.apiSessions",
        "application": "OpenIRN API",
        "version": APP_VERSION,
        "tenantId": tenant_id,
        "serverTime": _utc_now().isoformat(),
        "sessionCount": len(sessions),
        "activeCount": active_count,
        "expiredCount": expired_count,
        "revokedCount": revoked_count,
        "sessions": sessions,
    }


@app.delete("/auth/sessions/{session_id}")
def revoke_auth_session(
    session_id: str,
    request: Request,
    tenantId: str = Query(default="default", min_length=1, max_length=80),
) -> dict[str, Any]:
    tenant_id = _resolve_tenant_id_for_request(tenantId, DEFAULT_TENANT_ID)
    safe_session_id = str(session_id or "").strip()[:160]
    if not safe_session_id:
        raise HTTPException(status_code=400, detail="Missing session id")
    _require_admin_authorization(request, tenant_id)
    provided_token = _extract_bearer_token(request)
    current_token_hash = _secret_hash(provided_token) if provided_token.startswith("ost_") else ""
    now = _utc_now()

    with _db() as con:
        _ensure_tenant(con, tenant_id)
        row = con.execute(
            """
            SELECT session_id, token_hash, device_id, user_id, revoked_at
            FROM api_sessions
            WHERE tenant_id = ? AND session_id = ?
            """,
            (tenant_id, safe_session_id),
        ).fetchone()
        if row is None:
            raise HTTPException(status_code=404, detail="Session inconnue")
        if current_token_hash and hmac.compare_digest(row["token_hash"], current_token_hash):
            raise HTTPException(status_code=400, detail="La session courante ne peut pas être révoquée depuis cette action")
        if row["revoked_at"] is None:
            con.execute(
                """
                UPDATE api_sessions
                SET revoked_at = ?
                WHERE tenant_id = ? AND session_id = ?
                """,
                (now.isoformat(), tenant_id, safe_session_id),
            )
            _record_device_audit(
                con,
                tenant_id,
                "session.revoked",
                device_id=row["device_id"],
                payload={"sessionId": safe_session_id, "userId": row["user_id"]},
            )
            con.commit()

    return {
        "status": "ok",
        "type": "openirn.apiSessionRevoked",
        "application": "OpenIRN API",
        "version": APP_VERSION,
        "tenantId": tenant_id,
        "serverTime": _utc_now().isoformat(),
        "sessionId": safe_session_id,
        "message": "Session révoquée.",
    }


@app.get("/users")
def users(
    request: Request,
    tenantId: str = Query(default="default", min_length=1, max_length=80),
    allTenants: bool = Query(default=False),
) -> dict[str, Any]:
    tenant_id = _resolve_tenant_id_for_request(tenantId, DEFAULT_TENANT_ID)
    include_all_tenants = False
    if allTenants:
        auth_context = _require_admin_authorization(request, tenant_id)
        include_all_tenants = _is_administrator_context(auth_context)
        if not include_all_tenants:
            raise HTTPException(status_code=403, detail="La vue multi-espaces est réservée à l’administrateur")
    else:
        _require_device_or_authorized_read(request, tenant_id)

    with _db() as con:
        _ensure_tenant(con, tenant_id)
        if include_all_tenants:
            rows = con.execute(
                """
                SELECT u.tenant_id, COALESCE(NULLIF(t.display_name, ''), 'Espace de travail') AS tenant_display_name,
                       u.user_id, u.first_name, u.last_name, u.email, u.role,
                       u.active, u.created_at, u.updated_at, u.payload_json
                FROM users u
                LEFT JOIN tenants t ON t.id = u.tenant_id
                ORDER BY COALESCE(NULLIF(t.display_name, ''), 'Espace de travail') ASC, u.active DESC,
                         u.last_name ASC, u.first_name ASC, u.email ASC
                """
            ).fetchall()
            central_users = _sort_users([_row_to_user(row) for row in rows])
        else:
            central_users = _load_central_users(con, tenant_id)
            _ensure_user_credentials(con, tenant_id, central_users)
        tenant_display_name = _tenant_display_name(con, tenant_id)
        con.commit()

    return {
        "status": "ok",
        "type": "openirn.centralUsers",
        "application": "OpenIRN API",
        "version": APP_VERSION,
        "tenantId": tenant_id,
        "tenantDisplayName": tenant_display_name,
        "scope": "all_tenants" if include_all_tenants else "tenant",
        "serverTime": _utc_now().isoformat(),
        "userCount": len(central_users),
        "users": central_users,
    }


@app.post("/users/replace")
async def users_replace(request: Request) -> dict[str, Any]:
    try:
        payload = await request.json()
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail="Invalid JSON payload") from exc

    if not isinstance(payload, dict):
        raise HTTPException(status_code=400, detail="Payload must be a JSON object")

    tenant_id = _resolve_tenant_id_for_request(payload.get("tenantId"), DEFAULT_TENANT_ID)
    auth_context = _require_campaign_manager_authorization(request, tenant_id)
    triggered_by_user_id = str(auth_context.get("userId") or "").strip()
    raw_users = payload.get("users")
    if not isinstance(raw_users, list):
        raise HTTPException(status_code=400, detail="Missing users array")

    users_to_save = [user for raw_user in raw_users if (user := _sanitize_user(raw_user))]
    if _role_normalize(auth_context.get("userRole")) != "administrator" and any(
        user.get("role") == "administrator" for user in users_to_save
    ):
        raise HTTPException(status_code=403, detail="Un Pilote IRN ne peut pas créer ou modifier un profil Administrateur")
    protective_backup = _create_protective_backup(
        tenant_id,
        reason="pre_users_replace",
        triggered_by_user_id=triggered_by_user_id,
        context={"incomingUserCount": len(users_to_save)},
    )
    with _db() as con:
        with con:
            _ensure_tenant(con, tenant_id)
            resolved_users: list[dict[str, Any]] = []
            for user in users_to_save:
                original_user_id = str(user.get("id") or "").strip()
                resolved_user_id = _user_id_for_save(con, tenant_id, original_user_id)
                if resolved_user_id != original_user_id:
                    user = dict(user)
                    user["id"] = resolved_user_id
                resolved_users.append(user)
                _save_user(con, tenant_id, user)
            users_to_save = resolved_users
            user_ids = {user["id"] for user in users_to_save}
            if user_ids:
                placeholders = ",".join("?" for _ in user_ids)
                con.execute(
                    f"DELETE FROM users WHERE tenant_id = ? AND user_id NOT IN ({placeholders})",
                    (tenant_id, *sorted(user_ids)),
                )
            else:
                con.execute("DELETE FROM users WHERE tenant_id = ?", (tenant_id,))
            _ensure_user_credentials(con, tenant_id, users_to_save)

    return {
        "status": "accepted",
        "type": "openirn.centralUsersReplace",
        "application": "OpenIRN API",
        "version": APP_VERSION,
        "tenantId": tenant_id,
        "serverTime": _utc_now().isoformat(),
        "userCount": len(users_to_save),
        "protectiveBackup": protective_backup,
    }


@app.post("/users/pin")
async def users_pin(request: Request) -> dict[str, Any]:
    try:
        payload = await request.json()
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail="Invalid JSON payload") from exc

    if not isinstance(payload, dict):
        raise HTTPException(status_code=400, detail="Payload must be a JSON object")

    tenant_id = _resolve_tenant_id_for_request(payload.get("tenantId"), DEFAULT_TENANT_ID)
    auth_context = _require_campaign_manager_authorization(request, tenant_id)
    triggered_by_user_id = str(auth_context.get("userId") or "").strip()
    user_id = str(payload.get("userId") or "").strip()
    pin = str(payload.get("pin") or payload.get("newPin") or "").strip()
    if not user_id:
        raise HTTPException(status_code=400, detail="Missing userId")

    if _role_normalize(auth_context.get("userRole")) != "administrator":
        with _db() as con:
            target = con.execute(
                "SELECT role FROM users WHERE tenant_id = ? AND user_id = ?",
                (tenant_id, user_id),
            ).fetchone()
            if target is not None and _role_normalize(target["role"]) == "administrator":
                raise HTTPException(status_code=403, detail="Un Pilote IRN ne peut pas modifier le code d’un Administrateur")

    protective_backup = _create_protective_backup(
        tenant_id,
        reason="pre_user_pin_change",
        triggered_by_user_id=triggered_by_user_id,
        context={"userId": user_id},
    )

    with _db() as con:
        with con:
            _ensure_tenant(con, tenant_id)
            _set_user_pin(con, tenant_id, user_id, pin, requires_change=False)

    return {
        "status": "accepted",
        "type": "openirn.userPinUpdate",
        "application": "OpenIRN API",
        "version": APP_VERSION,
        "tenantId": tenant_id,
        "serverTime": _utc_now().isoformat(),
        "userId": user_id,
        "protectiveBackup": protective_backup,
    }


@app.get("/sync/status")
def sync_status(request: Request, tenantId: str = Query(default="default", min_length=1, max_length=80)) -> dict[str, Any]:
    tenant_id = _resolve_tenant_id_for_request(tenantId, DEFAULT_TENANT_ID)
    _require_sync_read_access(request, tenant_id)

    with _db() as con:
        latest_row = con.execute(
            """
            SELECT tenant_id, server_sync_id, device_id, received_at,
                   payload_sha256, campaign_count, payload_json
            FROM sync_snapshots
            WHERE tenant_id = ?
            ORDER BY received_at DESC, server_sync_id DESC
            LIMIT 1
            """,
            (tenant_id,),
        ).fetchone()
        snapshot_count = int(
            con.execute("SELECT COUNT(*) FROM sync_snapshots WHERE tenant_id = ?", (tenant_id,)).fetchone()[0]
        )
        device_count = int(
            con.execute("SELECT COUNT(DISTINCT device_id) FROM sync_snapshots WHERE tenant_id = ?", (tenant_id,)).fetchone()[0]
        )
        campaign_count = int(
            con.execute("SELECT COALESCE(SUM(campaign_count), 0) FROM sync_snapshots WHERE tenant_id = ?", (tenant_id,)).fetchone()[0]
        )
        current_campaign_count = int(
            con.execute("SELECT COUNT(*) FROM campaign_states WHERE tenant_id = ?", (tenant_id,)).fetchone()[0]
        )
        conflict_count = int(
            con.execute(
                "SELECT COUNT(*) FROM campaign_revisions WHERE tenant_id = ? AND conflict_detected = 1",
                (tenant_id,),
            ).fetchone()[0]
        )

    return {
        "status": "ok",
        "type": "openirn.syncStatus",
        "application": "OpenIRN API",
        "version": APP_VERSION,
        "storage": "mariadb",
        "tenantId": tenant_id,
        "serverTime": _utc_now().isoformat(),
        "snapshotCount": snapshot_count,
        "deviceCount": device_count,
        "campaignCount": campaign_count,
        "currentCampaignCount": current_campaign_count,
        "conflictCount": conflict_count,
        "conflictPolicy": "last_write_wins",
        "latestSnapshot": _snapshot_summary_from_row(latest_row),
    }


@app.get("/sync/pull")
def sync_pull(
    request: Request,
    tenantId: str = Query(default="default", min_length=1, max_length=80),
    limit: int = Query(default=10, ge=1, le=50),
) -> dict[str, Any]:
    tenant_id = _resolve_tenant_id_for_request(tenantId, DEFAULT_TENANT_ID)
    _require_sync_read_access(request, tenant_id)

    with _db() as con:
        rows = con.execute(
            """
            SELECT tenant_id, server_sync_id, device_id, received_at,
                   payload_sha256, campaign_count, payload_json
            FROM sync_snapshots
            WHERE tenant_id = ?
            ORDER BY received_at DESC, server_sync_id DESC
            LIMIT ?
            """,
            (tenant_id, limit),
        ).fetchall()
        available_snapshot_count = int(
            con.execute("SELECT COUNT(*) FROM sync_snapshots WHERE tenant_id = ?", (tenant_id,)).fetchone()[0]
        )

    snapshots = [_public_snapshot_from_row(row) for row in rows]
    return {
        "status": "ok",
        "type": "openirn.syncPull",
        "application": "OpenIRN API",
        "version": APP_VERSION,
        "storage": "mariadb",
        "tenantId": tenant_id,
        "serverTime": _utc_now().isoformat(),
        "snapshotCount": len(snapshots),
        "availableSnapshotCount": available_snapshot_count,
        "limit": limit,
        "snapshots": snapshots,
    }


@app.get("/campaigns")
def campaigns(
    request: Request,
    tenantId: str = Query(default="default", min_length=1, max_length=80),
    limit: int = Query(default=100, ge=1, le=500),
) -> dict[str, Any]:
    tenant_id = _resolve_tenant_id_for_request(tenantId, DEFAULT_TENANT_ID)
    _require_sync_read_access(request, tenant_id)

    with _db() as con:
        rows = con.execute(
            """
            SELECT tenant_id, campaign_id, server_revision, server_sync_id,
                   device_id, updated_at, received_at, payload_sha256,
                   payload_json, conflict_policy
            FROM campaign_states
            WHERE tenant_id = ?
            ORDER BY received_at DESC, updated_at DESC, campaign_id ASC
            LIMIT ?
            """,
            (tenant_id, limit),
        ).fetchall()
        revision_count = int(
            con.execute(
                "SELECT COUNT(*) FROM campaign_revisions WHERE tenant_id = ?",
                (tenant_id,),
            ).fetchone()[0]
        )
        conflict_count = int(
            con.execute(
                "SELECT COUNT(*) FROM campaign_revisions WHERE tenant_id = ? AND conflict_detected = 1",
                (tenant_id,),
            ).fetchone()[0]
        )

    items = [_public_campaign_state_from_row(row) for row in rows]
    return {
        "status": "ok",
        "type": "openirn.campaignStates",
        "application": "OpenIRN API",
        "version": APP_VERSION,
        "storage": "mariadb",
        "tenantId": tenant_id,
        "serverTime": _utc_now().isoformat(),
        "campaignCount": len(items),
        "revisionCount": revision_count,
        "conflictCount": conflict_count,
        "campaigns": items,
    }


@app.get("/campaigns/revisions")
def campaign_revisions(
    request: Request,
    tenantId: str = Query(default="default", min_length=1, max_length=80),
    campaignId: str = Query(min_length=1, max_length=240),
    limit: int = Query(default=50, ge=1, le=200),
) -> dict[str, Any]:
    tenant_id = _resolve_tenant_id_for_request(tenantId, DEFAULT_TENANT_ID)
    _require_campaign_manager_authorization(request, tenant_id)
    campaign_id = str(campaignId or "").strip()

    with _db() as con:
        state_row = con.execute(
            """
            SELECT tenant_id, campaign_id, server_revision, server_sync_id,
                   device_id, updated_at, received_at, payload_sha256,
                   payload_json, conflict_policy
            FROM campaign_states
            WHERE tenant_id = ? AND campaign_id = ?
            """,
            (tenant_id, campaign_id),
        ).fetchone()
        rows = con.execute(
            """
            SELECT tenant_id, campaign_id, server_revision, server_sync_id,
                   device_id, updated_at, received_at, payload_sha256,
                   payload_json, conflict_policy, conflict_detected, conflict_reason
            FROM campaign_revisions
            WHERE tenant_id = ? AND campaign_id = ?
            ORDER BY server_revision DESC
            LIMIT ?
            """,
            (tenant_id, campaign_id, limit),
        ).fetchall()

    if state_row is None and not rows:
        raise HTTPException(status_code=404, detail="Unknown campaign")

    return {
        "status": "ok",
        "type": "openirn.campaignRevisions",
        "application": "OpenIRN API",
        "version": APP_VERSION,
        "storage": "mariadb",
        "tenantId": tenant_id,
        "serverTime": _utc_now().isoformat(),
        "campaignId": campaign_id,
        "current": _public_campaign_state_from_row(state_row) if state_row else None,
        "revisionCount": len(rows),
        "revisions": [_public_campaign_revision_from_row(row) for row in rows],
    }


@app.get("/campaigns/conflicts")
def campaign_conflicts(
    request: Request,
    tenantId: str = Query(default="default", min_length=1, max_length=80),
    campaignId: str = Query(default="", max_length=240),
    limit: int = Query(default=50, ge=1, le=200),
) -> dict[str, Any]:
    tenant_id = _resolve_tenant_id_for_request(tenantId, DEFAULT_TENANT_ID)
    _require_campaign_manager_authorization(request, tenant_id)
    campaign_id = str(campaignId or "").strip()

    with _db() as con:
        if campaign_id:
            rows = con.execute(
                """
                SELECT tenant_id, campaign_id, server_revision, server_sync_id,
                       device_id, updated_at, received_at, payload_sha256,
                       payload_json, conflict_policy, conflict_detected, conflict_reason
                FROM campaign_revisions
                WHERE tenant_id = ? AND campaign_id = ? AND conflict_detected = 1
                ORDER BY received_at DESC, server_revision DESC
                LIMIT ?
                """,
                (tenant_id, campaign_id, limit),
            ).fetchall()
        else:
            rows = con.execute(
                """
                SELECT tenant_id, campaign_id, server_revision, server_sync_id,
                       device_id, updated_at, received_at, payload_sha256,
                       payload_json, conflict_policy, conflict_detected, conflict_reason
                FROM campaign_revisions
                WHERE tenant_id = ? AND conflict_detected = 1
                ORDER BY received_at DESC, server_revision DESC
                LIMIT ?
                """,
                (tenant_id, limit),
            ).fetchall()

    return {
        "status": "ok",
        "type": "openirn.campaignConflicts",
        "application": "OpenIRN API",
        "version": APP_VERSION,
        "storage": "mariadb",
        "tenantId": tenant_id,
        "serverTime": _utc_now().isoformat(),
        "campaignId": campaign_id or None,
        "conflictCount": len(rows),
        "conflicts": [_public_campaign_revision_from_row(row) for row in rows],
    }


@app.get("/campaigns/revision")
def campaign_revision(
    request: Request,
    tenantId: str = Query(default="default", min_length=1, max_length=80),
    campaignId: str = Query(min_length=1, max_length=240),
    serverRevision: int = Query(ge=1),
) -> dict[str, Any]:
    tenant_id = _resolve_tenant_id_for_request(tenantId, DEFAULT_TENANT_ID)
    _require_campaign_manager_authorization(request, tenant_id)
    campaign_id = str(campaignId or "").strip()

    with _db() as con:
        row = con.execute(
            """
            SELECT tenant_id, campaign_id, server_revision, server_sync_id,
                   device_id, updated_at, received_at, payload_sha256,
                   payload_json, conflict_policy, conflict_detected, conflict_reason
            FROM campaign_revisions
            WHERE tenant_id = ? AND campaign_id = ? AND server_revision = ?
            """,
            (tenant_id, campaign_id, serverRevision),
        ).fetchone()

    if row is None:
        raise HTTPException(status_code=404, detail="Unknown campaign revision")

    return {
        "status": "ok",
        "type": "openirn.campaignRevision",
        "application": "OpenIRN API",
        "version": APP_VERSION,
        "storage": "mariadb",
        "tenantId": tenant_id,
        "serverTime": _utc_now().isoformat(),
        "revision": _public_campaign_revision_from_row(row, include_payload=True),
    }


@app.post("/campaigns/restore")
async def campaign_restore(request: Request) -> dict[str, Any]:
    try:
        payload = await request.json()
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail="Invalid JSON payload") from exc

    if not isinstance(payload, dict):
        raise HTTPException(status_code=400, detail="Payload must be a JSON object")

    tenant_id = _resolve_tenant_id_for_request(payload.get("tenantId"), DEFAULT_TENANT_ID)
    auth_context = _require_campaign_manager_authorization(request, tenant_id)
    campaign_id = str(payload.get("campaignId") or "").strip()
    restored_by_user_id = str(payload.get("restoredByUserId") or "").strip()
    reason = str(payload.get("reason") or "admin_restore").strip()[:240] or "admin_restore"

    try:
        source_revision = int(payload.get("serverRevision"))
    except (TypeError, ValueError) as exc:
        raise HTTPException(status_code=400, detail="serverRevision must be an integer") from exc

    if not campaign_id:
        raise HTTPException(status_code=400, detail="Missing campaignId")
    if source_revision < 1:
        raise HTTPException(status_code=400, detail="serverRevision must be greater than zero")

    protective_backup = _create_protective_backup(
        tenant_id,
        reason="pre_campaign_restore",
        triggered_by_user_id=restored_by_user_id or str(auth_context.get("userId") or "").strip(),
        context={"campaignId": campaign_id, "sourceRevision": source_revision, "reason": reason},
    )

    restored_at = _utc_now().isoformat()
    server_sync_id = f"restore_{_utc_now().strftime('%Y%m%dT%H%M%SZ')}_{uuid.uuid4().hex[:12]}"
    device_id = _safe_segment(restored_by_user_id or "server-restore", "server-restore")

    with _db() as con:
        with con:
            _ensure_tenant(con, tenant_id)
            source_row = con.execute(
                """
                SELECT tenant_id, campaign_id, server_revision, server_sync_id,
                       device_id, updated_at, received_at, payload_sha256,
                       payload_json, conflict_policy, conflict_detected, conflict_reason
                FROM campaign_revisions
                WHERE tenant_id = ? AND campaign_id = ? AND server_revision = ?
                """,
                (tenant_id, campaign_id, source_revision),
            ).fetchone()
            if source_row is None:
                raise HTTPException(status_code=404, detail="Unknown campaign revision")

            state_row = con.execute(
                """
                SELECT server_revision, payload_sha256, device_id
                FROM campaign_states
                WHERE tenant_id = ? AND campaign_id = ?
                """,
                (tenant_id, campaign_id),
            ).fetchone()
            if state_row is None:
                raise HTTPException(status_code=404, detail="Unknown campaign state")

            current_revision = int(state_row["server_revision"] or 0)
            source_payload = _parse_json(source_row["payload_json"], {})
            if not isinstance(source_payload, dict):
                raise HTTPException(status_code=500, detail="Stored revision payload is invalid")

            source_sha256 = str(source_row["payload_sha256"] or _json_sha256(source_payload))
            if current_revision == source_revision and str(state_row["payload_sha256"] or "") == source_sha256:
                return {
                    "status": "no_change",
                    "type": "openirn.campaignRestore",
                    "application": "OpenIRN API",
                    "version": APP_VERSION,
                    "storage": "mariadb",
                    "tenantId": tenant_id,
                    "serverTime": _utc_now().isoformat(),
                    "campaignId": campaign_id,
                    "sourceRevision": source_revision,
                    "currentRevision": current_revision,
                    "message": "The requested revision is already current",
                }

            new_revision = current_revision + 1
            restore_metadata = {
                "campaignId": campaign_id,
                "sourceRevision": source_revision,
                "newRevision": new_revision,
                "restoredByUserId": restored_by_user_id or None,
                "reason": reason,
                "sourceServerSyncId": source_row["server_sync_id"],
            }
            restore_payload = {
                "type": "openirn.syncPush",
                "sync": {
                    "tenantId": tenant_id,
                    "deviceId": device_id,
                    "mode": "admin_restore",
                },
                "generatedAt": restored_at,
                "campaigns": [source_payload],
                "restore": restore_metadata,
            }
            restore_payload_sha256 = _json_sha256(restore_payload)
            envelope = {
                "serverSyncId": server_sync_id,
                "receivedAt": restored_at,
                "tenantId": tenant_id,
                "deviceId": device_id,
                "payloadSha256": restore_payload_sha256,
                "payload": restore_payload,
            }

            con.execute(
                """
                INSERT INTO sync_snapshots(
                    tenant_id, server_sync_id, device_id, received_at,
                    payload_sha256, campaign_count, payload_json, envelope_json
                ) VALUES (?, ?, ?, ?, ?, 1, ?, ?)
                """,
                (
                    tenant_id,
                    server_sync_id,
                    device_id,
                    restored_at,
                    restore_payload_sha256,
                    _canonical_json(restore_payload),
                    _canonical_json(envelope),
                ),
            )

            con.execute(
                """
                INSERT INTO campaign_revisions(
                    tenant_id, campaign_id, server_revision, server_sync_id,
                    device_id, updated_at, received_at, payload_sha256,
                    payload_json, conflict_policy, conflict_detected, conflict_reason
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'admin_restore', 0, ?)
                """,
                (
                    tenant_id,
                    campaign_id,
                    new_revision,
                    server_sync_id,
                    device_id,
                    restored_at,
                    restored_at,
                    source_sha256,
                    _canonical_json(source_payload),
                    f"restored_from_revision_{source_revision}",
                ),
            )

            con.execute(
                """
                INSERT INTO campaign_states(
                    tenant_id, campaign_id, server_revision, server_sync_id,
                    device_id, updated_at, received_at, payload_sha256,
                    payload_json, conflict_policy
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'admin_restore')
                ON CONFLICT(tenant_id, campaign_id) DO UPDATE SET
                    server_revision = excluded.server_revision,
                    server_sync_id = excluded.server_sync_id,
                    device_id = excluded.device_id,
                    updated_at = excluded.updated_at,
                    received_at = excluded.received_at,
                    payload_sha256 = excluded.payload_sha256,
                    payload_json = excluded.payload_json,
                    conflict_policy = excluded.conflict_policy
                """,
                (
                    tenant_id,
                    campaign_id,
                    new_revision,
                    server_sync_id,
                    device_id,
                    restored_at,
                    restored_at,
                    source_sha256,
                    _canonical_json(source_payload),
                ),
            )

            con.execute(
                """
                INSERT INTO sync_events(
                    tenant_id, event_type, server_sync_id, campaign_id,
                    device_id, created_at, payload_json
                ) VALUES (?, 'campaign_revision_restored', ?, ?, ?, ?, ?)
                """,
                (
                    tenant_id,
                    server_sync_id,
                    campaign_id,
                    device_id,
                    restored_at,
                    _canonical_json(restore_metadata),
                ),
            )

    return {
        "status": "accepted",
        "type": "openirn.campaignRestore",
        "application": "OpenIRN API",
        "version": APP_VERSION,
        "storage": "mariadb",
        "tenantId": tenant_id,
        "serverTime": _utc_now().isoformat(),
        "campaignId": campaign_id,
        "sourceRevision": source_revision,
        "serverRevision": new_revision,
        "serverSyncId": server_sync_id,
        "deviceId": device_id,
        "payloadSha256": source_sha256,
        "conflictPolicy": "admin_restore",
        "protectiveBackup": protective_backup,
    }


@app.get("/maintenance/status")
def maintenance_status(
    request: Request,
    tenantId: str = Query(default="default", min_length=1, max_length=80),
    limit: int = Query(default=10, ge=1, le=50),
) -> dict[str, Any]:
    tenant_id = _resolve_tenant_id_for_request(tenantId or request.headers.get("x-openirn-tenant-id"), DEFAULT_TENANT_ID)
    _require_admin_authorization(request, tenant_id)
    return _maintenance_status(limit=limit, tenant_id=tenant_id)


@app.post("/maintenance/backup")
async def maintenance_backup(request: Request) -> dict[str, Any]:
    try:
        payload = await request.json()
    except json.JSONDecodeError:
        payload = {}
    if not isinstance(payload, dict):
        payload = {}
    tenant_id = _resolve_tenant_id_for_request(payload.get("tenantId") or request.headers.get("x-openirn-tenant-id"), DEFAULT_TENANT_ID)
    auth_context = _require_admin_authorization(request, tenant_id)
    triggered_by_user_id = (
        str(payload.get("triggeredByUserId") or "").strip()
        or str(auth_context.get("userId") or "").strip()
        or None
    )
    backup = _create_database_backup(
        tenant_id=tenant_id,
        triggered_by_user_id=triggered_by_user_id,
        reason="manual",
        automatic=False,
    )
    return {
        "status": "ok",
        "type": "openirn.mariadbDumpBackupCreated",
        "application": "OpenIRN API",
        "version": APP_VERSION,
        "serverTime": _utc_now().isoformat(),
        "backup": backup,
        "maintenance": _maintenance_status(limit=10, tenant_id=tenant_id),
    }


@app.delete("/maintenance/backups/{backup_name}")
def maintenance_delete_backup(
    backup_name: str,
    request: Request,
    tenantId: str = Query(default="default", min_length=1, max_length=80),
) -> dict[str, Any]:
    tenant_id = _resolve_tenant_id_for_request(tenantId or request.headers.get("x-openirn-tenant-id"), DEFAULT_TENANT_ID)
    auth_context = _require_admin_authorization(request, tenant_id)
    deletion = _delete_database_backup(
        backup_name,
        tenant_id=tenant_id,
        triggered_by_user_id=str(auth_context.get("userId") or "").strip() or None,
    )
    return {
        "status": "ok",
        "type": "openirn.databaseBackupDeleted",
        "application": "OpenIRN API",
        "version": APP_VERSION,
        "serverTime": _utc_now().isoformat(),
        "deletion": deletion,
        "maintenance": _maintenance_status(limit=10, tenant_id=tenant_id),
    }


@app.get("/sync/events")
async def sync_events(
    request: Request,
    tenantId: str = Query(default="default", min_length=1, max_length=80),
    since: str = Query(default="", max_length=120),
    interval: float = Query(default=2.0, ge=1.0, le=30.0),
) -> StreamingResponse:
    tenant_id = _resolve_tenant_id_for_request(tenantId, DEFAULT_TENANT_ID)
    _require_sync_read_access(request, tenant_id)
    last_server_sync_id = str(since or "").strip()

    async def event_stream():
        nonlocal last_server_sync_id
        while True:
            if await request.is_disconnected():
                break

            with _db() as con:
                latest_row = con.execute(
                    """
                    SELECT tenant_id, server_sync_id, device_id, received_at,
                           payload_sha256, campaign_count, payload_json
                    FROM sync_snapshots
                    WHERE tenant_id = ?
                    ORDER BY received_at DESC, server_sync_id DESC
                    LIMIT 1
                    """,
                    (tenant_id,),
                ).fetchone()

            latest_snapshot = _snapshot_summary_from_row(latest_row)
            current_server_sync_id = str((latest_snapshot or {}).get("serverSyncId") or "").strip()
            server_time = _utc_now().isoformat()

            if latest_snapshot and current_server_sync_id != last_server_sync_id:
                last_server_sync_id = current_server_sync_id
                payload = {
                    "type": "openirn.syncEvent",
                    "event": "snapshot",
                    "application": "OpenIRN API",
                    "version": APP_VERSION,
                    "storage": "mariadb",
                    "tenantId": tenant_id,
                    "serverTime": server_time,
                    "latestSnapshot": latest_snapshot,
                }
                yield f"event: snapshot\ndata: {json.dumps(payload, ensure_ascii=False, separators=(',', ':'))}\n\n"
            else:
                payload = {
                    "type": "openirn.heartbeat",
                    "event": "heartbeat",
                    "application": "OpenIRN API",
                    "version": APP_VERSION,
                    "storage": "mariadb",
                    "tenantId": tenant_id,
                    "serverTime": server_time,
                    "latestServerSyncId": current_server_sync_id or None,
                }
                yield f"event: heartbeat\ndata: {json.dumps(payload, ensure_ascii=False, separators=(',', ':'))}\n\n"

            await asyncio.sleep(interval)

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )
