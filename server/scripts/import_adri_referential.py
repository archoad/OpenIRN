#!/usr/bin/env python3
"""
Importe le référentiel IRN officiel aDRI depuis le fichier Excel téléchargé depuis GitLab
et produit un JSON canonique versionné.

Usage :
  python server/scripts/import_adri_referential.py \
    --input Questionnaire_IRN_v.1.1.xlsx \
    --output canonical_irn_v1_1.json \
    --version v1.1

Notes :
- Le référentiel officiel reste la source de vérité.
- Le script ne modifie pas le référentiel ; il extrait une représentation JSON exploitable par l'application.
- Les identifiants RES-3-3, RES-3-4, RES-3-5 observés dans v1.1 sont normalisés en RES-3.3, RES-3.4, RES-3.5,
  tout en conservant l'identifiant source dans `sourceCode`.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit("Installez openpyxl : pip install openpyxl") from exc

PILLAR_RE = re.compile(r"^RES-[1-8]$")
CRITERION_RE = re.compile(r"^(RES-[1-8])([.-])([0-9]+)$")

SOURCE_URL = "https://gitlab.com/digitalresilienceinitiative/adri-irn"
PROJECT_PATH = "digitalresilienceinitiative/adri-irn"
DEFAULT_FILE_PATH = "Grille d'évaluation IRN (FR)/xlsx/Questionnaire_IRN_v.1.1.xlsx"
LICENSE = "CC BY-NC-ND 4.0"

# Libellés officiels visibles dans la grille et cohérents avec le README.
DEFAULT_PILLAR_LABELS = {
    "RES-1": "Résilience stratégique",
    "RES-2": "Résilience économique et juridique",
    "RES-3": "Résilience Data & IA",
    "RES-4": "Résilience opérationnelle",
    "RES-5": "Résilience Supply-Chain",
    "RES-6": "Résilience Technologique",
    "RES-7": "Sécurité & Résilience",
    "RES-8": "Résilience Environnementale et énergétique",
}

EXPECTED_HEADERS = {
    "Dimension": "pillarId",
    "ID": "sourceCode",
    "Intitulé du critère": "label",
    "Critère": "shortLabel",
    "Description (objectif)": "description",
    "Portée du critère": "sourceScope",
    "Références réglementaires (TBD)": "regulatoryReferences",
    "Recommandations": "recommendations",
}


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def norm(value: Any) -> str:
    return "" if value is None else str(value).strip()


def normalize_code(value: Any) -> str:
    """Normalise RES-x.y et RES-x-y vers RES-x.y."""
    s = norm(value)
    m = CRITERION_RE.match(s)
    if not m:
        return s
    return f"{m.group(1)}.{m.group(3)}"


def scope_from_label(value: Any) -> str:
    s = norm(value).lower()
    if "actif" in s:
        return "asset"
    if "système" in s or "systeme" in s:
        return "criticalSystem"
    if "fonction" in s:
        # Le libellé officiel est "Fonction ou organisation".
        # On le rattache à organization pour rester compatible avec le modèle MVP.
        return "organization"
    if "organisation" in s:
        return "organization"
    return "unknown"


def clean_multiline(value: Any) -> str:
    return re.sub(r"\s+", " ", norm(value)).strip()


def sort_key_code(code: str) -> tuple[int, int]:
    m = CRITERION_RE.match(code)
    if not m:
        pm = re.match(r"^RES-([1-8])$", code)
        if pm:
            return int(pm.group(1)), 0
        return 99, 99
    return int(m.group(1).split("-")[1]), int(m.group(3))


def find_header_row(ws) -> tuple[int, dict[int, str]]:
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        values = [norm(v) for v in row]
        if "Dimension" in values and "ID" in values and "Intitulé du critère" in values:
            return row_idx, {idx: value for idx, value in enumerate(values) if value}
    raise ValueError(f"Impossible de trouver la ligne d'en-tête dans l'onglet {ws.title!r}")


def parse_pillar_labels_from_grid(wb) -> dict[str, str]:
    labels = dict(DEFAULT_PILLAR_LABELS)
    if "Grille V1" not in wb.sheetnames:
        return labels
    ws = wb["Grille V1"]
    for row in ws.iter_rows(values_only=True):
        first_cell = norm(row[0] if row else None)
        if not first_cell or "RES-" not in first_cell:
            continue
        m = re.search(r"(RES-[1-8])", first_cell)
        if not m:
            continue
        code = m.group(1)
        label = re.sub(r"\(?\s*RES-[1-8]\s*\)?", "", first_cell)
        label = clean_multiline(label)
        if label:
            labels[code] = label
    return labels


def parse_referential_sheet(wb) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[str]]:
    if "Référentiel v1" not in wb.sheetnames:
        raise ValueError("Onglet 'Référentiel v1' introuvable dans le fichier officiel")

    warnings: list[str] = []
    pillar_labels = parse_pillar_labels_from_grid(wb)
    ws = wb["Référentiel v1"]
    header_row, header_by_index = find_header_row(ws)

    criteria_by_code: dict[str, dict[str, Any]] = {}
    seen_source_codes: set[str] = set()

    for row_number, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        raw: dict[str, str] = {}
        for idx, header in header_by_index.items():
            if header in EXPECTED_HEADERS:
                raw[EXPECTED_HEADERS[header]] = norm(row[idx] if idx < len(row) else None)

        source_code = raw.get("sourceCode", "")
        if not CRITERION_RE.match(source_code):
            continue

        code = normalize_code(source_code)
        pillar_id = raw.get("pillarId") or code.split(".")[0]
        if not PILLAR_RE.match(pillar_id):
            warnings.append(f"Ligne {row_number}: dimension invalide {pillar_id!r} pour {source_code}")
            pillar_id = code.split(".")[0]

        if source_code != code:
            warnings.append(f"Ligne {row_number}: identifiant normalisé {source_code!r} -> {code!r}")

        if source_code in seen_source_codes:
            warnings.append(f"Ligne {row_number}: doublon sourceCode {source_code!r}")
        seen_source_codes.add(source_code)

        label = clean_multiline(raw.get("label", ""))
        if not label:
            label = clean_multiline(raw.get("shortLabel", ""))
        if not label:
            warnings.append(f"Ligne {row_number}: critère {code} sans libellé")

        criteria_by_code[code] = {
            "id": code,
            "code": code,
            "sourceCode": source_code,
            "pillarId": pillar_id,
            "label": label,
            "shortLabel": clean_multiline(raw.get("shortLabel", "")),
            "description": clean_multiline(raw.get("description", "")),
            "scope": scope_from_label(raw.get("sourceScope", "")),
            "sourceScope": clean_multiline(raw.get("sourceScope", "")),
            "answerMode": "R_NR",
            "regulatoryReferences": clean_multiline(raw.get("regulatoryReferences", "")),
            "recommendations": clean_multiline(raw.get("recommendations", "")),
            "active": True,
            "source": {
                "sheet": ws.title,
                "row": row_number,
            },
        }

    pillars = []
    for code in sorted(pillar_labels.keys(), key=sort_key_code):
        if PILLAR_RE.match(code):
            pillars.append({
                "id": code,
                "code": code,
                "label": pillar_labels.get(code, code),
            })

    criteria = sorted(criteria_by_code.values(), key=lambda c: sort_key_code(c["code"]))
    if len(pillars) != 8:
        warnings.append(f"Nombre de piliers inattendu: {len(pillars)} au lieu de 8")
    if not criteria:
        warnings.append("Aucun critère extrait: vérifier la structure du fichier officiel et le mapping des colonnes")

    return pillars, criteria, warnings


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, help="Chemin vers le fichier Questionnaire_IRN_v.x.xlsx")
    parser.add_argument("--output", required=True, help="Chemin du JSON canonique produit")
    parser.add_argument("--version", default="v1.1")
    parser.add_argument("--commit-sha", default=None)
    parser.add_argument("--file-path", default=DEFAULT_FILE_PATH)
    args = parser.parse_args()

    input_path = Path(args.input)
    wb = load_workbook(input_path, data_only=True, read_only=True)
    pillars, criteria, warnings = parse_referential_sheet(wb)

    result = {
        "id": f"adri-irn-{args.version}",
        "version": args.version,
        "importedAt": datetime.now(timezone.utc).isoformat(),
        "source": {
            "type": "gitlab",
            "url": SOURCE_URL,
            "projectPath": PROJECT_PATH,
            "defaultBranch": "main",
            "filePath": args.file_path,
            "commitSha": args.commit_sha,
            "checksumSha256": sha256_file(input_path),
            "license": LICENSE,
        },
        "pillars": pillars,
        "criteria": criteria,
        "importWarnings": [
            "Ne pas modifier le contenu du référentiel officiel dans l'application.",
            "Le JSON canonique sert à l'indexation et à la synchronisation interne.",
            *warnings,
        ],
    }

    Path(args.output).write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {args.output} with {len(result['pillars'])} pillars and {len(result['criteria'])} criteria")
    if warnings:
        print("Warnings:")
        for warning in warnings:
            print(f"- {warning}")


if __name__ == "__main__":
    main()
