#!/usr/bin/env python3
"""
Convertit l'Excel interne `Evaluation IRN.xlsx` en JSON de cartographie, assets,
assignations et évaluations.

Usage :
  python server/scripts/import_company_excel.py \
    --input "Evaluation IRN.xlsx" \
    --output company_seed.json \
    --campaign-id campaign-initial-import \
    --referential-id adri-irn-v1.1
"""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit("Installez openpyxl : pip install openpyxl") from exc


def value(v: Any) -> str:
    return "" if v is None else str(v).strip()


def slug_key(v: Any) -> str:
    """Clé de rapprochement robuste pour les noms d'assets."""
    s = unicodedata.normalize("NFKD", value(v)).encode("ascii", "ignore").decode("ascii")
    s = s.lower()
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def normalize_entity_id(v: Any) -> str:
    """Corrige les confusions fréquentes O/0 dans les références entité, ex. EO6 -> E06."""
    s = value(v).upper()
    m = re.match(r"^E[O0]([1-9])$", s)
    if m:
        return f"E0{m.group(1)}"
    return s


def read_sheet_rows(wb, sheet_name: str) -> list[dict[str, str]]:
    ws = wb[sheet_name]
    headers = [value(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        item = {headers[i]: value(r[i]) for i in range(min(len(headers), len(r))) if headers[i]}
        if any(item.values()):
            rows.append(item)
    return rows


def unique_by_id(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    seen = {}
    for item in items:
        if item.get("id"):
            seen[item["id"]] = item
    return list(seen.values())


def import_cartography(wb) -> dict[str, list[dict[str, Any]]]:
    rows = read_sheet_rows(wb, "1-Desc Stack")
    entities = []
    business_functions = []
    critical_systems = []
    technical_functions = []
    assets = []

    for r in rows:
        entity_id = normalize_entity_id(r.get("Ref. ent", ""))
        entity_name = r.get("Entité (Ent)", "")
        bf_id = r.get("Ref. FM", "")
        bf_name = r.get("Fonction Métier (FM)", "")
        system_id = r.get("Ref. Sys", "")
        system_name = r.get("Système (Sys)", "")
        tf_id = r.get("Ref. FT", "")
        tf_name = r.get("Fonction Technique (FT)", "")
        asset_id = r.get("Ref. As", "")
        asset_name = r.get("Asset (As)", "")

        if entity_id:
            entities.append({"id": entity_id, "name": entity_name})
        if bf_id:
            business_functions.append({"id": bf_id, "name": bf_name, "entityId": entity_id})
        if system_id:
            critical_systems.append({"id": system_id, "name": system_name, "businessFunctionId": bf_id})
        if tf_id:
            technical_functions.append({"id": tf_id, "name": tf_name, "criticalSystemId": system_id})
        if asset_id:
            assets.append({"id": asset_id, "name": asset_name, "technicalFunctionId": tf_id})

    return {
        "entities": unique_by_id(entities),
        "businessFunctions": unique_by_id(business_functions),
        "criticalSystems": unique_by_id(critical_systems),
        "technicalFunctions": unique_by_id(technical_functions),
        "assets": unique_by_id(assets),
    }


def import_harmonized_assets(wb) -> list[dict[str, Any]]:
    rows = read_sheet_rows(wb, "2- Table Assets")
    grouped: dict[str, dict[str, Any]] = {}
    for r in rows:
        source_asset_id = r.get("Ref. As", "")
        harmonized_id = r.get("Ref. Ash", "") or source_asset_id
        harmonized_name = r.get("Asset Homogénéisé (ASH)", "") or r.get("Asset (As)", "")
        is_common = r.get("Asset Commun", "").lower() in {"oui", "yes", "true", "1"}
        if not harmonized_id:
            continue
        grouped.setdefault(harmonized_id, {
            "id": harmonized_id,
            "name": harmonized_name,
            "sourceAssetIds": [],
            "criticalityWeight": 1,
            "isCommon": is_common,
        })
        grouped[harmonized_id]["isCommon"] = grouped[harmonized_id].get("isCommon", False) or is_common
        if source_asset_id and source_asset_id not in grouped[harmonized_id]["sourceAssetIds"]:
            grouped[harmonized_id]["sourceAssetIds"].append(source_asset_id)
    return list(grouped.values())


def build_harmonized_lookup(harmonized_assets: list[dict[str, Any]]) -> dict[str, str]:
    lookup: dict[str, str] = {}
    for asset in harmonized_assets:
        asset_id = value(asset.get("id"))
        asset_name = value(asset.get("name"))
        if asset_id and asset_name:
            lookup.setdefault(slug_key(asset_name), asset_id)
    return lookup


def import_assignments(wb, campaign_id: str, harmonized_assets: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[str]]:
    ws = wb["3- Assignation Evaluation"]
    headers = [value(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assignments = []
    warnings: list[str] = []
    asset_name_to_id = build_harmonized_lookup(harmonized_assets)

    for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        values = [value(v) for v in row]
        raw_target_id = values[0] if len(values) > 0 else ""
        target_name = values[1] if len(values) > 1 else ""
        evaluator_entity = normalize_entity_id(values[2] if len(values) > 2 else "")
        if not raw_target_id and not target_name:
            continue

        target_id = raw_target_id
        if not target_id and target_name:
            resolved = asset_name_to_id.get(slug_key(target_name), "")
            if resolved:
                target_id = resolved
                warnings.append(
                    f"Ligne {row_index}: Ref. Ash vide pour {target_name!r}; résolu automatiquement en {resolved!r} depuis la table des assets."
                )
            else:
                target_id = target_name
                warnings.append(
                    f"Ligne {row_index}: Ref. Ash vide pour {target_name!r}; impossible de résoudre l'identifiant harmonisé."
                )

        for col_index, criterion_id in enumerate(headers):
            if criterion_id.startswith("RES-"):
                assignee = values[col_index] if col_index < len(values) else ""
                if assignee:
                    assignments.append({
                        "id": f"ass-{campaign_id}-{row_index}-{criterion_id}",
                        "campaignId": campaign_id,
                        "targetType": "asset",
                        "targetId": target_id,
                        "targetName": target_name,
                        "criterionId": criterion_id,
                        "assigneeId": assignee,
                        "evaluatorEntityId": evaluator_entity,
                        "status": "todo",
                    })
    return assignments, warnings


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--campaign-id", default="campaign-initial-import")
    parser.add_argument("--referential-id", default="adri-irn-v1.1")
    args = parser.parse_args()

    wb = load_workbook(args.input, data_only=True)
    cartography = import_cartography(wb)
    harmonized_assets = import_harmonized_assets(wb)
    assignments, assignment_warnings = import_assignments(wb, args.campaign_id, harmonized_assets)

    result = {
        "importedAt": datetime.now(timezone.utc).isoformat(),
        "sourceFile": Path(args.input).name,
        "company": {
            **cartography,
            "harmonizedAssets": harmonized_assets,
        },
        "campaign": {
            "id": args.campaign_id,
            "name": "Campagne initiale importée depuis Excel",
            "referentialId": args.referential_id,
            "status": "draft",
            "createdAt": datetime.now(timezone.utc).isoformat(),
            "assignments": assignments,
            "evaluations": [],
        },
        "importWarnings": [
            "Les évaluations de l'onglet 4 doivent être vérifiées : la structure est une grille et non une table normalisée.",
            "Les utilisateurs doivent être rapprochés d'un référentiel d'identités avant usage multi-utilisateurs.",
            *assignment_warnings,
        ],
    }

    Path(args.output).write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {args.output}")
    print(
        f"entities={len(cartography['entities'])}, assets={len(cartography['assets'])}, "
        f"harmonizedAssets={len(harmonized_assets)}, assignments={len(assignments)}"
    )
    if assignment_warnings:
        print("Warnings:")
        for warning in assignment_warnings:
            print(f"- {warning}")


if __name__ == "__main__":
    main()
